from email.mime import image
from urllib import request

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group 
from django.contrib import messages 
from django.contrib.auth import get_user_model 
from django.db.models import Q, Sum 
from orders.models import Worker, Order     
from datetime import date, timedelta, datetime
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
import csv 
from django.urls import reverse_lazy
from django.contrib.auth.views import LoginView
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import Material
from django.db.models import Sum, F
from django.conf import settings

# AUDIT LOG UCHUN IMPORTLAR
from django.contrib.contenttypes.models import ContentType
from django.contrib.admin.models import LogEntry, CHANGE, DELETION, ADDITION 

from .models import Order, Notification, Worker 
from .forms import OrderForm, StartImageUploadForm, FinishImageUploadForm, MaterialForm

from django.db.models import Count, Case, When, IntegerField


User = get_user_model()

# --- Yordamchi Funksiya: Foydalanuvchi qaysi guruhda ekanligini tekshirish ---
def is_in_group(user, group_name):
    """Foydalanuvchi berilgan guruhda mavjudligini tekshiradi."""
    if user.is_anonymous:
        return False
        
    if user.is_superuser and group_name == 'Glavniy Admin':
        return True
        
    try:
        return user.groups.filter(name=group_name).exists()
    except Group.DoesNotExist:
        return False

# --- Yangi: Kuzatuvchi funksiyalari ---
def is_observer(user):
    """Foydalanuvchi kuzatuvchi guruhida ekanligini tekshiradi."""
    if user.is_anonymous:
        return False
    return is_in_group(user, 'Kuzatuvchi')

def is_observer_or_above(user):
    """Kuzatuvchi yoki undan yuqori darajadagi foydalanuvchilarni tekshiradi."""
    return (
        is_observer(user) or 
        is_in_group(user, 'Glavniy Admin') or 
        is_in_group(user, 'Menejer') or 
        is_in_group(user, "Ishlab Chiqarish Boshlig'i") or
        user.is_superuser
    )

# ----------------------------------------------------------------------
# 💡 YORDAMCHI FUNKSIYA: MUDDAT BUZILISHINI TEKSHIRISH
# ----------------------------------------------------------------------
def check_and_create_overdue_alerts(order):
    """
    Berilgan buyurtma uchun muddat o'tgan bo'lsa va ogohlantirish yuborilmagan bo'lsa,
    Notification yaratadi.
    """
    if not order.deadline or order.deadline > timezone.now():
        return False 
    
    if order.status in ['BAJARILDI', 'RAD_ETILDI', 'TAYYOR']:
        return False 

    if hasattr(order, 'deadline_breach_alert_sent') and order.deadline_breach_alert_sent:
        return False 

    admin_users = User.objects.filter(
        Q(is_superuser=True) | 
        Q(groups__name='Glavniy Admin') | 
        Q(groups__name="Ishlab Chiqarish Boshlig'i")
    ).distinct()
    
    message = (
        f"🚨 URGENT: Buyurtma #{order.order_number} ning muddati {order.deadline.strftime('%d-%m %H:%M')} da O'TIB KETDI. "
        f"Status: {order.get_status_display()}."
    )
    
    for admin in admin_users:
        Notification.objects.create(
            user=admin,
            order=order,
            message=message
        )
    
    if hasattr(order, 'deadline_breach_alert_sent'):
        order.deadline_breach_alert_sent = True
        order.save(update_fields=['deadline_breach_alert_sent'])
    
    return True

# --- Yordamchi Funksiya: Hisobotni ko'rishga ruxsatni tekshirish ---
def is_report_viewer(user):
    """Admin, Menejer va Ishlab Chiqarish Boshlig'iga ruxsat beradi."""
    from django.conf import settings
    
    # Agar foydalanuvchi tizimga kirmagan bo'lsa, avtomatik rad etamiz
    if not user.is_authenticated:
        return False
        
    return (
        is_in_group(user, 'Glavniy Admin') or 
        is_in_group(user, 'Menejer') or 
        is_in_group(user, "Ishlab Chiqarish Boshlig'i")
    )

# --- Yangi: Hisobotlar uchun kengaytirilgan ruxsat tekshiruvi ---
def is_report_viewer_or_observer(user):
    """Admin, Menejer, Ishlab Chiqarish Boshlig'i yoki Kuzatuvchiga ruxsat beradi."""
    return is_report_viewer(user) or is_observer(user)

# ----------------------------------------------------------------------
# YANGI: RASM YUKLASH FUNKSIYASI
# ----------------------------------------------------------------------
@require_POST
@csrf_exempt
@login_required
def upload_order_image(request):
    """
    Usta tomonidan rasm yuklash uchun AJAX endpoint
    """
    try:
        order_id = request.POST.get('order_id')
        upload_type = request.POST.get('upload_type')  # 'qabul', 'start', 'finish'
        comment = request.POST.get('comment', '')
        
        if not order_id or not upload_type:
            return JsonResponse({'success': False, 'error': 'Ma\'lumotlar yetarli emas'})
        
        try:
            order = Order.objects.get(pk=order_id)
        except Order.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Buyurtma topilmadi'})
        
        # Ruxsatni tekshirish
        is_worker = is_in_group(request.user, 'Usta')
        is_assigned_worker = order.assigned_workers.filter(user=request.user).exists()
        
        if not is_worker or not is_assigned_worker:
            return JsonResponse({'success': False, 'error': 'Sizga ruxsat yo\'q'})
        
        # Rasm yuklash turiga qarab formani tanlash
        if upload_type == 'start':
            form = StartImageUploadForm(request.POST, request.FILES, instance=order)
            if order.status == 'TASDIQLANDI' and not order.start_image:
                if form.is_valid():
                    order = form.save(commit=False)
                    order.status = 'USTA_QABUL_QILDI'
                    order.start_image_uploaded_at = timezone.now()
                    if comment:
                        order.comment = f"{order.comment or ''}\n\nUsta izohi ({timezone.now().strftime('%Y-%m-%d %H:%M')}): {comment}"
                    order.save()
                    return JsonResponse({
                        'success': True, 
                        'message': 'Boshlash rasmi muvaffaqiyatli yuklandi',
                        'new_status': order.status
                    })
                else:
                    return JsonResponse({
                        'success': False, 
                        'error': form.errors.as_text()
                    })
            else:
                return JsonResponse({'success': False, 'error': 'Boshlash rasm yuklash uchun holat mos emas'})
                
        elif upload_type == 'finish':
            form = FinishImageUploadForm(request.POST, request.FILES, instance=order)
            if order.status in ['USTA_BOSHLA', 'ISHDA'] and not order.finish_image:
                if form.is_valid():
                    order = form.save(commit=False)
                    order.status = 'USTA_TUGATDI'
                    order.worker_finished_at = timezone.now()
                    order.finish_image_uploaded_at = timezone.now()
                    if comment:
                        order.comment = f"{order.comment or ''}\n\nUsta izohi ({timezone.now().strftime('%Y-%m-%d %H:%M')}): {comment}"
                    order.save()
                    return JsonResponse({
                        'success': True, 
                        'message': 'Tugatish rasmi muvaffaqiyatli yuklandi',
                        'new_status': order.status
                    })
                else:
                    return JsonResponse({
                        'success': False, 
                        'error': form.errors.as_text()
                    })
            else:
                return JsonResponse({'success': False, 'error': 'Tugatish rasm yuklash uchun holat mos emas'})
        else:
            return JsonResponse({'success': False, 'error': 'Noto\'g\'ri yuklash turi'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ----------------------------------------------------------------------
# MAXSUS LOGIN VIEW
# ----------------------------------------------------------------------
class CustomLoginView(LoginView):
    template_name = 'orders/login.html'
    redirect_authenticated_user = True
    success_url = reverse_lazy('order_list') 

@login_required 
def order_list(request):
    user = request.user
    now = timezone.now()
    
    # ================================================================
    # 1. GURUHLAR TEKSHIRUVI (BIR MARTA)
    # ================================================================
    is_glavniy_admin = user.is_superuser or is_in_group(user, 'Glavniy Admin')
    is_production_boss = is_in_group(user, "Ishlab Chiqarish Boshlig'i")
    is_manager = is_in_group(user, 'Menejer/Tasdiqlovchi')
    is_worker = is_in_group(user, 'Usta')
    is_observer = is_in_group(user, 'Kuzatuvchi')
    is_sales_manager = is_in_group(user, 'Sales Manager') or user.username.lower() == 'sales_manager'
    
    # ================================================================
    # 2. FILTR PARAMETRLARI
    # ================================================================
    search_query = request.GET.get('q', '').strip()
    filter_type = request.GET.get('filter', 'all')
    page_number = request.GET.get('page', 1)
    
    # ================================================================
    # 3. USTA UCHUN MAXSUS LOGIKA
    # ================================================================
    if is_worker and not (is_glavniy_admin or is_production_boss or is_manager or is_observer):
        # ============================================================
        # A. MAIN ORDERLAR - FAQAT KO'RISH UCHUN (BARCHA MAIN ORDERLAR)
        # ============================================================
        main_orders_qs = Order.objects.select_related('parent_order').prefetch_related(
            'assigned_workers__user'
        ).filter(
            parent_order__isnull=True  # MAIN ORDERLAR
        ).exclude(
            status__in=['BAJARILDI', 'USTA_TUGATDI', 'TAYYOR']
        ).order_by('-created_at')
        
        # Qidiruv
        if search_query:
            search_filter = (
                Q(order_number__icontains=search_query) |
                Q(customer_name__icontains=search_query) |
                Q(product_name__icontains=search_query) |
                Q(customer_unique_id__icontains=search_query)
            )
            main_orders_qs = main_orders_qs.filter(search_filter)
        
        # Filtr turlari
        if filter_type == 'completed':
            main_orders_qs = main_orders_qs.filter(status__in=['TAYYOR', 'BAJARILDI'])
        elif filter_type == 'in_progress':
            main_orders_qs = main_orders_qs.filter(
                ~Q(status__in=['TAYYOR', 'BAJARILDI', 'RAD_ETILDI']) & 
                (Q(deadline__isnull=True) | Q(deadline__gte=now))
            )
        elif filter_type == 'overdue':
            main_orders_qs = main_orders_qs.filter(
                Q(deadline__lt=now) & 
                ~Q(status__in=['BAJARILDI', 'RAD_ETILDI', 'TAYYOR'])
            )
        
        # Pagination MAIN orderlar uchun
        from django.core.paginator import Paginator
        main_paginator = Paginator(main_orders_qs, 50)
        main_page_obj = main_paginator.get_page(page_number)
        
        # ============================================================
        # B. CHILD ORDERLAR - USTA O'ZIGA BIRIKTIRILGANLARI BILAN ISHLAYDI
        # ============================================================
        child_orders_qs = Order.objects.select_related('parent_order').prefetch_related(
            'assigned_workers__user'
        ).filter(
            parent_order__isnull=False,  # CHILD ORDERLAR
            assigned_workers__user=user  # FAQAT O'ZIGA BIRIKTIRILGANLAR
        ).exclude(
            status__in=['BAJARILDI', 'USTA_TUGATDI', 'TAYYOR']
        ).order_by('-created_at')
        
        # Child orderlarni guruhlash
        panel_child_orders = []
        ugul_child_orders = []
        other_child_orders = []
        
        for order in child_orders_qs:
            product_lower = order.product_name.lower() if order.product_name else ''
            if 'panel' in product_lower or 'панель' in product_lower or 'панел' in product_lower:
                panel_child_orders.append(order)
            elif 'ugul' in product_lower or 'угол' in product_lower or 'уголь' in product_lower:
                ugul_child_orders.append(order)
            else:
                other_child_orders.append(order)
        
        # ============================================================
        # C. STATISTIKA (MAIN ORDERLAR UCHUN)
        # ============================================================
        total_orders = main_orders_qs.count()
        completed_orders = main_orders_qs.filter(status__in=['TAYYOR', 'BAJARILDI']).count()
        in_progress_orders = main_orders_qs.filter(
            ~Q(status__in=['TAYYOR', 'BAJARILDI', 'RAD_ETILDI']) & 
            (Q(deadline__isnull=True) | Q(deadline__gte=now))
        ).count()
        overdue_orders_count = main_orders_qs.filter(
            Q(deadline__lt=now) & ~Q(status__in=['BAJARILDI', 'RAD_ETILDI', 'TAYYOR'])
        ).count()
        
        # ============================================================
        # D. CHILD ORDERLAR STATISTIKASI
        # ============================================================
        panel_child_count = len(panel_child_orders)
        ugul_child_count = len(ugul_child_orders)
        other_child_count = len(other_child_orders)
        
        panel_completed = sum(1 for o in panel_child_orders if o.status in ['TAYYOR', 'BAJARILDI'])
        ugul_completed = sum(1 for o in ugul_child_orders if o.status in ['TAYYOR', 'BAJARILDI'])
        
        panel_in_progress = panel_child_count - panel_completed
        ugul_in_progress = ugul_child_count - ugul_completed
        
        panel_progress_percentage = (panel_completed / panel_child_count * 100) if panel_child_count > 0 else 0
        ugul_progress_percentage = (ugul_completed / ugul_child_count * 100) if ugul_child_count > 0 else 0
        
        # ============================================================
        # E. ARXIV BUYURTMALAR
        # ============================================================
        archived_orders_qs = Order.objects.filter(
            status__in=['BAJARILDI', 'USTA_TUGATDI', 'TAYYOR'],
            assigned_workers__user=user
        ).select_related('parent_order').distinct()
        archived_count = archived_orders_qs.count()
        
        # ============================================================
        # F. NOTIFICATIONLAR
        # ============================================================
        user_notifications = Notification.objects.filter(user=user, is_read=False)[:5]
        
        # ============================================================
        # G. CONTEXT (USTA UCHUN)
        # ============================================================
        context = {
            # MAIN orderlar (faqat ko'rish)
            'main_orders': main_page_obj,
            'page_obj': main_page_obj,
            
            # CHILD orderlar (ishlash uchun)
            'panel_child_orders': panel_child_orders,
            'ugul_child_orders': ugul_child_orders,
            'other_child_orders': other_child_orders,
            
            # Child orderlar soni
            'panel_child_count': panel_child_count,
            'ugul_child_count': ugul_child_count,
            'other_child_count': other_child_count,
            
            # Child orderlar statistikasi
            'panel_completed': panel_completed,
            'ugul_completed': ugul_completed,
            'panel_in_progress': panel_in_progress,
            'ugul_in_progress': ugul_in_progress,
            'panel_progress_percentage': round(panel_progress_percentage, 1),
            'ugul_progress_percentage': round(ugul_progress_percentage, 1),
            
            # MAIN statistika
            'total_orders': total_orders,
            'completed_orders': completed_orders,
            'in_progress_orders': in_progress_orders,
            'overdue_orders_count': overdue_orders_count,
            
            # Arxiv
            'archived_count': archived_count,
            'archived_orders': archived_orders_qs[:100],
            
            # Rollar
            'is_glavniy_admin': is_glavniy_admin,
            'is_manager': is_manager,
            'is_production_boss': is_production_boss,
            'is_worker': is_worker,
            'is_observer': is_observer,
            'is_sales_manager': is_sales_manager,
            'is_storekeeper': False,
            'can_view_orders': True,
            
            # Filtrlar
            'search_query': search_query,
            'filter_type': filter_type,
            'now': now,
            
            # Boshqa
            'notifications': user_notifications,
            'customers_count': 0,
            'unpaid_orders_count': 0,
            'total_unpaid_amount': 0,
        }
        
        return render(request, 'orders/order_list.html', context)
    
    # ================================================================
    # 4. BOSHQA ROLLAR UCHUN (ADMIN, MANAGER, OBSERVER)
    # ================================================================
    # ARXIV BUYURTMALAR
    archived_orders_qs = Order.objects.filter(
        status__in=['BAJARILDI', 'USTA_TUGATDI', 'TAYYOR']
    ).select_related('customer').prefetch_related('assigned_workers__user')
    
    if is_worker and not (is_glavniy_admin or is_production_boss or is_manager or is_observer):
        archived_orders_qs = archived_orders_qs.filter(assigned_workers__user=user).distinct()
    
    archived_count = archived_orders_qs.count()
    
    # FAOL BUYURTMALAR
    base_qs = Order.objects.select_related('parent_order').prefetch_related(
        'assigned_workers__user'
    ).exclude(status__in=['BAJARILDI', 'USTA_TUGATDI', 'TAYYOR'])
    
    # Qidiruv
    if search_query:
        search_filter = (
            Q(order_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(product_name__icontains=search_query) |
            Q(customer_unique_id__icontains=search_query)
        )
        base_qs = base_qs.filter(search_filter)
    
    # Rol bo'yicha filtr
    if is_worker and not (is_glavniy_admin or is_production_boss or is_manager or is_observer):
        base_qs = base_qs.filter(assigned_workers__user=user).exclude(status='RAD_ETILDI').distinct()
    
    # Order tipini aniqlash
    from django.db.models import Case, When, Value, CharField
    
    orders_with_type = base_qs.annotate(
        order_type=Case(
            When(parent_order__isnull=True, then=Value('MAIN')),
            When(
                Q(product_name__icontains='panel') | 
                Q(product_name__icontains='панель') |
                Q(product_name__icontains='панел'), 
                then=Value('PANEL_CHILD')
            ),
            When(
                Q(product_name__icontains='ugul') | 
                Q(product_name__icontains='угол') |
                Q(product_name__icontains='уголь'), 
                then=Value('UGUL_CHILD')
            ),
            default=Value('OTHER_CHILD'),
            output_field=CharField()
        )
    ).order_by('-created_at')
    
    # Filtr turlari
    filter_conditions = {
        'completed': Q(status__in=['TAYYOR', 'BAJARILDI']),
        'in_progress': ~Q(status__in=['TAYYOR', 'BAJARILDI', 'RAD_ETILDI']) & (Q(deadline__isnull=True) | Q(deadline__gte=now)),
        'overdue': Q(deadline__lt=now) & ~Q(status__in=['BAJARILDI', 'RAD_ETILDI', 'TAYYOR']),
    }
    
    if filter_type in filter_conditions:
        orders_with_type = orders_with_type.filter(filter_conditions[filter_type])
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(orders_with_type, 50)
    page_obj = paginator.get_page(page_number)
    
    # Guruhlash
    main_orders = []
    panel_child_orders = []
    ugul_child_orders = []
    other_child_orders = []
    
    for order in page_obj:
        if order.order_type == 'MAIN':
            main_orders.append(order)
        elif order.order_type == 'PANEL_CHILD':
            panel_child_orders.append(order)
        elif order.order_type == 'UGUL_CHILD':
            ugul_child_orders.append(order)
        else:
            other_child_orders.append(order)
    
    # Statistika
    from django.db.models import Count, Sum, F
    
    main_qs = Order.objects.filter(parent_order__isnull=True)
    if search_query:
        main_qs = main_qs.filter(search_filter)
    
    if is_worker and not (is_glavniy_admin or is_production_boss or is_manager or is_observer):
        main_qs = main_qs.filter(assigned_workers__user=user).exclude(status='RAD_ETILDI')
    
    stats = main_qs.aggregate(
        total_orders=Count('id'),
        completed_orders=Count('id', filter=Q(status__in=['TAYYOR', 'BAJARILDI'])),
        in_progress_orders=Count('id', filter=~Q(status__in=['TAYYOR', 'BAJARILDI', 'RAD_ETILDI']) & (Q(deadline__isnull=True) | Q(deadline__gte=now))),
        overdue_orders_count=Count('id', filter=Q(deadline__lt=now) & ~Q(status__in=['BAJARILDI', 'RAD_ETILDI', 'TAYYOR'])),
    )
    
    # Child statistika
    child_stats = Order.objects.filter(parent_order__isnull=False).aggregate(
        all_child_orders_count=Count('id'),
        panel_child_count=Count('id', filter=Q(product_name__icontains='panel') | Q(product_name__icontains='панель') | Q(product_name__icontains='панел')),
        ugul_child_count=Count('id', filter=Q(product_name__icontains='ugul') | Q(product_name__icontains='угол') | Q(product_name__icontains='уголь')),
        panel_completed=Count('id', filter=(Q(product_name__icontains='panel') | Q(product_name__icontains='панель') | Q(product_name__icontains='панел')) & Q(status__in=['TAYYOR', 'BAJARILDI'])),
        ugul_completed=Count('id', filter=(Q(product_name__icontains='ugul') | Q(product_name__icontains='угол') | Q(product_name__icontains='уголь')) & Q(status__in=['TAYYOR', 'BAJARILDI'])),
    )
    
    # To'lanmaganlar
    unpaid_orders = Order.objects.none()
    total_unpaid_amount = 0
    unpaid_orders_count = 0
    
    if is_glavniy_admin or is_manager:
        unpaid_orders = Order.objects.filter(
            parent_order__isnull=True,
            total_price__gt=F('prepayment')
        ).exclude(status='BEKOR_QILINDI').only('order_number', 'customer_name', 'total_price', 'prepayment')
        unpaid_orders_count = unpaid_orders.count()
        total_unpaid_amount = unpaid_orders.aggregate(
            total=Sum(F('total_price') - F('prepayment'))
        )['total'] or 0
    
    # Notifikatsiyalar
    user_notifications = Notification.objects.filter(user=user, is_read=False)[:5]
    
    # Muddati o'tganlarni tekshirish
    if is_glavniy_admin or is_production_boss:
        overdue_check_orders = main_orders[:20]
        for order in overdue_check_orders:
            if order.deadline and order.deadline < now and order.status not in ['BAJARILDI', 'RAD_ETILDI', 'TAYYOR']:
                check_and_create_overdue_alerts(order)
    
    # Mijozlar soni
    customers_count = Order.objects.values('customer_unique_id').distinct().count()
    
    # Progress foizlari
    panel_progress_percentage = 0
    ugul_progress_percentage = 0
    panel_in_progress = 0
    ugul_in_progress = 0
    
    if child_stats['panel_child_count'] > 0:
        panel_progress_percentage = (child_stats['panel_completed'] / child_stats['panel_child_count']) * 100
        panel_in_progress = child_stats['panel_child_count'] - child_stats['panel_completed']
    
    if child_stats['ugul_child_count'] > 0:
        ugul_progress_percentage = (child_stats['ugul_completed'] / child_stats['ugul_child_count']) * 100
        ugul_in_progress = child_stats['ugul_child_count'] - child_stats['ugul_completed']
    
    # Context (admin/manager/observer)
    context = {
        'page_obj': page_obj,
        'orders': page_obj,
        'main_orders': main_orders,
        'panel_child_orders': panel_child_orders,
        'ugul_child_orders': ugul_child_orders,
        'other_child_orders': other_child_orders,
        'archived_count': archived_count,
        'archived_orders': archived_orders_qs[:100],
        'unpaid_orders_count': unpaid_orders_count,
        'total_unpaid_amount': total_unpaid_amount,
        'is_glavniy_admin': is_glavniy_admin,
        'is_manager': is_manager,
        'is_production_boss': is_production_boss,
        'is_worker': is_worker,
        'is_observer': is_observer,
        'is_sales_manager': is_sales_manager,
        'is_storekeeper': user.username.lower() == 'omborchi' or 'store' in user.username.lower(),
        'can_view_orders': any([is_glavniy_admin, is_production_boss, is_manager, is_worker, is_observer]),
        'search_query': search_query,
        'filter_type': filter_type,
        'now': now,
        'total_orders': stats['total_orders'],
        'completed_orders': stats['completed_orders'],
        'in_progress_orders': stats['in_progress_orders'],
        'overdue_orders_count': stats['overdue_orders_count'],
        'all_child_orders_count': child_stats['all_child_orders_count'],
        'panel_child_count': child_stats['panel_child_count'],
        'ugul_child_count': child_stats['ugul_child_count'],
        'other_child_count': child_stats['all_child_orders_count'] - child_stats['panel_child_count'] - child_stats['ugul_child_count'],
        'panel_completed': child_stats['panel_completed'],
        'ugul_completed': child_stats['ugul_completed'],
        'panel_in_progress': panel_in_progress,
        'ugul_in_progress': ugul_in_progress,
        'panel_progress_percentage': round(panel_progress_percentage, 1),
        'ugul_progress_percentage': round(ugul_progress_percentage, 1),
        'customers_count': customers_count,
        'notifications': user_notifications,
    }
    
    return render(request, 'orders/order_list.html', context)

@login_required
def order_receive_warehouse(request, pk):
    """Omborchi buyurtmani omborga qabul qiladi (USTA_TUGATDI -> TAYYOR)"""
    
    if not (request.user.username.lower() == 'omborchi' or 'warehouse' in request.user.username.lower() or 'store' in request.user.username.lower()):
        messages.error(request, "Sizga bu amalni bajarish uchun ruxsat yo'q!")
        return redirect('order_list')
    
    order = get_object_or_404(Order, pk=pk)
    
    if order.status == 'USTA_TUGATDI':
        order.status = 'TAYYOR'
        order.save()
        
        from django.contrib.admin.models import LogEntry, CHANGE
        from django.contrib.contenttypes.models import ContentType
        LogEntry.objects.create(
            user_id=request.user.id,
            content_type_id=ContentType.objects.get_for_model(order).pk,
            object_id=order.pk,
            object_repr=str(order),
            action_flag=CHANGE,
            change_message=f"Omborchi tomonidan qabul qilindi: USTA_TUGATDI -> TAYYOR"
        )
        
        messages.success(request, f"✅ Buyurtma №{order.order_number} omborga qabul qilindi! (TAYYOR)")
    else:
        messages.warning(request, f"Bu buyurtma 'USTA_TUGATDI' holatida emas!")
    
    return redirect('order_list')
from django.db.models import Q
from django.core.paginator import Paginator
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Order  # ChildOrder ni bu yerdan olib tashladik

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.models import User
from datetime import datetime, timedelta, date
from django.utils import timezone
from django.db.models import Q
from django.db.models.functions import Coalesce

def is_in_group(user, group_name):
    """Foydalanuvchi berilgan guruhga tegishli ekanligini tekshiradi"""
    return user.groups.filter(name=group_name).exists()

@login_required
def order_archive(request):
    """Arxivlangan (bajarilgan) buyurtmalar ro'yxati"""
    
    is_glavniy_admin = request.user.is_superuser or is_in_group(request.user, 'Glavniy Admin')
    is_manager = is_in_group(request.user, 'Menejer/Tasdiqlovchi')
    is_worker = is_in_group(request.user, 'Usta') or is_in_group(request.user, 'Eshik Ustasi')
    
    archived_statuses = ['BAJARILDI', 'USTA_TUGATDI', 'TAYYOR']
    
    # ========== ASOSIY QUERYSET - TARTIB TUZATILDI ==========
    from django.db.models.functions import Coalesce
    from django.db.models import Q
    
    main_orders = Order.objects.filter(
        status__in=archived_statuses
    ).annotate(
        # worker_finished_at NULL bo'lsa created_at ishlatiladi
        sort_date=Coalesce('worker_finished_at', 'created_at')
    ).order_by('-sort_date')  # Eng yangisi birinchi
    
    # ==================== FILTRLAR ====================
    search_query = request.GET.get('q', '').strip()
    worker_filter = request.GET.get('worker_type', '').strip()
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str = request.GET.get('date_to', '').strip()
    custom_period = request.GET.get('custom_period', '').strip()
    
    date_from = None
    date_to = None
    
    if custom_period:
        today = timezone.now().date()
        
        if custom_period == 'today':
            date_from = today
            date_to = today
        elif custom_period == 'yesterday':
            yesterday = today - timedelta(days=1)
            date_from = yesterday
            date_to = yesterday
        elif custom_period == 'week':
            date_from = today - timedelta(days=today.weekday())
            date_to = today
        elif custom_period == 'month':
            date_from = today.replace(day=1)
            date_to = today
        elif custom_period == 'year':
            date_from = today.replace(month=1, day=1)
            date_to = today
        elif custom_period == 'last_week':
            date_from = today - timedelta(days=today.weekday() + 7)
            date_to = date_from + timedelta(days=6)
        elif custom_period == 'last_month':
            first_day_of_month = today.replace(day=1)
            last_day_of_last_month = first_day_of_month - timedelta(days=1)
            date_from = last_day_of_last_month.replace(day=1)
            date_to = last_day_of_last_month
    else:
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                date_from = None
        
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                date_to = None
    
    # --- SANA FILTRI (sort_date bo'yicha) ---
    if date_from:
        main_orders = main_orders.filter(sort_date__date__gte=date_from)
    
    if date_to:
        main_orders = main_orders.filter(sort_date__date__lte=date_to)
    
    # --- USTA TURI FILTRI ---
    if not is_worker and worker_filter:
        worker_usernames = {
            'list': 'list_usta', 'panel': 'panel_usta',
            'eshik': 'eshik_usta', 'ugol': 'ugol_usta'
        }
        if worker_filter in worker_usernames:
            try:
                worker_user = User.objects.get(username=worker_usernames[worker_filter])
                main_orders = main_orders.filter(
                    assigned_workers__user=worker_user
                ).distinct()
            except User.DoesNotExist:
                main_orders = main_orders.none()
    
    elif is_worker and not (is_glavniy_admin or is_manager):
        main_orders = main_orders.filter(
            assigned_workers__user=request.user
        ).distinct()
    
    # --- QIDIRUV ---
    if search_query:
        main_orders = main_orders.filter(
            Q(order_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(product_name__icontains=search_query) |
            Q(customer_unique_id__icontains=search_query)
        )
    
    total_count = main_orders.count()
    
    # PAGINATION
    paginator = Paginator(main_orders, 12)
    page_number = request.GET.get('page', '1')
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # STATISTIKA
    worker_stats = {'list': 0, 'panel': 0, 'eshik': 0, 'ugol': 0}
    if not is_worker:
        worker_usernames = {
            'list': 'list_usta', 'panel': 'panel_usta',
            'eshik': 'eshik_usta', 'ugol': 'ugol_usta'
        }
        for key, username in worker_usernames.items():
            try:
                worker_user = User.objects.get(username=username)
                count = Order.objects.filter(
                    status__in=archived_statuses,
                    assigned_workers__user=worker_user
                ).distinct().count()
                worker_stats[key] = count
            except User.DoesNotExist:
                worker_stats[key] = 0
    
    personal_stats = {}
    if is_worker:
        worker_orders = Order.objects.filter(
            assigned_workers__user=request.user,
            status__in=archived_statuses
        ).distinct()
        personal_stats = {
            'total': worker_orders.count(),
            'bajarildi': worker_orders.filter(status='BAJARILDI').count(),
            'usta_tugatdi': worker_orders.filter(status='USTA_TUGATDI').count(),
            'tayyor': worker_orders.filter(status='TAYYOR').count(),
        }
    
    context = {
        'main_orders': page_obj,
        'main_orders_count': total_count,
        'search_query': search_query,
        'worker_filter': worker_filter,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'selected_custom_date': custom_period,
        'is_worker': is_worker,
        'is_manager': is_manager,
        'is_glavniy_admin': is_glavniy_admin,
        'worker_stats': worker_stats,
        'personal_stats': personal_stats,
    }
    
    return render(request, 'orders/order_archive.html', context)



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import DriverTrip, TripPoint
from django.utils import timezone

@login_required
def driver_dashboard(request):
    # Faqat haydovchi yoki admin kira oladi
    is_driver = "haydovchi" in request.user.username.lower() or request.user.is_staff
    if not is_driver:
        messages.error(request, "Kirish taqiqlangan!")
        return redirect('home')

    # Faqat SHU haydovchiga tegishli oxirgi faol reys
    active_trip = DriverTrip.objects.filter(
        driver=request.user, 
        is_active=True
    ).last()

    # Haydovchining oxirgi 5 ta yopilgan reysi (Tarixi)
    trip_history = DriverTrip.objects.filter(
        driver=request.user, 
        is_active=False
    ).order_by('-start_time')[:5]

    return render(request, 'orders/driver_dashboard.html', {
        'active_trip': active_trip,
        'trip_history': trip_history
    })
@csrf_exempt
@login_required
def track_location(request):
    if request.method == "POST":
        data = json.loads(request.body)
        lat = data.get('lat')
        lng = data.get('lng')
        
        # Faol reysni qidirish yoki yangisini yaratish
        trip, created = DriverTrip.objects.get_or_create(
            driver=request.user, 
            is_active=True,
            defaults={'car_number': "MASHINA-01"} # Buni profilidan olsa ham bo'ladi
        )
        
        # Yangi nuqtani saqlash
        last_point = trip.points.last()
        is_stop = False
        
        if last_point:
            # Agar oxirgi nuqtadan beri 3 minut o'tgan bo'lsa va masofa o'zgarmagan bo'lsa
            # (Bu yerda mantiqni kengaytirish mumkin)
            pass

        TripPoint.objects.create(
            trip=trip,
            latitude=lat,
            longitude=lng,
            is_stop=is_stop
        )
        
        return JsonResponse({"status": "ok", "trip_id": trip.id})

# orders/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from .models import Order, Material, MaterialTransaction


from django.db.models import Count, Sum, F, Q
from decimal import Decimal
from django.core.paginator import Paginator

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Sum, F, Q, Value, DecimalField
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from decimal import Decimal
import json

from .models import Material, Category


@login_required
@staff_member_required
def warehouse_dashboard(request):
    """Ombordagi barcha materiallar qoldig'i - Asosiy sahifa"""
    
    # Kategoriyalarni olish
    categories = Category.objects.annotate(
        material_count=Count('material'),
        total_quantity=Coalesce(Sum('material__quantity'), Value(Decimal('0'), output_field=DecimalField()))
    ).order_by('name')
    
    # Statistik ma'lumotlar
    total_materials = Material.objects.count()
    total_quantity = Material.objects.aggregate(total=Coalesce(Sum('quantity'), Value(Decimal('0'), output_field=DecimalField())))['total']
    low_stock_count = Material.objects.filter(quantity__lte=F('min_stock_level')).count()
    
    context = {
        'categories': categories,
        'total_categories': categories.count(),
        'total_materials': total_materials,
        'total_quantity': float(total_quantity) if total_quantity else 0,
        'low_stock_count': low_stock_count,
    }
    
    return render(request, 'orders/warehouse_dashboard.html', context)


# ==================== API ENDPOINTLAR ====================

@login_required
@staff_member_required
def api_statistics(request):
    """API: Statistika ma'lumotlari (JSON)"""
    try:
        total_materials = Material.objects.count()
        total_quantity = Material.objects.aggregate(
            total=Coalesce(Sum('quantity'), Value(Decimal('0'), output_field=DecimalField()))
        )['total']
        low_stock_count = Material.objects.filter(quantity__lte=F('min_stock_level')).count()
        total_categories = Category.objects.count()
        
        return JsonResponse({
            'success': True,
            'total_materials': total_materials,
            'total_quantity': float(total_quantity) if total_quantity else 0,
            'low_stock_count': low_stock_count,
            'total_categories': total_categories,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@staff_member_required
def api_categories(request):
    """API: Kategoriyalar ro'yxati (JSON)"""
    try:
        categories = Category.objects.annotate(
            material_count=Count('material'),
            total_quantity=Coalesce(Sum('material__quantity'), Value(Decimal('0'), output_field=DecimalField()))
        ).order_by('name')
        
        total_materials = Material.objects.count()
        total_quantity = Material.objects.aggregate(
            total=Coalesce(Sum('quantity'), Value(Decimal('0'), output_field=DecimalField()))
        )['total']
        
        data = {
            'success': True,
            'total_materials': total_materials,
            'total_quantity': float(total_quantity) if total_quantity else 0,
            'categories': []
        }
        
        for cat in categories:
            data['categories'].append({
                'id': cat.id,
                'name': cat.name,
                'material_count': cat.material_count,
                'total_quantity': float(cat.total_quantity) if cat.total_quantity else 0,
                'icon': 'fa-tag'
            })
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# orders/views.py - api_materials funksiyasini yangilang

@login_required
@staff_member_required
def api_materials(request):
    """API: Materiallar ro'yxati (filtrlangan)"""
    try:
        search_query = request.GET.get('search', '').strip()
        status_filter = request.GET.get('status', '')
        category_name = request.GET.get('category', '')
        
        materials = Material.objects.select_related('category').all()
        
        # Kategoriya bo'yicha filtr
        if category_name:
            materials = materials.filter(category__name=category_name)
        
        # Qidiruv bo'yicha filtr (izohni ham qo'shish)
        if search_query:
            materials = materials.filter(
                Q(name__icontains=search_query) |
                Q(category__name__icontains=search_query) |
                Q(note__icontains=search_query)  # ✅ IZOH BO'YICHA QIDIRUV
            )
        
        # Holat bo'yicha filtr
        if status_filter == 'danger':
            materials = materials.filter(quantity__lte=F('min_stock_level'))
        elif status_filter == 'success':
            materials = materials.filter(quantity__gt=F('min_stock_level'))
        elif status_filter == 'has_note':
            materials = materials.filter(note__isnull=False).exclude(note='')  # ✅ IZOHLI MATERIALLAR
        
        materials = materials.order_by('name')
        
        data = {
            'success': True,
            'materials': []
        }
        
        for m in materials:
            is_low = m.quantity <= m.min_stock_level
            data['materials'].append({
                'id': m.id,
                'name': m.name,
                'category_name': m.category.name if m.category else 'Kategoriyasiz',
                'category_id': m.category.id if m.category else None,
                'quantity': float(m.quantity),
                'unit': m.unit,
                'min_stock_level': float(m.min_stock_level),
                'is_low': is_low,
                'qr_code_url': m.qr_code.url if m.qr_code else None,
                'note': m.note or '',  # ✅ IZOHNI QAYTARISH
            })
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@staff_member_required
@require_http_methods(["POST"])
def api_material_add(request):
    """API: Yangi material qo'shish"""
    try:
        name = request.POST.get('name', '').strip()
        category_id = request.POST.get('category_id')
        quantity = float(request.POST.get('quantity', 0))
        unit = request.POST.get('unit', 'dona')
        min_stock_level = float(request.POST.get('min_stock_level', 0))
        
        if not name:
            return JsonResponse({'success': False, 'message': 'Material nomi kiritilishi shart!'})
        
        # Kategoriya
        category = None
        if category_id:
            try:
                category = Category.objects.get(id=category_id)
            except Category.DoesNotExist:
                pass
        
        material = Material.objects.create(
            name=name,
            category=category,
            quantity=quantity,
            unit=unit,
            min_stock_level=min_stock_level
        )
        
        return JsonResponse({
            'success': True,
            'message': f'"{name}" muvaffaqiyatli qo\'shildi!',
            'material': {
                'id': material.id,
                'name': material.name,
                'category_name': material.category.name if material.category else 'Kategoriyasiz',
                'quantity': float(material.quantity),
                'unit': material.unit,
                'min_stock_level': float(material.min_stock_level),
                'is_low': material.quantity <= material.min_stock_level,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Xatolik: {str(e)}'}, status=500)


# orders/views.py - api_material_get funksiyasini yangilang

@login_required
@staff_member_required
def api_material_get(request, material_id):
    """API: Bitta material ma'lumotlarini olish"""
    try:
        material = get_object_or_404(Material, id=material_id)
        return JsonResponse({
            'success': True,
            'id': material.id,
            'name': material.name,
            'category_id': material.category.id if material.category else None,
            'quantity': float(material.quantity),
            'unit': material.unit,
            'min_stock_level': float(material.min_stock_level),
            'note': material.note or '',  # ✅ IZOHN QAYTARISH
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
@login_required
@staff_member_required
@require_http_methods(["POST"])
def api_material_edit(request, material_id):
    """API: Materialni tahrirlash"""
    try:
        material = get_object_or_404(Material, id=material_id)
        
        name = request.POST.get('name', '').strip()
        category_id = request.POST.get('category_id')
        quantity = float(request.POST.get('quantity', 0))
        unit = request.POST.get('unit', 'dona')
        min_stock_level = float(request.POST.get('min_stock_level', 0))
        
        if not name:
            return JsonResponse({'success': False, 'message': 'Material nomi kiritilishi shart!'})
        
        # Kategoriya
        category = None
        if category_id:
            try:
                category = Category.objects.get(id=category_id)
            except Category.DoesNotExist:
                pass
        
        material.name = name
        material.category = category
        material.quantity = quantity
        material.unit = unit
        material.min_stock_level = min_stock_level
        material.save()
        
        return JsonResponse({
            'success': True,
            'message': f'"{material.name}" muvaffaqiyatli tahrirlandi!',
            'material': {
                'id': material.id,
                'name': material.name,
                'category_name': material.category.name if material.category else 'Kategoriyasiz',
                'quantity': float(material.quantity),
                'unit': material.unit,
                'min_stock_level': float(material.min_stock_level),
                'is_low': material.quantity <= material.min_stock_level,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Xatolik: {str(e)}'}, status=500)


@login_required
@staff_member_required
@require_http_methods(["DELETE"])
def api_material_delete(request, material_id):
    """API: Materialni o'chirish"""
    try:
        material = get_object_or_404(Material, id=material_id)
        material_name = material.name
        material.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'"{material_name}" muvaffaqiyatli o\'chirildi!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Xatolik: {str(e)}'}, status=500)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Material, Category

@login_required
@staff_member_required
def edit_material(request, material_id):
    """Materialni tahrirlash"""
    material = get_object_or_404(Material, id=material_id)
    
    if request.method == "POST":
        name = request.POST.get('name', '').strip()
        quantity = float(request.POST.get('quantity', 0))
        unit = request.POST.get('unit', '')
        min_stock = float(request.POST.get('min_stock', 0))
        
        # Kategoriya yangilash (agar kerak bo'lsa)
        category_name = request.POST.get('category_name', '')
        if category_name:
            category_obj, _ = Category.objects.get_or_create(name=category_name.strip())
            material.category = category_obj
        
        material.name = name
        material.quantity = quantity
        material.unit = unit
        material.min_stock_level = min_stock
        material.save()
        
        messages.success(request, f'"{material.name}" muvaffaqiyatli tahrirlandi!')
        return redirect('warehouse_dashboard')
    
    return render(request, 'orders/edit_material.html', {'material': material})


@login_required
@staff_member_required
def delete_material(request, material_id):
    """Materialni o'chirish"""
    material = get_object_or_404(Material, id=material_id)
    material_name = material.name
    
    if request.method == "POST":
        material.delete()
        messages.success(request, f'"{material_name}" muvaffaqiyatli o\'chirildi!')
    
    return redirect('warehouse_dashboard')

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from .models import Material, MaterialOutput, Category
import json

# orders/views.py - material_output funksiyasini to'liq yangilang

from decimal import Decimal
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from .models import Material, MaterialOutput

@login_required
@staff_member_required
def material_output(request):
    """Materialni ombordan chiqarish"""
    
    if request.method == "POST":
        m_id = request.POST.get('material_id')
        
        if not m_id:
            messages.error(request, "Iltimos, materialni ro'yxatdan tanlang!")
            return redirect('material_output')
        
        try:
            material = get_object_or_404(Material, id=m_id)
            
            # Miqdorni olish
            quantity_str = request.POST.get('quantity', '0').strip()
            if not quantity_str:
                messages.error(request, "Iltimos, miqdorni kiriting!")
                return redirect('material_output')
            
            try:
                quantity = Decimal(quantity_str)
            except:
                messages.error(request, "Miqdor noto'g'ri formatda! Iltimos, son kiriting.")
                return redirect('material_output')
            
            # Forma ma'lumotlari
            recipient = request.POST.get('recipient', '').strip()  # ✅ Qabul qilgan shaxs
            reason = request.POST.get('reason', '').strip()
            
            # Sana va vaqt
            output_date = request.POST.get('output_date')
            output_time = request.POST.get('output_time')
            
            # Validatsiya
            if quantity <= 0:
                messages.error(request, "Chiqarish miqdori 0 dan katta bo'lishi kerak!")
                return redirect('material_output')
            
            # Ombordagi joriy qoldiqni tekshirish
            current_quantity = material.quantity
            if quantity > current_quantity:
                messages.error(
                    request, 
                    f"Omborda yetarli emas! Joriy qoldiq: {current_quantity:.3f} {material.unit}. "
                    f"So'ralgan: {quantity:.3f}"
                )
                return redirect('material_output')

            # Ombordan ayirish
            material.quantity = material.quantity - quantity
            material.save()
            
            # Tarixga yozish
            MaterialOutput.objects.create(
                material=material,
                quantity=quantity,
                recipient=recipient,  # ✅ Qabul qilgan shaxs
                reason=reason,
                user=request.user,
                output_date=output_date,
                output_time=output_time,
            )
            
            messages.success(
                request, 
                f"✅ {material.name} dan {quantity:.3f} {material.unit} muvaffaqiyatli chiqarildi! "
                f"Yangi qoldiq: {material.quantity:.3f} {material.unit}"
            )
            return redirect('warehouse_dashboard')
            
        except Material.DoesNotExist:
            messages.error(request, "Material topilmadi!")
            return redirect('material_output')
            
        except Exception as e:
            print(f"❌ Xatolik: {str(e)}")
            messages.error(request, f"Tizimda xatolik: {str(e)}")
            return redirect('material_output')
    
    # GET so'rovi: materiallarni yuborish
    materials = Material.objects.filter(quantity__gt=0).order_by('name')
    return render(request, 'orders/material_output.html', {'materials': materials})









from django.shortcuts import render
from django.contrib.auth.decorators import login_required
# views.py - output_history funksiyasini yangilang

from django.db.models import Sum
from itertools import groupby
from datetime import datetime

@login_required
@staff_member_required
def output_history(request):
    """
    Chiqarish tarixi - sanalar bo'yicha guruhlangan
    """
    outputs = MaterialOutput.objects.select_related('material', 'user').order_by('-created_at')
    
    # Sanalar bo'yicha guruhlash
    grouped = {}
    user_set = set()
    
    for output in outputs:
        date_key = output.created_at.date()
        if date_key not in grouped:
            grouped[date_key] = []
        grouped[date_key].append(output)
        if output.user:
            user_set.add(output.user.username)
    
    # Guruhlarni sorted list ga o'tkazish (eng yangi sana birinchi)
    grouped_outputs = []
    for date, items in sorted(grouped.items(), reverse=True):
        total_quantity = sum(item.quantity for item in items)
        grouped_outputs.append({
            'date': date,
            'items': items,
            'total_quantity': total_quantity,
        })
    
    context = {
        'grouped_outputs': grouped_outputs,
        'outputs': outputs,
        'user_count': len(user_set),
    }
    
    return render(request, 'orders/output_history.html', context)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from .models import MaterialOutput

def export_outputs_excel(request):
    # 1. Yangi Excel kitobi yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chiqarish Tarixi"

    # --- STILLAR ---
    # Sarlavha stili (To'q yashil fon, oq yozuv, qalin)
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Markazlashtirish va Hoshiya (Border)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 2. Sarlavha ustunlarini yozish
    columns = ['№', 'Sana', 'Vaqt', 'Material Nomi', 'Kategoriya', 'Miqdor', 'Birlik', 'Kimga / Sabab', 'Mas’ul Admin']
    ws.append(columns)

    # Sarlavha dizaynini qo'llash
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # 3. Ma'lumotlarni bazadan olish
    outputs = MaterialOutput.objects.select_related('material', 'user').all().order_by('-created_at')
    
    for index, output in enumerate(outputs, start=1):
        row = [
            index,
            output.created_at.strftime('%d.%m.%Y'),
            output.created_at.strftime('%H:%M'),
            output.material.name,
            output.material.category.name if output.material.category else "-",
            output.quantity,
            output.material.unit.upper(),
            output.reason if output.reason else "-",
            output.user.username
        ]
        ws.append(row)
        
        # Har bir satrga hoshiya va tekstni tekislashni qo'shish
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="left" if isinstance(cell.value, str) else "center")

    # 4. Ustun kengligini avtomatik sozlash (Auto-fit)
    column_widths = {
        'A': 5, 'B': 12, 'C': 10, 'D': 30, 'E': 20, 
        'F': 12, 'G': 10, 'H': 35, 'I': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 5. Faylni yuborish
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=chiqarish_tarixi.xlsx'
    
    wb.save(response)
    return response

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from .models import Material

def export_inventory_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ombor Qoldig'i"

    # --- RANG STRIFTLARI ---
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid") # To'q ko'k
    low_stock_fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid") # Och qizil
    ok_stock_font = Font(color="059669", bold=True) # Yashil yozuv
    danger_font = Font(color="DC2626", bold=True) # Qizil yozuv
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Sarlavha (Katta Header)
    ws.merge_cells('A1:G1')
    ws['A1'] = "OMBORXONA JORIY QOLDIQ HISOBOTI"
    ws['A1'].font = Font(bold=True, size=16, color="1E40AF")
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # 2. Ustun nomlari (2-qator)
    columns = ['№', 'Material Nomi', 'Kategoriya', 'Joriy Qoldiq', 'Birlik', 'Minimal Limit', 'Holat']
    ws.append(columns)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 20

    # 3. Ma'lumotlar
    materials = Material.objects.select_related('category').all().order_by('name')
    
    for index, m in enumerate(materials, start=3):
        is_low = m.quantity <= m.min_stock_level
        status_text = "KAM QOLDI" if is_low else "YETARLI"
        
        row = [
            index - 2,
            m.name,
            m.category.name if m.category else "-",
            m.quantity,
            m.unit.upper(),
            m.min_stock_level,
            status_text
        ]
        ws.append(row)
        
        # Har bir katakni formatlash
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = center_align if cell.column != 2 else Alignment(horizontal="left", vertical="center", indent=1)
            
            # Agar mahsulot kam qolsa, butun qatorni och qizil qilish
            if is_low:
                cell.fill = low_stock_fill
        
        # Holat ustunini rangli qilish
        status_cell = ws.cell(row=ws.max_row, column=7)
        status_cell.font = danger_font if is_low else ok_stock_font

    # 4. Ustun kengliklarini avtomatik va aniq sozlash
    widths = [5, 40, 20, 15, 10, 15, 15]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 5. Faylni yuborish
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Ombor_Hisoboti_{timezone.now().strftime("%d_%m_%Y")}.xlsx'
    
    wb.save(response)
    return response
@login_required
@staff_member_required
def bulk_output(request):
    """Ko'p materiallarni bir vaqtda chiqarish"""
    if request.method == "POST":
        outputs_data = json.loads(request.POST.get('outputs_data', '[]'))
        
        for item in outputs_data:
            material_id = item.get('material_id')
            quantity = float(item.get('quantity', 0))
            reason = item.get('reason', '')
            
            try:
                material = Material.objects.get(id=material_id)
                if quantity <= material.quantity:
                    material.quantity -= quantity
                    material.save()
                    
                    MaterialOutput.objects.create(
                        material=material,
                        quantity=quantity,
                        reason=reason,
                        user=request.user
                    )
            except Material.DoesNotExist:
                continue
        
        messages.success(request, "Materiallar muvaffaqiyatli chiqarildi!")
        return redirect('warehouse_dashboard')
    
    materials = Material.objects.filter(quantity__gt=0).order_by('name')
    return render(request, 'orders/bulk_output.html', {'materials': materials})
@staff_member_required
@login_required
def order_picking_list(request, order_id):
    """
    Buyurtma uchun 'Sarflash ro'yxati' (Picking List)
    Bu funksiya ombordagi materiallar yetarli yoki yo'qligini tekshiradi.
    """
    order = get_object_or_404(Order, id=order_id)
    
    # 1. Kalkulyatordan hisob-kitoblarni olamiz
    calc_data = order.calculate_materials()
    
    if not calc_data:
        return render(request, 'orders/picking_list.html', {
            'order': order,
            'error': "Ushbu buyurtma turi uchun hisob-kitob mantiqi topilmadi."
        })

    # 2. Ombor bilan solishtirish (Report tayyorlash)
    report = []
    can_produce = True
    
    # Materiallar xaritasi: Kalkulyatordagi kalit so'z -> Bazadagi qidiriladigan nom
    check_list = [
        {'key': 'foam_volume', 'search': 'siryo', 'label': 'Siryo (Foam)'},
        {'key': 'sheets_area', 'search': 'list', 'label': 'List (Metal)'},
    ]

    for item in check_list:
        needed = calc_data.get(item['key'], 0)
        # Bazadan nomiga qarab qidiramiz
        material = Material.objects.filter(name__icontains=item['search']).first()
        
        available = material.quantity if material else 0
        is_enough = available >= needed if material else False
        
        if not is_enough:
            can_produce = False
            
        report.append({
            'label': item['label'],
            'needed': needed,
            'available': available,
            'status': is_enough,
            'unit': material.unit if material else '?'
        })

    context = {
        'order': order,
        'report': report,
        'can_produce': can_produce,
        'calc_data': calc_data
    }
    return render(request, 'orders/picking_list.html', context)
from django.shortcuts import render, redirect
from .models import Material, Category
# orders/views.py - add_material funksiyasini yangilang

@login_required
def add_material(request):
    if request.method == "POST":
        name = request.POST.get('name', '').strip()
        category_name = request.POST.get('category_name', '').strip()
        quantity = float(request.POST.get('quantity', 0))
        unit = request.POST.get('unit', 'dona')
        min_stock = request.POST.get('min_stock', 0)
        note = request.POST.get('note', '').strip()  # ✅ IZOHNI QABUL QILISH

        # Kategoriya mantiqi
        category_obj = None
        if category_name:
            category_obj, created = Category.objects.get_or_create(name=category_name.strip())

        # MAHSULOTNI TEKSHIRISH VA SAQLASH
        material, created = Material.objects.get_or_create(
            name=name,
            defaults={
                'category': category_obj,
                'quantity': quantity,
                'unit': unit,
                'min_stock_level': min_stock,
                'note': note  # ✅ IZOHNI SAQLASH
            }
        )

        # Agar mahsulot allaqachon bor bo'lsa, miqdorini qo'shamiz va izohni yangilaymiz
        if not created:
            material.quantity = float(material.quantity) + quantity
            if note:  # Agar izoh yozilgan bo'lsa, yangilaymiz
                material.note = note
            material.save()

        return redirect('warehouse_dashboard')

    categories = Category.objects.all()
    return render(request, 'orders/add_material.html', {'categories': categories})

@login_required
def guard_dashboard(request):
    user_lower = request.user.username.lower()
    if not ("qorovul" in user_lower or "guard" in user_lower or request.user.is_superuser):
        return HttpResponseForbidden("Sizda qorovul paneliga kirish huquqi yo'q!")

    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        action = request.POST.get('action')
        order = get_object_or_404(Order, id=order_id)
        img = request.FILES.get('guard_img')
        now = timezone.now()

        if action == 'enter':
            status_text = "KIRDI (Yukxonaga)"
            status_emoji = "📥"
            # KIRISH VAQTINI MUHRLASH
            order.work_started_at = now 
            order.save()
        elif action == 'exit':
            status_text = "CHIQDI (Zavoddan)"
            status_emoji = "📤"
            # CHIQISH VAQTINI MUHRLASH VA STATUSNI O'ZGARTIRISH
            order.status = 'YUK_CHIQDI'
            order.work_finished_at = now
            order.save()

        # Telegramga yuborish (vaqt bilan)
        if img:
            caption = (
                f"🛡️ #QOROVUL_NAZORATI\n"
                f"{status_emoji} {status_text}\n"
                f"📦 Buyurtma: #{order.id}\n"
                f"🚛 Moshina: {order.worker_comment}\n"
                f"🕒 Vaqt: {now.strftime('%H:%M')}"
            )
            try:
                url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendPhoto"
                img.seek(0)
                files = {'photo': (img.name, img.read(), img.content_type)}
                requests.post(url, data={'chat_id': TELEGRAM_GROUP_ID, 'caption': caption}, files=files)
            except Exception as e:
                print(f"Telegram error: {e}")

        messages.success(request, f"#{order.id} {status_text} tasdiqlandi (Vaqt: {now.strftime('%H:%M')}).")
        return redirect('guard_dashboard')
    


    # 1. KUTILAYOTGANLAR (Hali chiqmagan moshinalar)
    pending_orders = Order.objects.filter(
        status='BAJARILDI'
    ).exclude(
        Q(worker_comment="") | Q(worker_comment__isnull=True)
    ).order_by('-work_finished_at')

    # 2. BUGUNGI TARIX (Kirgan va chiqqan moshinalar jadvali)
    today = timezone.now().date()
    today_history = Order.objects.filter(
        Q(work_started_at__date=today) | Q(work_finished_at__date=today)
    ).filter(
        status__in=['BAJARILDI', 'YUK_CHIQDI']
    ).order_by('-id')

    context = {
        'orders': pending_orders,
        'history': today_history,
    }
    return render(request, 'orders/guard_dashboard.html', context)










from datetime import datetime  # BU MUHIM: importni shunday o'zgartiring
import requests  
from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import GuardPatrol

# Telegram bot sozlamalari
BOT_TOKEN = '8559719741:AAGa5BnxXt2rxjC-gKFnzboBiJQgPUY2GzU' # O'zingizniki bilan almashtiring
CHAT_ID = '-1002338157363' # Guruh ID sini qo'ying
import json
import requests
from datetime import datetime
from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import GuardPatrol

# Sozlamalarni o'zgaruvchilarga chiqaramiz
BOT_TOKEN = "8559719741:AAGa5BnxXt2rxjC-gKFnzboBiJQgPUY2GzU"
CHAT_ID = "-1003274223599"

import json
import requests
from datetime import datetime
from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import GuardPatrol

# --- TELEGRAM SOZLAMALARI ---
BOT_TOKEN = "8559719741:AAGa5BnxXt2rxjC-gKFnzboBiJQgPUY2GzU"
CHAT_ID = "-1002338157363"

import json
import requests

def send_patrol_to_telegram(patrol):
    """Hisobotni va rasmlarni bitta albom qilib Telegramga yuborish"""
    map_url = f"https://www.google.com/maps?q={patrol.latitude},{patrol.longitude}"

    caption = (
        f"🚨 *YANGI PATRUL HISOBOTI*\n\n"
        f"👤 *Qorovul:* {patrol.guard.get_full_name() or patrol.guard.username}\n"
        f"⏰ *Vaqt:* {patrol.patrol_time_slot}\n"
        f"📅 *Sana:* {patrol.created_at.strftime('%d.%m.%Y %H:%M')}\n"
        f"📍 [Xaritada ko'rish]({map_url})"
    )

    media_url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMediaGroup"
    files = {}
    media = []

    # ✅ 4 ta rasm
    image_fields = [patrol.image1, patrol.image2, patrol.image3, patrol.image4]

    for i, img_field in enumerate(image_fields, 1):
        if img_field and img_field.name:
            file_key = f"pic{i}"
            try:
                # local storage bo‘lsa path bo‘ladi
                files[file_key] = open(img_field.path, "rb")
                media.append({
                    "type": "photo",
                    "media": f"attach://{file_key}",
                    "caption": caption if i == 1 else "",
                    "parse_mode": "Markdown",
                })
            except Exception as e:
                print(f"Rasm ochishda xato ({file_key}): {e}")

    if not media:
        print("Telegramga yuborish bekor: rasm topilmadi (media bo'sh).")
        return None

    try:
        response = requests.post(
            media_url,
            data={"chat_id": CHAT_ID, "media": json.dumps(media)},
            files=files,
            timeout=40
        )

        # ✅ debug: aynan nima xato ekanini ko‘rasan
        print("TELEGRAM STATUS =", response.status_code)
        print("TELEGRAM TEXT =", response.text)

        return response.json()

    except requests.exceptions.Timeout:
        print("Telegram xatosi: Timeout (rasmlar katta bo'lishi mumkin).")
    except Exception as e:
        print(f"Telegram yuborishda kutilmagan xato: {e}")
    finally:
        for f in files.values():
            try:
                f.close()
            except:
                pass

    return None

            
from datetime import datetime
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.utils import timezone

from .models import GuardPatrol  # sizda qaysi appda bo'lsa shu yo'lni to'g'rilang


@login_required
def guard_patrol_view(request):
    if not (request.user.is_staff or request.user.username == 'Qorovul'):
        messages.error(request, "Sizda bu sahifaga kirish ruxsati yo'q!")
        return redirect('home')

    now_dt = timezone.localtime()
    now_time = now_dt.time()
    today = now_dt.date()

    # ✅ Patrul vaqtlar jadvali (to'g'ri formatda)
    patrol_slots = [
        ("02:00", "02:20"),
        ("04:00", "04:20"),
        ("05:00", "05:20"),
        ("12:00", "12:20"),  # ⚠️ sizda 12:30-12:20 xato edi
        ("14:00", "14:20"),
        ("18:00", "18:20"),
        ("22:00", "22:20"),
    ]

    # Bugun topshirilgan slotlarni 1 martada olib olamiz (tez)
    completed_qs = GuardPatrol.objects.filter(
        guard=request.user,
        created_at__date=today
    ).values_list('patrol_time_slot', flat=True)
    completed_set = set(completed_qs)

    # Hozir active slotni topamiz + hamma slotlar uchun status tayyorlaymiz
    active_slot = None
    slots_ui = []

    for start, end in patrol_slots:
        start_t = datetime.strptime(start, "%H:%M").time()
        end_t = datetime.strptime(end, "%H:%M").time()

        slot_label = f"{start} - {end}"
        is_active = (start_t <= now_time <= end_t)

        if is_active and active_slot is None:
            active_slot = slot_label

        slots_ui.append({
            "label": slot_label,
            "start": start,
            "end": end,
            "is_active": is_active,
            "is_completed": slot_label in completed_set,
        })

    # POST faqat active slot bo'lsa ishlasin
    if request.method == "POST":
        slot_from_post = request.POST.get("slot")  # qaysi slotdan yuborildi

        if not active_slot or slot_from_post != active_slot:
            messages.error(request, "Hozir patrul vaqti emas yoki noto‘g‘ri vaqt tanlandi!")
            return redirect('guard_patrol')

        if active_slot in completed_set:
            messages.warning(request, "Siz bu vaqt oralig'i uchun hisobot topshirib bo'lgansiz!")
            return redirect('guard_patrol')

        img1 = request.FILES.get('img1')
        img2 = request.FILES.get('img2')
        img3 = request.FILES.get('img3')
        img4 = request.FILES.get('img4')
        lat = request.POST.get('lat')
        lng = request.POST.get('lng')

        if img1 and img2 and img3 and img4:
            patrol = GuardPatrol.objects.create(
                guard=request.user,
                checkpoint_name="Umumiy nazorat",
                patrol_time_slot=active_slot,
                image1=img1, image2=img2, image3=img3, image4=img4,   # ✅ shu qo‘shildi
                latitude=float(lat) if lat and lat != "undefined" else 0.0,
                longitude=float(lng) if lng and lng != "undefined" else 0.0
            )

            result = send_patrol_to_telegram(patrol)  # ✅ resultni ushlab qolamiz
            print("TELEGRAM RESULT =", result)         # ✅ konsolda ko‘rasan

            messages.success(request, "Patrul hisoboti muvaffaqiyatli topshirildi!")
            return redirect('guard_patrol')
        else:
            messages.error(request, "Xatolik: 4 ta rasm yuklash majburiy!")


         

    return render(request, 'orders/patrol.html', {
        "slots": slots_ui,
        "active_slot": active_slot,
        "current_time": now_dt,  # template'da vaqt ko'rsatish uchun
    })



def rankings_view(request):
    """Ustalar reytingi sahifasi"""
    # models.Q o'rniga Q o'zi ishlatildi (import qismiga qarang)
    workers_list = Worker.objects.annotate(
        total_finished=Count('orders', filter=Q(orders__status='TUGATILDI')),
        total_kvadrat=Sum('orders__panel_kvadrat', filter=Q(orders__status='TUGATILDI'))
    ).order_by('-total_kvadrat')

    context = {
        'workers': workers_list,
    }
    return render(request, 'orders/rankings.html', context)
# ----------------------------------------------------------------------
# BUYURTMA TAHSILOTLARI
# ----------------------------------------------------------------------
import requests
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse
from .models import Order
from .forms import OrderForm

# TELEGRAM BOT CONFIG
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from .models import Order
from .forms import OrderForm
import requests
from django.utils import timezone
from django.contrib.admin.models import LogEntry, CHANGE
from django.contrib.contenttypes.models import ContentType

# TELEGRAM CONFIG (Siz so'ragandek shu yerda qoldi)
TELEGRAM_BOT_TOKEN = "8559719741:AAGa5BnxXt2rxjC-gKFnzboBiJQgPUY2GzU"
TELEGRAM_GROUP_ID = "-1002338157363"

def is_in_group(user, group_name):
    return user.groups.filter(name=group_name).exists()

@login_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)

    # 1. Ruxsatlarni aniqlash (O'zgarishsiz)
    is_glavniy_admin = request.user.is_superuser or is_in_group(request.user, 'Glavniy Admin')
    is_manager = is_in_group(request.user, 'Menejer/Tasdiqlovchi')
    is_production_boss = is_in_group(request.user, "Ishlab Chiqarish Boshlig'i")
    is_worker = is_in_group(request.user, 'Usta') 
    is_observer_user = is_in_group(request.user, 'Kuzatuvchi')

    # Kuzatuvchi uchun faqat ko'rish rejimi
    if is_observer_user:
        return render(request, 'orders/order_detail.html', {'order': order, 'readonly': True})

    # 2. Usta ruxsatini tekshirish
    is_assigned_worker = False
    if is_worker:
        try:
            worker_profile = request.user.worker_profile
            is_assigned_worker = order.assigned_workers.filter(pk=worker_profile.pk).exists()
        except Exception:
            is_assigned_worker = False

    if is_worker and not is_assigned_worker and not is_production_boss and not is_glavniy_admin:
        messages.error(request, "Siz faqat o'zingizga tayinlangan buyurtmalarni ko'rishingiz mumkin.")
        return redirect('order_list')

    # 3. Formani boshqarish (Admin/Manager uchun)
    order_form = None
    if is_glavniy_admin or is_manager or is_production_boss:
        order_form = OrderForm(request.POST or None, request.FILES or None, instance=order)
        if request.method == 'POST' and 'upload_type' not in request.POST:
            if order_form.is_valid():
                order_form.save()
                messages.success(request, "Buyurtma muvaffaqiyatli yangilandi.")
                return redirect('order_detail', pk=order.pk)

    # 4. POST: Rasm yuklash va Telegram (TO'G'RILANGAN QISM)
    # Ruxsat: Agar usta biriktirilgan bo'lsa YOKI admin/boshliq bo'lsa rasm yuklay oladi
    can_upload = is_assigned_worker or is_glavniy_admin or is_production_boss

    if request.method == 'POST' and 'upload_type' in request.POST and can_upload:
        upload_type = request.POST.get('upload_type')
        image = request.FILES.get(upload_type)

        if image and upload_type in ['start_image', 'finish_image']:
            # Emojilarsiz xavfsiz matn
            status_text = "ISH BOSHLANDI" if upload_type == 'start_image' else "ISH YAKUNLANDI"
            debt = order.total_price - order.prepayment
            payment_info = "Toliq tolangan" if debt <= 0 else f"Qarz: {debt} USD"
            customer_name = order.customer.name if hasattr(order, 'customer') and order.customer else "Nomalum"

            caption = (
                f"{status_text}\n"
                f"-------------------------------\n"
                f"Buyurtma: #{order.id}\n"
                f"Mijoz: {order.customer_unique_id} / {customer_name}\n"
                f"Mahsulot: {order.product_name or 'Aniqlanmagan'}\n"
                f"Olcham: {order.panel_thickness or '0'} sm | {order.panel_kvadrat or '0'} m2\n"
                f"Jami: {order.total_price} USD\n"
                f"Holat: {payment_info}\n"
                f"-------------------------------\n"
                f"Masul: {request.user.get_full_name() or request.user.username}\n"
                f"Vaqt: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
            )

            try:
                # Telegramga yuborish
                response = requests.post(
                    f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendPhoto",
                    data={
                        "chat_id": TELEGRAM_GROUP_ID,
                        "caption": caption
                    },
                    files={"photo": image},
                    timeout=20
                )

                if response.status_code == 200:
                    # Telegramga muvaffaqiyatli ketgandagina bazani yangilaymiz
                    if upload_type == 'start_image':
                        order.start_image = image
                        order.start_confirmed = True
                        order.started_by = request.user
                        order.work_started_at = timezone.now()
                        order.status = 'USTA_QABUL_QILDI'
                    else:
                        order.finish_image = image
                        order.finish_confirmed = True
                        order.finished_by = request.user
                        order.work_finished_at = timezone.now()
                        order.status = 'USTA_TUGATDI'

                    order.save()

                    LogEntry.objects.create(
                        user_id=request.user.id,
                        content_type_id=ContentType.objects.get_for_model(order).pk,
                        object_id=order.pk,
                        object_repr=str(order),
                        action_flag=CHANGE,
                        change_message=f"Telegramga hisobot yuborildi: {status_text}"
                    )
                    messages.success(request, "Hisobot Telegramga yuborildi va saqlandi.")
                else:
                    messages.error(request, f"Telegram bot rad etdi: {response.text}")

            except Exception as e:
                messages.error(request, f"Aloqa xatoligi: {str(e)}")
            
            return redirect('order_detail', pk=order.pk)
        else:
            messages.error(request, "Rasm tanlanmagan!")

    # 5. Sahifani yuklash uchun context
    context = {
        'order': order,
        'order_form': order_form,
        'is_worker': is_worker,
        'is_assigned_worker': is_assigned_worker,
        'is_privileged': is_glavniy_admin or is_manager or is_production_boss,
    }
    return render(request, 'orders/order_detail.html', context)

# ----------------------------------------------------------------------
# USTA HARAKATLARI FUNKSIYALARI
# ----------------------------------------------------------------------
@login_required
@user_passes_test(lambda u: is_in_group(u, 'Usta') or u.is_superuser, login_url='/login/')
def order_worker_accept(request, pk):
    """Usta buyurtmani qabul qilish."""
    # Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    order = get_object_or_404(Order, pk=pk)
    
    if not request.user.is_superuser and not order.assigned_workers.filter(user=request.user).exists():
        messages.error(request, "Siz bu buyurtmaga tayinlanmagansiz.")
        return redirect('order_list')

    if order.status == 'TASDIQLANDI':
        if not order.start_image:
             messages.error(request, "Ishni qabul qilishdan oldin, **Boshlanish Rasmini** yuklashingiz kerak.")
             return redirect('order_detail', pk=order.pk)

        order.status = 'USTA_QABUL_QILDI'
        order.save(update_fields=['status'])
        messages.success(request, f"Buyurtma #{order.order_number} ustalar tomonidan qabul qilindi. Endi 'Boshlash' tugmasini bosing.")
    else:
        messages.warning(request, f"Buyurtma #{order.order_number} faqat 'Tasdiqlandi' statusida qabul qilinishi mumkin.")
        
    return redirect('order_list')

@login_required
@user_passes_test(lambda u: is_in_group(u, 'Usta') or u.is_superuser, login_url='/login/')
def order_worker_start(request, pk):
    """Usta ishni boshlash."""
    # Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    order = get_object_or_404(Order, pk=pk)
    
    if not request.user.is_superuser and not order.assigned_workers.filter(user=request.user).exists():
        messages.error(request, "Siz bu operatsiyani bajarishga ruxsat etilmagansiz.")
        return redirect('order_list')

    if order.status == 'USTA_QABUL_QILDI':
        order.status = 'USTA_BOSHLA'
        order.worker_started_at = timezone.now() 
        order.save(update_fields=['status', 'worker_started_at'])
        messages.success(request, f"Buyurtma #{order.order_number} bo'yicha ish boshlandi. Status: USTA BOSHLADI.")
    else:
        messages.warning(request, f"Ishni boshlash uchun buyurtma 'Usta Qabul Qildi' statusida bo'lishi kerak.")
        
    return redirect('order_list')


@login_required
@user_passes_test(lambda u: is_in_group(u, 'Usta') or u.is_superuser, login_url='/login/')
def order_worker_finish(request, pk):
    """Usta ishni yakunlash va avtomatik ravishda keyingi bosqich ustalari uchun buyurtma ochish."""
    
    # 1. Kuzatuvchi (Observer) tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    order = get_object_or_404(Order, pk=pk)

    # 2. Huquqlarni tekshirish (faqat biriktirilgan usta yoki superuser)
    if not request.user.is_superuser and not order.assigned_workers.filter(user=request.user).exists():
        messages.error(request, "Siz bu operatsiyani bajarishga ruxsat etilmagansiz.")
        return redirect('order_list')

    # 3. Rasm yuklanganligini tekshirish
    if not order.finish_image:
         messages.error(request, "Ishni tugatishdan oldin, yakuniy rasm (Tugatish Rasmi) yuklashingiz kerak.")
         return redirect('order_detail', pk=order.pk)
        
    # 4. Statusni yangilash
    if order.status in ['USTA_BOSHLA', 'ISHDA']: 
        current_time = timezone.now()
        order.status = 'USTA_TUGATDI'
        order.worker_finished_at = current_time 
        
        # Muddatdan o'tib ketgan bo'lsa ogohlantirish
        if order.deadline and current_time > order.deadline:
            # Agar funksiya mavjud bo'lsa chaqiriladi
            if 'check_and_create_overdue_alerts' in globals():
                check_and_create_overdue_alerts(order)
            messages.warning(request, f"⚠️ Buyurtma #{order.order_number} muddatidan kech yakunlandi.")
            
        order.save(update_fields=['status', 'worker_finished_at'])

        # ================================================================
        # YANGI ZANJIRSIMON ALGORITM (List usta -> Panel/Ugol usta)
        # ================================================================
        
        # Agar hozirgi foydalanuvchi "List usta" guruhida bo'lsa
        if is_in_group(request.user, "List usta"):
            # Keyingi bosqich ustalari (Panel va Ugol) guruhlarini bazadan topamiz
            next_workers = Worker.objects.filter(
                Q(user__groups__name="Panel usta") | Q(user__groups__name="Ugol usta")
            )
            
            if next_workers.exists():
                # Yangi order raqami (masalan: ORD-100 bo'lsa, ORD-100-PU bo'ladi)
                new_order_number = f"{order.order_number}-PU"
                
                # Agar bunaqa raqamli buyurtma hali ochilmagan bo'lsa (dublikat bo'lmasligi uchun)
                if not Order.objects.filter(order_number=new_order_number).exists():
                    new_order = Order.objects.create(
                        order_number=new_order_number,
                        material=order.material,
                        quantity=order.quantity,
                        drawings_pdf=order.drawings_pdf, # List usta ishlatgan chizmani o'tkazamiz
                        status='TASDIQLANDI',           # Avtomatik tasdiqlangan holatda
                        created_by=order.created_by,
                        deadline=timezone.now() + timedelta(days=1), # 1 kun muddat
                        notes=f"List usta #{order.order_number} ishini yakunlagani uchun avtomatik yaratildi."
                    )
                    
                    # Topilgan barcha Panel va Ugol ustalarni yangi buyurtmaga biriktiramiz
                    for worker in next_workers:
                        new_order.assigned_workers.add(worker)
                        
                        # Har biriga bildirishnoma yuboramiz
                        Notification.objects.create(
                            user=worker.user,
                            order=new_order,
                            message=f"Yangi ish: List usta #{order.order_number} chizmasini bitirdi. Panel/Ugol bosqichini boshlang."
                        )
                    
                    messages.success(request, "Panel va Ugol ustalari uchun avtomatik buyurtma yaratildi.")
        # ================================================================

        messages.success(request, f"Buyurtma #{order.order_number} yakunlandi.")
    else:
        messages.warning(request, "Ishni yakunlash uchun buyurtma 'Usta Boshladi' yoki 'Ishda' statusida bo'lishi kerak.")
        
    return redirect('order_list')

# ----------------------------------------------------------------------
# YANGI: USTALAR PANELI FUNKSIYALARI
# ----------------------------------------------------------------------
@login_required
@user_passes_test(lambda u: is_in_group(u, 'Usta') or u.is_superuser or 
                   is_in_group(u, "Ishlab Chiqarish Boshlig'i") or 
                   is_in_group(u, 'Kuzatuvchi'), login_url='/login/')  # ✅ YANGI
def worker_panel(request):
    """
    Ustalar paneli - barcha ustalar ro'yxati
    """
    # Faqat admin, ishlab chiqarish boshliqlari va kuzatuvchilar ko'ra oladi
    is_glavniy_admin = request.user.is_superuser or is_in_group(request.user, 'Glavniy Admin')
    is_production_boss = is_in_group(request.user, "Ishlab Chiqarish Boshlig'i")
    is_observer = is_in_group(request.user, 'Kuzatuvchi')  # ✅ YANGI
    
    if not (is_glavniy_admin or is_production_boss or is_observer):  # ✅ YANGI
        messages.error(request, "Sizda bu sahifani ko'rish uchun ruxsat yo'q.")
        return redirect('order_list')
    
    # Barcha ustalarni olish
    workers = Worker.objects.all().select_related('user').annotate(
        completed_orders_count=Count(
            'orders', 
            filter=Q(orders__status__in=['TAYYOR', 'BAJARILDI'])
        ),
        total_kvadrat=Sum(
            'orders__panel_kvadrat', 
            filter=Q(orders__status__in=['TAYYOR', 'BAJARILDI'])
        )
    )

    context = {
        'workers': workers,
        'is_glavniy_admin': is_glavniy_admin,
        'is_production_boss': is_production_boss,
        'is_observer': is_observer,
    }
    
    return render(request, 'orders/worker_panel.html', context)

@login_required
@user_passes_test(lambda u: is_in_group(u, 'Usta') or u.is_superuser or 
                   is_in_group(u, "Ishlab Chiqarish Boshlig'i") or 
                   is_in_group(u, 'Kuzatuvchi'), login_url='/login/')  # ✅ YANGI
def worker_orders(request, worker_id):
    """
    Muayyan ustaning barcha buyurtmalari
    """
    worker = get_object_or_404(Worker, id=worker_id)
    
    # Ruxsatni tekshirish
    is_glavniy_admin = request.user.is_superuser or is_in_group(request.user, 'Glavniy Admin')
    is_production_boss = is_in_group(request.user, "Ishlab Chiqarish Boshlig'i")
    is_worker_self = request.user == worker.user
    is_observer = is_in_group(request.user, 'Kuzatuvchi')  # ✅ YANGI
    
    if not (is_glavniy_admin or is_production_boss or is_worker_self or is_observer):  # ✅ YANGI
        messages.error(request, "Sizda bu sahifani ko'rish uchun ruxsat yo'q.")
        return redirect('order_list')
    
    # Filtrlash parametrlari
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status_filter = request.GET.get('status', '')
    
    # Ustaning buyurtmalari
    # orders = Order.objects.filter(assigned_workers=worker).order_by('-created_at')
    # select_related - bog'langan model ma'lumotlarini bitta so'rovda oladi
# prefetch_related - ManyToMany (ustalar) bog'liqligini tezlashtiradi
    orders = Order.objects.filter(assigned_workers=worker)\
        .select_related('parent_order')\
        .prefetch_related('assigned_workers')\
        .order_by('-created_at')
    
    # Filtrlash
    if start_date:
        orders = orders.filter(created_at__gte=start_date)
    if end_date:
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        orders = orders.filter(created_at__lt=end_datetime)
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    # Statistikani hisoblash
    total_orders = orders.count()
    completed_orders = orders.filter(status__in=['TAYYOR', 'BAJARILDI']).count()
    total_kvadrat = orders.filter(status__in=['TAYYOR', 'BAJARILDI']).aggregate(
        Sum('panel_kvadrat')
    )['panel_kvadrat__sum'] or 0
    
    context = {
        'worker': worker,
        'orders': orders,
        'total_orders': total_orders,
        'completed_orders': completed_orders,
        'total_kvadrat': total_kvadrat,
        'start_date': start_date,
        'end_date': end_date,
        'status_filter': status_filter,
        'is_glavniy_admin': is_glavniy_admin,
        'is_production_boss': is_production_boss,
        'is_worker_self': is_worker_self,
        'is_observer': is_observer,  # ✅ YANGI
    }
    
    return render(request, 'orders/worker_orders.html', context)

# ----------------------------------------------------------------------
# QOLGAN FUNKSIYALAR
# views.py - order_create funksiyasini yangilang
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from .models import Order, Customer, User # Kerakli modellarni import qiling
from .forms import OrderForm

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Order, Customer, User
from .forms import OrderForm
@login_required
@user_passes_test(
    lambda u: u.is_superuser or is_in_group(u, 'Glavniy Admin') or is_in_group(u, 'Manager'),
    login_url='/login/'
)
def order_create(request):
    """Buyurtma yaratish - Yangi va mavjud IDlar bilan ishlash"""
    
    # 1. Barcha manbalardan ID va ismlarni olish (Bo'sh joylarni tozalab)
    db_customers = {str(c.unique_id).strip(): c.name for c in Customer.objects.all()}
    order_ids = Order.objects.values_list('customer_unique_id', flat=True).distinct()
    
    all_unique_ids = set(list(db_customers.keys()) + [str(uid).strip() for uid in order_ids if uid])
    
    customers_data = []
    for uid in all_unique_ids:
        name = db_customers.get(uid)
        display_text = f"{uid} - {name}" if name else f"{uid}"
        customers_data.append({
            'id': uid,
            'text': display_text
        })

    if request.method == 'POST':
        form = OrderForm(request.POST, request.FILES)
        
        customer_unique_id = request.POST.get('customer_unique_id', '').strip()
        customer_name = request.POST.get('customer_name', '').strip()
        customer_phone = request.POST.get('customer_phone', '').strip() or None

        if form.is_valid():
            try:
                order = form.save(commit=False)
                order.created_by = request.user
                
                if not customer_unique_id:
                    messages.error(request, "❌ Iltimos, mijoz uchun unikal raqam kiriting.")
                else:
                    customer, created = Customer.objects.get_or_create(
                        unique_id=customer_unique_id,
                        defaults={'name': customer_name, 'phone': customer_phone}
                    )
                    
                    if not created and customer_name and (not customer.name or customer.name == "Noma'lum mijoz"):
                        customer.name = customer_name
                        customer.save()

                    order.customer = customer
                    order.customer_unique_id = customer_unique_id
                    order.status = 'KIRITILDI'

                    # MUHIM: eshik_turi va komplekt_* maydonlari HiddenInput orqali keladi,
                    # shuning uchun ularni to'g'ridan-to'g'ri POST'dan olamiz (select2 tag muammosi bo'lmasin)
                    worker_type = form.cleaned_data.get('worker_type', 'LIST')

                    # ================= ESHIK LOGIKASI =================
                    if worker_type in ['ESHIK', 'LIST_ESHIK']:
                        eshik_turi = request.POST.get('eshik_turi', '').strip() or None
                        zamokli_eshik = form.cleaned_data.get('zamokli_eshik', False)
                        if eshik_turi and '(' not in str(eshik_turi):
                            zamok_status = "Zamokli" if zamokli_eshik else "Zamoksiz"
                            order.eshik_turi = f"{eshik_turi} ({zamok_status})"
                        else:
                            order.eshik_turi = eshik_turi

                    # ================= KOMPLEKT LOGIKASI (YANGI) =================
                    if worker_type == 'KOMPLEKT':
                        komplekt_turi = request.POST.get('komplekt_turi', '').strip() or None
                        komplekt_custom_raw = request.POST.get('komplekt_custom', 'False')
                        komplekt_custom = str(komplekt_custom_raw).strip().lower() in ('true', '1', 'on', 'yes')

                        order.komplekt_turi = komplekt_turi
                        order.komplekt_custom = komplekt_custom

                        if komplekt_custom:
                            order.komplekt_kenglik = request.POST.get('komplekt_kenglik') or None
                            order.komplekt_balandligi = request.POST.get('komplekt_balandligi') or None
                            order.komplekt_kvadrat = request.POST.get('komplekt_kvadrat') or None
                        else:
                            order.komplekt_kenglik = None
                            order.komplekt_balandligi = None
                            order.komplekt_kvadrat = None

                    # ================= PANEL LOGIKASI =================
                    if worker_type == 'PANEL':
                        try:
                            u1 = User.objects.get(username='panel_usta')
                            u2 = User.objects.get(username='panel_usta2')
                            last_panel = Order.objects.filter(worker_type='PANEL').order_by('-id').first()
                            order.assigned_to = u2 if last_panel and last_panel.id % 2 != 0 else u1
                        except User.DoesNotExist:
                            pass 

                    order.save()
                    form.save_m2m()
                    
                    try:
                        from .utils import send_notifications_for_new_order 
                        send_notifications_for_new_order(order)
                    except Exception:
                        pass

                    messages.success(request, f"✅ Buyurtma №{order.order_number} kiritildi!")
                    return redirect('order_list')
                
            except Exception as e:
                messages.error(request, f"❌ Tizim xatoligi: {str(e)}")
    else:
        form = OrderForm()

    context = {
        'form': form, 
        'title': 'Yangi Buyurtma Kiritish',
        'customers_list': customers_data,
        'existing_customer_ids': list(all_unique_ids)
    }
    return render(request, 'orders/order_create.html', context)

# orders/views.py - completed_orders_for_loading funksiyasini tuzatamiz

@login_required
def completed_orders_for_loading(request):
    """Omborchi uchun tugatilgan va yuklashga tayyor buyurtmalar"""
    
    # Faqat omborchi yoki adminlar ko'ra oladi
    if not request.user.is_superuser and "omborchi" not in request.user.username.lower():
        messages.error(request, "Bu sahifaga faqat omborchi kirishi mumkin!")
        return redirect('order_list')
    
    try:
        # BARCHA TUGATILGAN STATUSLAR: USTA_TUGATDI, TAYYOR, BAJARILDI
        completed_orders = Order.objects.filter(
            Q(status='USTA_TUGATDI') | Q(status='TAYYOR') | Q(status='BAJARILDI')
        ).order_by('-work_finished_at')
        
        # Ortib bo'lingan buyurtmalar
        loaded_orders = Order.objects.filter(
            Q(status='ORTILDI') | Q(is_loaded=True)
        ).order_by('-loaded_at')[:50]
        
        context = {
            'completed_orders': completed_orders,
            'loaded_orders': loaded_orders,
            'completed_count': completed_orders.count(),
            'loaded_count': loaded_orders.count(),
        }
        return render(request, 'orders/completed_orders_for_loading.html', context)
    
    except Exception as e:
        messages.error(request, f"Xatolik yuz berdi: {str(e)}")
        return redirect('order_list')


@login_required
def mark_order_as_loaded(request, order_id):
    """Buyurtmani ortildi deb belgilash"""
    
    # Faqat omborchi yoki adminlar
    if not request.user.is_superuser and "omborchi" not in request.user.username.lower():
        messages.error(request, "Bu amalni faqat omborchi bajarishi mumkin!")
        return redirect('order_list')
    
    if request.method == 'POST':
        try:
            order = get_object_or_404(Order, id=order_id)
            
            # BARCHA TUGATILGAN STATUSLARNI TEKSHIRAMIZ
            if order.status not in ['USTA_TUGATDI', 'TAYYOR', 'BAJARILDI']:
                messages.error(request, f"{order.order_number} buyurtmasi hali tugatilmagan! Status: {order.status}")
                return redirect('completed_orders_for_loading')
            
            # Ortish amalini bajarish
            order.status = 'ORTILDI'  # Statusni yangilaymiz
            order.is_loaded = True
            order.loaded_at = timezone.now()
            order.loaded_by = request.user
            order.loaded_notes = request.POST.get('loaded_notes', '').strip()
            
            order.save()
            
            messages.success(request, f"✅ {order.order_number} buyurtmasi ortildi deb belgilandi!")
            return redirect('completed_orders_for_loading')
        
        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")
            return redirect('completed_orders_for_loading')
    
    return redirect('completed_orders_for_loading')


@login_required
def unload_order(request, order_id):
    """Ortilgan buyurtmani qaytarish (xato bo'lsa)"""
    
    # Faqat omborchi yoki adminlar
    if not request.user.is_superuser and "omborchi" not in request.user.username.lower():
        messages.error(request, "Bu amalni faqat omborchi bajarishi mumkin!")
        return redirect('order_list')
    
    if request.method == 'POST':
        try:
            order = get_object_or_404(Order, id=order_id)
            
            if order.status != 'ORTILDI' and not order.is_loaded:
                messages.warning(request, f"{order.order_number} buyurtmasi hali ortilmagan!")
                return redirect('completed_orders_for_loading')
            
            # Ortishni bekor qilish - TAYYOR holatiga qaytaramiz
            order.status = 'TAYYOR'
            order.is_loaded = False
            order.loaded_at = None
            order.loaded_by = None
            order.loaded_notes = ""
            
            order.save()
            
            messages.info(request, f"{order.order_number} buyurtmasi ortishdan qaytarildi!")
            return redirect('completed_orders_for_loading')
        
        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")
            return redirect('completed_orders_for_loading')
    
    return redirect('completed_orders_for_loading')



@login_required
def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        form = OrderForm(request.POST, request.FILES, instance=order)
        if form.is_valid():
            form.save()
            messages.success(request, f"#{order.customer_unique_id} buyurtma muvaffaqiyatli yangilandi.")
            return redirect('order_detail', pk=order.pk)
    else:
        form = OrderForm(instance=order)
        # Faqat kerakli ustalarni filtrlab ko'rsatish
        form.fields['assigned_workers'].queryset = Worker.objects.filter(
            Q(user__groups__name="List usta") | Q(user__groups__name="Eshik usta")
        ).distinct()
        
    # Fayl nomi order_edit.html ga o'zgardi
    return render(request, 'orders/order_edit.html', {'form': form, 'is_edit': True})


from django.contrib.admin.models import LogEntry, CHANGE
from django.contrib.contenttypes.models import ContentType

@login_required
def order_confirm(request, pk):
    """2-Bosqich: Buyurtmani tasdiqlash."""
    
    # 1. Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    order = get_object_or_404(Order, pk=pk)
    
    # 2. Ruxsatlarni tekshirish (Admin yoki Menejer bo'lishi shart)
    is_privileged = request.user.is_superuser or is_in_group(request.user, 'Menejer/Tasdiqlovchi')
    
    if not is_privileged: 
        messages.error(request, "Sizda bu buyurtmani tasdiqlash uchun ruxsat yo'q.")
        return redirect('order_list')

    # 3. Statusni yangilash
    if order.status == 'KIRITILDI':
        order.status = 'TASDIQLANDI'
        order.save()
        
        # LogEntry xatosini tuzatish (log_action o'rniga create)
        LogEntry.objects.create(
            user_id=request.user.id,
            content_type_id=ContentType.objects.get_for_model(order).pk,
            object_id=order.pk,
            object_repr=str(order),
            action_flag=CHANGE,
            change_message=f"Status o'zgartirildi: KIRITILDI -> TASDIQLANDI"
        )
        
        messages.success(request, f"Buyurtma №{order.order_number} Tasdiqlandi.")
        
        # 4. Bildirishnomalar yuborish
        # Buyurtma yaratuvchisiga
        if order.created_by:
            Notification.objects.create(
                user=order.created_by,
                order=order,
                message=f"Siz kiritgan buyurtma №{order.order_number} Muvaffaqiyatli Tasdiqlandi."
            )
        
        # Ishlab chiqarish boshlig'iga
        try:
            boss_group = Group.objects.get(name="Ishlab Chiqarish Boshlig'i")
            for boss in boss_group.user_set.all():
                Notification.objects.create(
                    user=boss,
                    order=order,
                    message=f"Yangi buyurtma №{order.order_number} Tasdiqlandi. Ishlab chiqarishni boshlashingiz mumkin."
                )
        except Group.DoesNotExist:
            pass # Xabar berish shart emas bo'lsa

        # Tayinlangan ustalarga
        if order.assigned_workers.exists():
            for worker in order.assigned_workers.all():
                Notification.objects.create(
                    user=worker.user,
                    order=order,
                    message=f"Tayinlangan buyurtma №{order.order_number} Tasdiqlandi! Ishni boshlashingiz mumkin."
                )
            
    else:
        messages.warning(request, "Bu buyurtma allaqachon tasdiqlangan yoki boshqa bosqichda.")
        
    return redirect('order_list')
@login_required
def order_reject(request, pk):
    """Buyurtmani Rad Etish."""
    # Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    order = get_object_or_404(Order, pk=pk)
    
    if not is_in_group(request.user, 'Menejer/Tasdiqlovchi'): 
        messages.error(request, "Sizda bu buyurtmani rad etish uchun ruxsat yo'q.")
        return redirect('order_list')

    if order.status == 'KIRITILDI':
        order.status = 'RAD_ETILDI'
        order.save()
        
        LogEntry.objects.log_action(
            user_id=request.user.id,
            content_type_id=ContentType.objects.get_for_model(order).pk,
            object_id=order.pk,
            object_repr=str(order),
            action_flag=CHANGE,
            change_message=f"Status o'zgartirildi: KIRITILDI -> RAD ETILDI"
        )
        
        messages.error(request, f"Buyurtma №{order.order_number} **Rad Etildi**.")
        
        if order.created_by:
            Notification.objects.create(
                user=order.created_by,
                order=order,
                message=f"Siz kiritgan buyurtma №{order.order_number} Menejer tomonidan **RAD ETILDI**."
            )

        if order.assigned_workers.exists():
            for worker in order.assigned_workers.all():
                Notification.objects.create(
                    user=worker.user,
                    order=order,
                    message=f"Sizga tayinlangan buyurtma №{order.order_number} RAD ETILDI."
                )
        
    else:
        messages.warning(request, "Rad etishni faqat 'Kiritildi' statusidagi buyurtmadan boshlash mumkin.")
        
    return redirect('order_list')

from django.contrib.admin.models import LogEntry, CHANGE
from django.contrib.contenttypes.models import ContentType

@login_required
def order_start_production(request, pk):
    """3-Bosqich: Ishlab chiqarishga berish."""
    
    # 1. Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    order = get_object_or_404(Order, pk=pk)
    
    # 2. Ruxsatlarni tekshirish (Admin yoki Ishlab chiqarish boshlig'i)
    is_privileged = request.user.is_superuser or is_in_group(request.user, "Ishlab Chiqarish Boshlig'i")
    
    if not is_privileged:
        messages.error(request, "Ishlab chiqarishni boshlash uchun ruxsat yo'q.")
        return redirect('order_list')

    # 3. Statusni yangilash
    if order.status == 'TASDIQLANDI':
        order.status = 'ISHDA'
        order.save()
        
        # LogEntry.objects.log_action o'rniga standart .create ishlatamiz
        LogEntry.objects.create(
            user_id=request.user.id,
            content_type_id=ContentType.objects.get_for_model(order).pk,
            object_id=order.pk,
            object_repr=str(order),
            action_flag=CHANGE,
            change_message=f"Status o'zgartirildi: TASDIQLANDI -> ISHDA"
        )
        
        messages.info(request, f"Buyurtma №{order.order_number} ishlab chiqarishga berildi.")
        
        # 4. Bildirishnomalar
        if order.assigned_workers.exists():
            for worker in order.assigned_workers.all():
                Notification.objects.create(
                    user=worker.user,
                    order=order,
                    message=f"Buyurtma №{order.order_number} ISHGA TUSHDI. O'z ishingizni boshlashingiz mumkin."
                )
        
    else:
        messages.warning(request, "Ishlab chiqarishni faqat Tasdiqlangan buyurtmadan boshlash mumkin.")
        
    return redirect('order_list')
from django.contrib.admin.models import LogEntry, CHANGE
from django.contrib.contenttypes.models import ContentType

@login_required
def order_finish(request, pk):
    """4-Bosqich: Buyurtmani yakunlash."""
    # 1. Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    order = get_object_or_404(Order, pk=pk)
    
    # 2. Ruxsatlarni tekshirish (Admin yoki Ishlab chiqarish boshlig'i)
    is_privileged = request.user.is_superuser or is_in_group(request.user, "Ishlab Chiqarish Boshlig'i")
    
    if not is_privileged:
        messages.error(request, "Buyurtmani yakunlash uchun ruxsat yo'q.")
        return redirect('order_list')

    # 3. Statusni yangilash
    if order.status in ['ISHDA', 'USTA_TUGATDI']:
        # DIQQAT: Zanjir ishlashi uchun statusni USTA_TUGATDI qilib saqlash kerak
        order.status = 'USTA_TUGATDI' 
        order.save() # Modeldagi save() avtomatik yangi buyurtma ochishi mumkin
        
        # LogEntry xatosini tuzatish (log_action -> create)
        LogEntry.objects.create(
            user_id=request.user.id,
            content_type_id=ContentType.objects.get_for_model(order).pk,
            object_id=order.pk,
            object_repr=str(order),
            action_flag=CHANGE,
            change_message=f"Status o'zgartirildi: {order.get_status_display()}"
        )
        
        if order.worker_type in ['LIST', 'ESHIK', 'LIST_ESHIK']:
            messages.info(request, "Usta ishini tugatdi. Navbatdagi bosqich (Panel) avtomatik yaratildi.")

        messages.success(request, f"Buyurtma №{order.order_number} yakunlandi.")
        
        # 4. Bildirishnomalar
        try:
            manager_group = Group.objects.get(name='Menejer/Tasdiqlovchi') 
            for manager in manager_group.user_set.all():
                Notification.objects.create(
                    user=manager,
                    order=order,
                    message=f"Buyurtma №{order.order_number} usta tomonidan tugatildi."
                )
        except Group.DoesNotExist:
            pass

    else:
        messages.warning(request, "Buyurtmani yakunlash uchun u jarayonda bo'lishi kerak.")
        
    return redirect('order_list')

@login_required
def order_complete(request, pk):
    """Yakuniy bosqich: Buyurtmani to'liq Bajarildi deb belgilash."""
    # Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    order = get_object_or_404(Order, pk=pk)
    
    if not is_in_group(request.user, 'Menejer/Tasdiqlovchi'): 
        messages.error(request, "Buyurtmani yakunlash uchun ruxsat yo'q.")
        return redirect('order_list')

    if order.status == 'TAYYOR':
        order.status = 'BAJARILDI' 
        order.save()
        
        LogEntry.objects.log_action(
            user_id=request.user.id,
            content_type_id=ContentType.objects.get_for_model(order).pk,
            object_id=order.pk,
            object_repr=str(order),
            action_flag=CHANGE,
            change_message=f"Status o'zgartirildi: TAYYOR -> BAJARILDI (Yakuniy amal)"
        )
        
        messages.success(request, f"Buyurtma №{order.order_number} **BAJARILDI** deb belgilandi. Jarayon to'liq yakunlandi.")
        
        if order.created_by:
            Notification.objects.create(
                user=order.created_by,
                order=order,
                message=f"Siz kiritgan buyurtma №{order.order_number} Muvaffaqiyatli **BAJARILDI**."
            )
        
    else:
        messages.warning(request, "Buyurtma Bajarildi deb belgilanishi uchun u avval 'Tayyor' bo'lishi kerak.")
        
    return redirect('order_list')
@login_required
@user_passes_test(lambda u: u.is_superuser or is_in_group(u, 'Glavniy Admin'), login_url='/login/')
def order_delete(request, pk):
    """Buyurtmani Glavniy Admin tomonidan o'chirish."""
    # Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    order = get_object_or_404(Order, pk=pk)
    
    if request.method == 'POST':
        order_num = order.order_number
        
        LogEntry.objects.log_action(
            user_id=request.user.id,
            content_type_id=ContentType.objects.get_for_model(order).pk,
            object_id=order.pk, 
            object_repr=str(order),
            action_flag=DELETION,
            change_message=f"Buyurtma №{order_num} tizimdan o'chirildi."
        )
        
        order.delete()
        messages.error(request, f"Buyurtma №{order_num} tizimdan butunlay **O'CHIRILDI**.")
        return redirect('order_list')
        
    return render(request, 'orders/order_confirm_delete.html', {'order': order})

from decimal import Decimal
from datetime import datetime, timedelta
from django.db.models import Q
from django.utils import timezone
# ... (boshqa importlar)

@login_required
@user_passes_test(is_report_viewer_or_observer, login_url='/login/')
def weekly_report_view(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Vaqt zonasidan xabardor sanalarni saqlash uchun o'zgaruvchilar
    start_datetime_aware = None
    end_datetime_aware = None

    # ----------------------------------------------------------------------
    # 💡 MUHIM TUZATISH: Sanani TZ-Aware qilish
    # ----------------------------------------------------------------------

    if start_date_str:
        try:
            # 1. Tanlangan sanani Naive Datetime obyektiga o'tkazamiz (00:00:00)
            start_datetime_naive = datetime.strptime(start_date_str, '%Y-%m-%d')
            # 2. Uni loyihaning TIME_ZONE zonasidan xabardor qilamiz (Masalan, Toshkent vaqti)
            start_datetime_aware = timezone.make_aware(start_datetime_naive)
        except ValueError:
             messages.error(request, "Boshlanish sana formati noto'g'ri.")
             start_date_str = None # Noto'g'ri bo'lsa filtrlashni to'xtatamiz

    if end_date_str:
        try:
            # 1. Tanlangan sanani Naive Datetime obyektiga o'tkazamiz (+1 kun)
            end_datetime_naive = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
            # 2. Uni loyihaning TIME_ZONE zonasidan xabardor qilamiz
            end_datetime_aware = timezone.make_aware(end_datetime_naive)
        except ValueError:
            messages.error(request, "Tugash sana formati noto'g'ri.")
            end_date_str = None
            
    # ----------------------------------------------------------------------
    # 1. Umumiy Buyurtmalar (created_at bo'yicha) filtrlash
    # ----------------------------------------------------------------------
    orders = Order.objects.all().select_related('created_by')
    filter_q = Q()
    
    if start_datetime_aware:
        # created_at__gte ni TZ-aware qiymat bilan solishtiramiz
        filter_q &= Q(created_at__gte=start_datetime_aware) 
    if end_datetime_aware:
        # created_at__lt ni TZ-aware qiymat bilan solishtiramiz
        filter_q &= Q(created_at__lt=end_datetime_aware) 
            
    report_orders = orders.filter(filter_q).order_by('-created_at')
    
    # ... (Umumiy statistikani hisoblash)
    
    # ----------------------------------------------------------------------
    # 2. Ustalar Ish Faoliyati Hisoboti (worker_finished_at bo'yicha)
    # ----------------------------------------------------------------------
    worker_report_orders = Order.objects.filter(
        status__in=['TAYYOR', 'BAJARILDI']
    ).filter(
        worker_finished_at__isnull=False 
    ).prefetch_related(
        'assigned_workers__user'
    ) 
    
    worker_filter_q = Q()
    # Yuqorida tayyorlangan TZ-aware obyektlarni qayta ishlatamiz!
    if start_datetime_aware:
        worker_filter_q &= Q(worker_finished_at__gte=start_datetime_aware)
    if end_datetime_aware:
        worker_filter_q &= Q(worker_finished_at__lt=end_datetime_aware)
            
    worker_report_orders = worker_report_orders.filter(worker_filter_q)
    
    # ... (Qolgan loop va kontekst mantig'i)
    
    # Kontekstni yangilash
    context = {
        # ... (boshqa kontekstlar)
        # Template uchun sanalarni string formatida qaytarish
        'start_date': start_date_str,
        'end_date': end_date_str,
        # ...
    }

    return render(request, 'orders/weekly_report_view.html', context)
# orders/views.py

from django.shortcuts import render, redirect 
from django.contrib import messages
from django.contrib.auth.decorators import login_required
# ... (Boshqa importlar qolaversin) ...

# from .forms import MaterialInflowForm, MaterialOutflowForm # Bularni o'chirib tashlaymiz
from .forms import MaterialTransactionForm # YANGI formani import qilamiz

# ... (material_sarfi_report funksiyasi qolaversin) ...

# ===================================================================
# 🔄 YAGONA VIEW: MATERIAL HARAKATINI YARATISH
# ===================================================================

from .forms import (
    OrderForm, 
    StartImageUploadForm, 
    FinishImageUploadForm, 
    OrderStatusForm, 
    MaterialTransactionForm,)
import json
# views.py

# views.py
from django.db import transaction as db_transaction
from django.contrib import messages
import json

# views.py
# views.py - material_transaction_create view ni yangilash
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction as db_transaction
import json
from .forms import MaterialTransactionForm
from .models import Material

import json
import uuid  # ⬅️ Unikal kod uchun shart
from django.db import transaction as db_transaction
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages

import uuid
import json
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction as db_transaction  # Nomini moslashtirdik
from django.contrib.auth.decorators import login_required
from .models import Material, MaterialTransaction
from .forms import MaterialTransactionForm
from django.http import JsonResponse
from django.views.decorators.http import require_GET
import json

@login_required
def material_transaction_create(request):
    """Material kirim/chiqim tranzaksiyasini yaratish."""
    
    if request.method == 'POST':
        form = MaterialTransactionForm(request.POST)
        
        if form.is_valid():
            try:
                with db_transaction.atomic():
                    transaction_obj = form.save(commit=False)
                    transaction_obj.performed_by = request.user
                    
                    # Materialni olish
                    material = transaction_obj.material
                    quantity = transaction_obj.quantity_change
                    transaction_type = transaction_obj.transaction_type
                    
                    # DEBUG
                    print(f"Material: {material.name} (ID: {material.id})")
                    print(f"Quantity: {quantity}")
                    print(f"Type: {transaction_type}")
                    
                    # Barcode yaratish
                    create_barcode = request.POST.get('create_batch_barcode') == 'on'
                    if transaction_type == 'IN' and create_barcode:
                        import uuid
                        new_code = f"P-{uuid.uuid4().hex[:8].upper()}"
                        transaction_obj.transaction_barcode = new_code
                    
                    # Qoldiqni yangilash
                    if transaction_type == 'IN':
                        material.quantity += quantity
                        message_type = "✅ Kirim"
                    else:  # OUT
                        if material.quantity < quantity:
                            raise ValueError(
                                f"Omborda yetarli qoldiq yo'q! "
                                f"Mavjud: {material.quantity} {material.unit}, "
                                f"So'ralgan: {quantity}"
                            )
                        material.quantity -= quantity
                        message_type = "📤 Chiqim"
                    
                    material.save()
                    transaction_obj.save()
                    
                    messages.success(request, 
                        f"{message_type} muvaffaqiyatli bajarildi. "
                        f"Material: {material.name}, "
                        f"Yangi qoldiq: {material.quantity} {material.unit}"
                    )
                    return redirect('material_list')
                    
            except ValueError as e:
                messages.error(request, f"⚠️ {str(e)}")
            except Exception as e:
                messages.error(request, f"❌ Texnik xatolik: {str(e)}")
        else:
            for field, errors in form.errors.items():
                field_name = form.fields[field].label if field in form.fields else field
                messages.error(request, f"{field_name}: {', '.join(errors)}")
    
    else:
        form = MaterialTransactionForm()
    
    # Material ma'lumotlarini JSON formatda yuborish
    materials = Material.objects.all().select_related('category').order_by('name')
    material_data = {}
    
    for mat in materials:
        material_data[str(mat.id)] = {
            'name': mat.name,
            'quantity': float(mat.quantity),
            'unit': mat.unit,
            'category': mat.category.name if mat.category else 'Kategoriyasiz',
            'product_name': mat.product_name if hasattr(mat, 'product_name') else '',
        }
    
    return render(request, 'orders/material_transaction_create.html', {
        'form': form,
        'material_data_json': json.dumps(material_data, ensure_ascii=False),
    })


# ✅ AJAX endpoint material ma'lumotlari uchun
@require_GET
@login_required
def get_material_details(request, material_id):
    """Material ma'lumotlarini JSON formatda qaytarish."""
    try:
        material = Material.objects.select_related('category').get(id=material_id)
        
        data = {
            'id': material.id,
            'name': material.name,
            'code': material.code,
            'quantity': float(material.quantity),
            'unit': material.unit,
            'category': material.category.name if material.category else 'Kategoriyasiz',
            'product_name': material.product_name if hasattr(material, 'product_name') else '',
            'price_per_unit': float(material.price_per_unit) if material.price_per_unit else 0,
            'min_stock_level': float(material.min_stock_level) if material.min_stock_level else 0,
            'success': True
        }
        return JsonResponse(data)
    except Material.DoesNotExist:
        return JsonResponse({'error': 'Material topilmadi', 'success': False}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e), 'success': False}, status=500)


# 🔴 YANGI: Material yaratish view
@login_required
def material_create(request):
    """Yangi material yaratish."""
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            material = form.save()
            messages.success(request, f"✅ Material '{material.name}' muvaffaqiyatli yaratildi.")
            return redirect('material_list')
    else:
        form = MaterialForm()
    
    context = {'form': form}
    return render(request, 'orders/material_form.html', context)


# 🔴 YANGI: Material tahrirlash view
@login_required
def material_edit(request, pk):
    """Materialni tahrirlash."""
    material = get_object_or_404(Material, pk=pk)
    
    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            material = form.save()
            messages.success(request, f"✅ Material '{material.name}' muvaffaqiyatli yangilandi.")
            return redirect('material_list')
    else:
        form = MaterialForm(instance=material)
    
    context = {'form': form, 'material': material}
    return render(request, 'orders/material_form.html', context)

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import F, ExpressionWrapper, DecimalField, Sum, Q, Avg, Count, IntegerField, Case, When
from django.db.models.functions import Coalesce # Coalesce uchun import
from datetime import timedelta
from django.utils import timezone
from django.core.paginator import Paginator
from decimal import Decimal # Decimal ishlatish uchun

# Modellar importi (Material va MaterialTransaction)
from .models import Material, MaterialTransaction 
# Eslatma: Agar modellar boshqa joyda bo'lsa, uni yuqoriga qo'ying

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q, F, DecimalField, IntegerField, Count, Avg, ExpressionWrapper, Case, When
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import timedelta
from django.core.paginator import Paginator
from decimal import Decimal # DecimalField uchun kerak

# Material va MaterialTransaction modellarini import qiling (agar yuqorida bo'lmasa)
# from .models import Material, MaterialTransaction 

@login_required
def material_list(request):
    """
    Omborxona materiallari va tranzaksiyalarini ko'rsatish:
    Barcode va to'g'rilangan related_name bilan.
    """
    filter_low_stock = request.GET.get('low_stock') == 'true'
    filter_has_stock = request.GET.get('has_stock') == 'true'
    filter_type = request.GET.get('type', '') 
    
    current_date = timezone.now()
    today_start = current_date.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = current_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    thirty_days_ago = current_date - timedelta(days=30)
    week_start = current_date - timedelta(days=7)
    
    # 1. MATERIALLAR RO'YXATI
    # ✅ related_name='transactions' bo'lgani uchun 'transactions' so'zi ishlatiladi
    base_materials_qs = Material.objects.annotate(
        difference=ExpressionWrapper(
            F('quantity') - F('min_stock_level'),
            output_field=DecimalField(max_digits=15, decimal_places=3)
        ),
        total_value=ExpressionWrapper(
            F('quantity') * F('price_per_unit'),
            output_field=DecimalField(max_digits=15, decimal_places=2)
        ), 
    ).order_by('name') 

    materials = base_materials_qs
    if filter_low_stock:
        materials = materials.filter(quantity__lt=F('min_stock_level'))
    if filter_has_stock:
        materials = materials.filter(quantity__gt=0)
    
    paginator = Paginator(materials, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 2. KIRIM TRANZAKSIYALARI (Barcode qo'shilgan)
    incoming_transactions_raw = MaterialTransaction.objects.filter(
        transaction_type='IN'
    ).select_related('material').order_by('-timestamp')[:100]
    
    income_materials = []
    for tx in incoming_transactions_raw:
        material = tx.material
        quantity = tx.quantity_change
        unit_price = material.price_per_unit if material else Decimal('0') 
        
        income_materials.append({
            'id': tx.id,
            'date': tx.timestamp,
            'barcode': tx.transaction_barcode,  # 🔴 BARCODE SHU YERDA
            'material': {
                'name': material.name if material else 'Noma\'lum',
                'product_name': material.product_name if material and hasattr(material, 'product_name') else '',
                'unit': material.unit if material else ''
            },
            'quantity': quantity,
            'price': unit_price,
            'total': quantity * unit_price,
            'notes': tx.notes or '',
            'supplier': tx.received_by or ''
        })
    
    # 3. CHIQIM TRANZAKSIYALARI
    outgoing_transactions_raw = MaterialTransaction.objects.filter(
        transaction_type='OUT'
    ).select_related('material').order_by('-timestamp')[:100]

    outcome_materials = []
    for tx in outgoing_transactions_raw:
        material = tx.material
        quantity = abs(tx.quantity_change)
        unit_price = material.price_per_unit if material else Decimal('0') 
        outcome_materials.append({
            'id': tx.id,
            'date': tx.timestamp,
            'barcode': tx.transaction_barcode,
            'material': {
                'name': material.name if material else 'Noma\'lum',
                'unit': material.unit if material else ''
            },
            'quantity_change': tx.quantity_change,
            'unit_price': unit_price,
            'total_value': quantity * unit_price,
            'notes': tx.notes or '',
            'department': tx.received_by or '' 
        })
        
    # 4. TOP MATERIALLAR (Xatolik tuzatilgan qism)
    # ✅ materialtransaction o'rniga transactions ishlatildi
    # views.py 1441-qator atrofini shunday o'zgartiring:

    top_materials_qs = Material.objects.annotate(
        # total_incoming hisobi
        total_incoming=Coalesce(
            Sum('materialtransaction__quantity_change', 
                filter=Q(materialtransaction__transaction_type='IN')), 
            Decimal('0'), 
            output_field=DecimalField(max_digits=15, decimal_places=3)
        ),
        # total_outgoing hisobi
        total_outgoing=Coalesce(
            Sum('materialtransaction__quantity_change', 
                filter=Q(materialtransaction__transaction_type='OUT')), 
            Decimal('0'), 
            output_field=DecimalField(max_digits=15, decimal_places=3)
        ),
        # incoming_value hisobi
        incoming_value=ExpressionWrapper(
            Coalesce(
                Sum('materialtransaction__quantity_change', 
                    filter=Q(materialtransaction__transaction_type='IN')), 
                Decimal('0')
            ) * F('price_per_unit'),
            output_field=DecimalField(max_digits=15, decimal_places=2) 
        ),
        # outgoing_value hisobi
        outgoing_value=ExpressionWrapper(
            Coalesce(
                Sum('materialtransaction__quantity_change', 
                    filter=Q(materialtransaction__transaction_type='OUT')), 
                Decimal('0')
            ) * F('price_per_unit'),
            output_field=DecimalField(max_digits=15, decimal_places=2) 
        )
    )
    
    top_incoming_materials = top_materials_qs.filter(total_incoming__gt=0).order_by('-total_incoming')[:10]
    top_outgoing_materials = top_materials_qs.filter(total_outgoing__gt=0).order_by('-total_outgoing')[:10]
    
    context = {
        'materials': page_obj,
        'page_obj': page_obj,
        'income_materials': income_materials,
        'outcome_materials': outcome_materials,
        'top_incoming_materials': top_incoming_materials,
        'top_outgoing_materials': top_outgoing_materials,
        'current_date': current_date,
        'title': 'Omborxona Boshqaruvi',
    }
    
    return render(request, 'orders/material_list.html', context)
# orders/views.py

from django.shortcuts import render
# from .models import Material # Modellar ham import qilingan bo'lishi kerak

# ... (Boshqa view funksiyalaringiz)

# orders/views.py

from django.shortcuts import render
from django.db.models import Sum
from django.utils import timezone
from datetime import timedelta
from .models import Material, MaterialTransaction # Modellar shu yerda import qilinishi kerak

# orders/views.py

from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
def fast_scanner_view(request):
    """
    Tezkor Skanerlash uchun alohida sahifani render qiladi.
    Bu sahifada faqat Kirim/Chiqim rejimi va Skanerlash maydoni bo'ladi.
    """
    context = {
        'title': 'Tezkor Skanerlash Markazi',
    }
    # orders/fast_scanner.html shablonini chaqirish
    return render(request, 'orders/fast_scanner.html', context)


# views.py ga qo'shing
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

# orders/views.py
from orders.models import Material
from django.http import JsonResponse
from django.db import transaction as db_transaction
from django.contrib.auth.decorators import login_required

@login_required
def find_material_by_code_api(request):
    if request.method == 'GET':
        code = request.GET.get('code', '').strip()
        
        if not code:
            return JsonResponse({'success': False, 'error': 'Kod kiritilmadi.'}, status=400)
        
        try:
            # 1. Materialni topishga urinish (product_name yoki code orqali)
            material = Material.objects.filter(
                product_name__iexact=code
            ).first() or Material.objects.filter(code__iexact=code).first()
            if material:
                is_new = False
            else:
                raise Material.DoesNotExist
        except Material.DoesNotExist:
            # 2. Agar topilmasa, uni avtomatik yaratish!
            try:
                with db_transaction.atomic():
                    material = Material.objects.create(
                        name=f"Yangi Material (Kod: {code})",
                        product_name=code, # Kodni bu yerga saqlash
                        unit='dona',
                        quantity=0, # Boshlang'ich qoldiq
                        price_per_unit=0
                        # Agar modelda boshqa majburiy maydonlar bo'lsa, ularni qo'shing (masalan, category_id)
                    )
                is_new = True
            except Exception as create_error:
                return JsonResponse({
                    'success': False, 
                    'error': f"Avtomatik yaratishda xato: {create_error}"
                }, status=500)
        
        # 3. Natijani qaytarish
        return JsonResponse({
            'success': True,
            'material_id': material.id,
            'material_name': material.name,
            'material_code': material.product_name, # Kod sifatida product_name ni yuborish
            'material_unit': material.unit,
            'scanned_raw_code': code, # <-- Shu yerda code o'zgaruvchisi yuborilmoqda
            # ... boshqa maydonlar
            'is_new': is_new
        })
        
    return JsonResponse({'success': False, 'error': 'Faqat GET so\'rovi qabul qilinadi.'}, status=405)

@login_required
@csrf_exempt  # Faqat test uchun
def save_scanned_transactions_api(request):
    """API: Skanerlangan tranzaksiyalarni saqlash"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Bu yerda ma'lumotlarni saqlash logikasi
            # Misol:
            # for item in data['items']:
            #     Transaction.objects.create(...)
            
            return JsonResponse({
                'success': True,
                'message': f"{len(data.get('items', []))} ta element saqlandi"
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
# orders/views.py

from django.shortcuts import render, redirect, get_object_or_404
from .forms import MaterialTransactionForm # Forma mavjudligini taxmin qilamiz
# ... (boshqa importlar)

@login_required
def add_transaction_view(request):
    """
    Yangi omborxona harakatini (kirim yoki chiqim) qo'shish.
    """
    if request.method == 'POST':
        form = MaterialTransactionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('material_list') # Muvaffaqiyatli saqlangandan keyin inventarizatsiya sahifasiga qaytish
    else:
        form = MaterialTransactionForm()
        
    context = {
        'title': 'Yangi Harakat Qo\'shish (Kirim/Chiqim)',
        'form': form
    }
    return render(request, 'orders/add_transaction.html', context)



# orders/views.py

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from .models import Material, MaterialTransaction
from django.shortcuts import get_object_or_404



@require_POST
@login_required
def remove_transaction_view(request):
    """
    Omborxona materialini chiqim qilish (omborxona zaxirasidan olib tashlash)
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Invalid Request'}, status=400)
    
    try:
        material_id = request.POST.get('material_id')
        quantity = float(request.POST.get('quantity'))
        reason = request.POST.get('reason', 'Chiqim sababi ko\'rsatilmadi')
        
        material = get_object_or_404(Material, pk=material_id)
        
        if quantity <= 0:
            return JsonResponse({'success': False, 'error': 'Noto\'g\'ri miqdor kiritildi'}, status=400)
        
        with transaction.atomic():
            # Zaxira yetarli yoki yo'qligini tekshirish
            if material.quantity < quantity:
                return JsonResponse({'success': False, 'error': f'Zaxirada yetarli {material.unit} mavjud emas. (Mavjud: {material.quantity})'}, status=400)
            
            # 1. Zaxirani yangilash
            material.quantity -= quantity
            material.save()
            
            # 2. Tranzaksiyani yaratish (Chiqim)
            MaterialTransaction.objects.create(
                material=material,
                transaction_type='OUT', # Chiqim
                quantity_change=quantity,
                performed_by=request.user if request.user.is_authenticated else None,
                notes=reason
            )
        
        return JsonResponse({'success': True, 'message': 'Chiqim muvaffaqiyatli amalga oshirildi.'})
        
    except Material.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Material topilmadi.'}, status=404)
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Miqdor noto\'g\'ri formatda.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Kutilmagan xato: {str(e)}'}, status=500)
    


@login_required
def material_transaction_delete(request, pk):
    """Material tranzaksiyasini o'chirish."""
    transaction = get_object_or_404(MaterialTransaction, pk=pk)
    
    if request.method == 'POST':
        try:
            # Qaytarish logikasi (agar kerak bo'lsa)
            material = transaction.material
            if transaction.transaction_type == 'IN':
                material.quantity -= transaction.quantity_change
            else:  # OUT
                material.quantity += transaction.quantity_change
            
            material.save()
            transaction.delete()
            
            messages.success(request, "✅ Tranzaksiya muvaffaqiyatli o'chirildi.")
            return redirect('material_list')
            
        except Exception as e:
            messages.error(request, f"❌ Xatolik: {str(e)}")
    
    context = {
        'transaction': transaction,
    }
    
    return render(request, 'orders/material_transaction_confirm_delete.html', context)
def get_material_data():
    """
    Material ma'lumotlarini JSON uchun tayyorlash
    """
    material_objects = Material.objects.all().select_related('category').values(
        'id', 'name', 'unit', 'quantity', 'category__name'
    )
    
    material_data = {
        str(m['id']): {
            'name': m['name'], 
            'unit': m['unit'], 
            'quantity': float(m['quantity']) if m['quantity'] is not None else 0,
            'category': m['category__name'] if m['category__name'] else 'Kategoriyasiz'
        } 
        for m in material_objects
    }
    
    return material_data


# orders/views.py


from django.db.models.functions import Coalesce # ⬅️ Mana bu qatorni qo'shing
from decimal import Decimal # ✅ Decimal to'g'ri import qilindi
from decimal import Decimal
from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.shortcuts import render
from .models import Order
from django.db.models import Sum, Q, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal
from django.shortcuts import render
from .models import Order

def material_sarfi_report(request):
    """
    Asosiy buyurtmalar (parent) asosida materiallar sarfini 
    aniq koeffitsientlar bilan hisoblaydigan hisobot.
    """
    
    # 1. FAQAT ASOSIY BUYURTMALARNI FILTRLASH (Dublikat bo'lmasligi uchun)
    parent_orders = Order.objects.filter(parent_order__isnull=True)

    # 2. BAZAGA FAQAT 1 MARTA MUROJAAT QILIB, BARCHA SUMMALARNI OLAMIZ
    # Bu usul filter(thickness=...).aggregate() dan 5 baravar tezroq ishlaydi
    stats = parent_orders.aggregate(
        kv_5=Coalesce(Sum('panel_kvadrat', filter=Q(panel_thickness='5')), Decimal('0')),
        kv_8=Coalesce(Sum('panel_kvadrat', filter=Q(panel_thickness='8')), Decimal('0')),
        kv_10=Coalesce(Sum('panel_kvadrat', filter=Q(panel_thickness='10')), Decimal('0')),
        kv_15=Coalesce(Sum('panel_kvadrat', filter=Q(panel_thickness='15')), Decimal('0')),
        total_kv=Coalesce(Sum('panel_kvadrat'), Decimal('0'))
    )

    # 3. STATS'DAN QIYMATLARNI ALOHIDA O'ZGARUVCHILARGA OLAMIZ
    sum_kv_5 = stats['kv_5']
    sum_kv_8 = stats['sum_8'] if 'sum_8' in stats else stats['kv_8'] # xavfsizlik uchun
    sum_kv_10 = stats['kv_10']
    sum_kv_15 = stats['kv_15']
    total_kvadrat = stats['total_kv']

    # 4. SIRYO SARFINI QALINLIK BO'YICHA ANIQ HISOBLASH
    # Formulalar: 5cm->x2, 8cm->x3, 10cm->x4, 15cm->x6
    siryo_5_sarfi = sum_kv_5 * Decimal('2')
    siryo_8_sarfi = sum_kv_8 * Decimal('3')
    siryo_10_sarfi = sum_kv_10 * Decimal('4')
    siryo_15_sarfi = sum_kv_15 * Decimal('6')

    # Jami Siryo (kg)
    jami_siryo_sarfi = siryo_5_sarfi + siryo_8_sarfi + siryo_10_sarfi + siryo_15_sarfi

    # 5. JAMI LIST SARFI (m²): (Total Kvadrat * 2) + 10 metr zapas
    # Faqat buyurtma mavjud bo'lsagina zapas qo'shiladi
    if total_kvadrat > 0:
        jami_list_sarfi = (total_kvadrat * Decimal('2')) + Decimal('10')
    else:
        jami_list_sarfi = Decimal('0')

    # 6. CONTEXT - HAMMA MA'LUMOTLARNI TEMPLATE'GA YUBORAMIZ
    context = {
        # Jami natijalar
        'total_kvadrat': total_kvadrat,
        'jami_list_sarfi': jami_list_sarfi,
        'jami_siryo_sarfi': jami_siryo_sarfi,
        
        # Har bir qalinlik bo'yicha kvadratlar
        'sum_kv_5': sum_kv_5,
        'sum_kv_8': sum_kv_8,
        'sum_kv_10': sum_kv_10,
        'sum_kv_15': sum_kv_15,

        # Har bir qalinlik bo'yicha siryo sarfi
        'siryo_5_sarfi': siryo_5_sarfi,
        'siryo_8_sarfi': siryo_8_sarfi,
        'siryo_10_sarfi': siryo_10_sarfi,
        'siryo_15_sarfi': siryo_15_sarfi,
    }
    
    return render(request, 'orders/material_sarfi_report.html', context)

from django.db.models import Q, Sum, Count, F
from django.utils.dateparse import parse_date
from datetime import timedelta

from django.db.models import Sum, Count, F

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, Q
from .models import Order

from django.db.models import Sum, Count, F, Q

@login_required
def worker_activity_report_view(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # 1. Archive dagi kabi statuslar
    archive_statuses = ['BAJARILDI', 'USTA_TUGATDI', 'TAYYOR']
    
    # 2. Barcha buyurtmalar (Asosiy va ichki)
    # Xuddi Archive dagi kabi filtrlaymiz
    orders = Order.objects.filter(status__in=archive_statuses)

    # 3. Sana bo'yicha filtr (Agar sana tanlangan bo'lsa)
    if start_date:
        orders = orders.filter(worker_finished_at__date__gte=start_date)
    if end_date:
        orders = orders.filter(worker_finished_at__date__lte=end_date)

    # 4. Ustalar bo'yicha guruhlash
    # Har bir orderning 'assigned_workers' maydonidan foydalanamiz
    worker_report_list = orders.values(
        'assigned_workers__id'
    ).annotate(
        first_name=F('assigned_workers__user__first_name'),
        last_name=F('assigned_workers__user__last_name'),
        username=F('assigned_workers__user__username'),
        total_order_count=Count('id'),
        total_finished_kvadrat=Sum('panel_kvadrat')
    ).filter(assigned_workers__id__isnull=False).order_by('-total_finished_kvadrat')

    # 5. Jami summa
    total_kv = orders.aggregate(Sum('panel_kvadrat'))['panel_kvadrat__sum'] or 0

    context = {
        "title": "Ustalar Ish Faoliyati Hisoboti",
        'worker_report_list': worker_report_list,
        'total_finished_kvadrat': total_kv,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'orders/weekly_report_view.html', context)

@login_required
@user_passes_test(is_report_viewer_or_observer, login_url='/')  # ✅ YANGI
def export_worker_activity_csv(request):
    """
    Ustalarning ish faoliyati hisobotini CSV fayl shaklida eksport qiladi.
    """
    # 1. CSV javobini tayyorlash
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="usta_faoliyat_hisoboti_{datetime.now().strftime("%Y-%m-%d")}.csv"'
    
    writer = csv.writer(response)
    
    # Jadval sarlavhalari
    writer.writerow([
        'T/r', 
        'Usta F.I.Sh.', 
        'Bajarilgan Kvadratura (m²)', 
        'Bajarilgan Buyurtmalar Soni'
    ])

    # 2. Filtrlash shartlarini yaratish
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # ✅ Xato tuzatildi: Statusni to'g'ri string qiymati bilan filtrlash
    worker_filter_q = Q(status='FINISHED') 
    
    start_date = None
    end_date = None

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            worker_filter_q &= Q(updated_at__date__gte=start_date)
        except ValueError:
            pass 

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            worker_filter_q &= Q(updated_at__date__lte=end_date)
        except ValueError:
            pass

    # 3. Hisobot ma'lumotlarini olish
    worker_report_list = Order.objects.filter(worker_filter_q).exclude(assigned_workers__isnull=True).values(
        'assigned_workers__user__first_name',
        'assigned_workers__user__last_name'
    ).annotate(
    total_finished_kvadrat=Sum('panel_kvadrat', default=0), 
    total_order_count=Count('id')

    ).order_by('assigned_workers__user__first_name')
    
    # 4. CSV ga ma'lumotlarni yozish
    for i, worker in enumerate(worker_report_list):
        writer.writerow([
            i + 1,
            f"{worker['assigned_workers__user__first_name']} {worker['assigned_workers__user__last_name']}",
            f"{worker['total_finished_kvadrat']:.2f}",
            worker['total_order_count']
        ])
        
    return response

@login_required
@user_passes_test(lambda u: u.is_superuser or is_in_group(u, 'Glavniy Admin'), login_url='/login/')
def export_orders_csv(request):
    """Buyurtmalarni CSV formatida eksport qilish (oxirgi 7 kunlik)."""
    # Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="EcoProm_Buyurtmalar_Hisboti_7kunlik.csv"'

    writer = csv.writer(response, delimiter=';') 

    writer.writerow([
        "Buyurtma Raqami", 
        "Xaridor Nomi", 
        "Kvadrat (m²)", 
        "Summa (so'm)", 
        "Status", 
        "Kiritilgan Sana",
        "Kiritgan Xodim"
    ])

    seven_days_ago = date.today() - timedelta(days=7)
    orders = Order.objects.filter(
        created_at__gte=seven_days_ago
    ).order_by('-created_at')

    for order in orders:
        writer.writerow([
            order.order_number,
            order.customer_name,
            order.panel_kvadrat,
            order.total_price,
            order.get_status_display(),
            order.created_at.strftime("%Y-%m-%d %H:%M"),
            order.created_by.get_full_name() if order.created_by else "Noma'lum", 
        ])

    return response

@login_required
def sales_report_view(request):
    import logging
    logger = logging.getLogger(__name__)
    
    # Foydalanuvchi ma'lumotlarini logga yozish
    logger.info(f"User: {request.user.username}")
    logger.info(f"Is authenticated: {request.user.is_authenticated}")
    logger.info(f"Groups: {[g.name for g in request.user.groups.all()]}")
    logger.info(f"Is report viewer: {is_report_viewer(request.user)}")
    logger.info(f"Is observer: {is_observer(request.user)}")
    
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    today = date.today()
    
    try:
        if start_date_str and end_date_str:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
        else:
            start_date = today - timedelta(days=30)
            end_date = today
    except ValueError:
        messages.error(request, "Noto'g'ri sana formati kiritildi. Iltimos, YYYY-MM-DD formatida kiriting.")
        start_date = today - timedelta(days=30)
        end_date = today

    # 🔴 Asosiy buyurtmalar
    main_orders = Order.objects.filter(
        parent_order__isnull=True,  # Faqat asosiy buyurtmalar
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).order_by('-created_at')

    # 🔴 Child buyurtmalar (alohida)
    child_orders = Order.objects.filter(
        parent_order__isnull=False,  # Faqat child buyurtmalar
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).order_by('-created_at')

    # 🔴 Barcha buyurtmalar (umumiy ko'rish uchun)
    all_orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).order_by('-created_at')

    total_orders_count = main_orders.count()
    total_square = main_orders.aggregate(Sum('panel_kvadrat'))['panel_kvadrat__sum'] or 0
    total_revenue = main_orders.aggregate(Sum('total_price'))['total_price__sum'] or 0

    context = {
        "title": "Sotuv Hisoboti (Vaqt Oralig'i)",
        'report_orders': main_orders,  # 🔴 Faqat asosiylar ko'rsatiladi
        'child_orders': child_orders,  # 🔴 Child buyurtmalar (alohida)
        'all_orders': all_orders,  # 🔴 Barcha buyurtmalar
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'total_orders_count': total_orders_count,
        'total_square': total_square,
        'total_revenue': total_revenue,
        'is_glavniy_admin': True,
        'today': timezone.now().date(),
        'is_observer': is_observer(request.user),
    }
    return render(request, 'orders/sales_report.html', context)

@login_required
def export_sales_report_excel(request):
    """Excel formatda savdo hisobotini eksport qilish"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # Ruxsatni tekshirish
    if not request.user.is_superuser and not request.user.is_staff:
        messages.error(request, "Sizda bu amalni bajarish uchun ruxsat yo'q!")
        return redirect('sales_report_view')
    
    # Filter parametrlarini olish
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Asosiy buyurtmalarni olish (filter bilan)
    orders = Order.objects.filter(parent_order__isnull=True)
    
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass
    
    # Excel fayl yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savdo Hisoboti"
    
    # Stil sozlamalari
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sarlavha qatori
    headers = [
        "№", 
        "Buyurtma Raqami", 
        "Xaridor", 
        "Mahsulot", 
        "Ish Turi", 
        "Kvadrat (m²)", 
        "Summa (usd)", 
        "Holat", 
        "Sana",
        "Chizma (PDF)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Ma'lumotlar
    total_square = 0
    total_revenue = 0
    
    # Saytning asosiy URL'ini olish
    base_url = request.build_absolute_uri('/')[:-1]
    
    for idx, order in enumerate(orders, 1):
        row = idx + 1
        
        # Holatni o'zbekcha ko'rinish
        status_dict = dict(Order.STATUS_CHOICES)
        status_text = status_dict.get(order.status, order.status)
        
        # PDF URL'ini yaratish
        pdf_url = ''
        display_text = "Mavjud emas"
        
        try:
            # 1. Orderda pdf_file field'i bo'lsa (media fayl)
            if hasattr(order, 'pdf_file') and order.pdf_file:
                if hasattr(order.pdf_file, 'url'):
                    pdf_url = base_url + order.pdf_file.url
                    display_text = "PDF ni ko'rish"
                    print(f"PDF URL (order.pdf_file): {pdf_url}")
            
            # 2. Orderda drawing field'i bo'lsa (PDF fayl)
            elif hasattr(order, 'drawing') and order.drawing:
                if hasattr(order.drawing, 'url'):
                    pdf_url = base_url + order.drawing.url
                    display_text = "PDF ni ko'rish"
                    print(f"PDF URL (order.drawing): {pdf_url}")
            
            # 3. Orderda drawing_file field'i bo'lsa
            elif hasattr(order, 'drawing_file') and order.drawing_file:
                if hasattr(order.drawing_file, 'url'):
                    pdf_url = base_url + order.drawing_file.url
                    display_text = "PDF ni ko'rish"
                    print(f"PDF URL (order.drawing_file): {pdf_url}")
            
            # 4. Project orqali PDF
            elif hasattr(order, 'project') and order.project:
                project = order.project
                
                # Projectda pdf_file bo'lsa
                if hasattr(project, 'pdf_file') and project.pdf_file:
                    if hasattr(project.pdf_file, 'url'):
                        pdf_url = base_url + project.pdf_file.url
                        display_text = "PDF ni ko'rish"
                        print(f"PDF URL (project.pdf_file): {pdf_url}")
                
                # Projectda drawing bo'lsa
                elif hasattr(project, 'drawing') and project.drawing:
                    if hasattr(project.drawing, 'url'):
                        pdf_url = base_url + project.drawing.url
                        display_text = "PDF ni ko'rish"
                        print(f"PDF URL (project.drawing): {pdf_url}")
            
            # 5. Agar PDF fayl bo'lmasa, "Mavjud emas" deb qoldiramiz
            
        except Exception as e:
            print(f"PDF URL xatosi (Order {order.order_number}): {e}")
            pdf_url = ''
            display_text = "Xatolik"
        
        # Ma'lumotlarni yozish
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=order.order_number)
        ws.cell(row=row, column=3, value=order.customer_name)
        ws.cell(row=row, column=4, value=order.product_name)
        ws.cell(row=row, column=5, value=order.get_worker_type_display())
        ws.cell(row=row, column=6, value=float(order.panel_kvadrat) if order.panel_kvadrat else 0)
        ws.cell(row=row, column=7, value=float(order.total_price) if order.total_price else 0)
        ws.cell(row=row, column=8, value=status_text)
        ws.cell(row=row, column=9, value=order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else '')
        
        # Chizma ustuni - giperhavola qo'shish
        cell = ws.cell(row=row, column=10, value=display_text)
        if pdf_url:
            cell.hyperlink = pdf_url
            cell.font = Font(color="0563C1", underline="single")
        else:
            cell.font = Font(color="9CA3AF")
        
        # Yig'indilarni hisoblash
        total_square += order.panel_kvadrat or 0
        total_revenue += order.total_price or 0
        
        # Har bir qatorga border
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = border
    
    # Jami qatori
    total_row = len(orders) + 2
    
    # "JAMI" yozuvi
    cell = ws.cell(row=total_row, column=1, value="JAMI:")
    cell.font = Font(bold=True, size=11)
    cell.border = border
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    
    # Jami kvadrat
    cell = ws.cell(row=total_row, column=6, value=float(total_square))
    cell.font = Font(bold=True, color="10B981")
    cell.border = border
    cell.alignment = Alignment(horizontal="center")
    
    # Jami summa
    cell = ws.cell(row=total_row, column=7, value=float(total_revenue))
    cell.font = Font(bold=True, color="F59E0B")
    cell.border = border
    cell.alignment = Alignment(horizontal="center")
    
    # Buyurtmalar soni
    cell = ws.cell(row=total_row, column=8, value=f"{orders.count()} ta")
    cell.font = Font(bold=True, color="4F46E5")
    cell.border = border
    cell.alignment = Alignment(horizontal="center")
    
    # Qolgan ustunlar
    for col in range(9, len(headers) + 1):
        ws.cell(row=total_row, column=col).border = border
    
    # Ustun kengliklarini avtomatik sozlash
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, total_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Chizma ustunini kengroq qilamiz
    ws.column_dimensions['J'].width = 20
    
    # Sarlavha qatorini muzlatish
    ws.freeze_panes = 'A2'
    
    # HTTP response tayyorlash
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'savdo_hisoboti_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
@user_passes_test(lambda u: u.is_superuser or is_in_group(u, 'Glavniy Admin'), login_url='/login/')
def product_audit_log_view(request):
    """Mahsulot/Buyurtma o'zgarishlari jurnalini ko'rish funksiyasi."""
    # Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    try:
        order_content_type = ContentType.objects.get(app_label='orders', model='order')
        
        log_entries = LogEntry.objects.filter(
            content_type=order_content_type
        ).select_related('user').order_by('-action_time')
        
        for log in log_entries:
            try:
                log.related_object_name = Order.objects.get(pk=log.object_id).order_number
            except Order.DoesNotExist:
                log.related_object_name = log.object_repr
    except ContentType.DoesNotExist:
        log_entries = []
        messages.error(request, "Buyurtma (Order) modeli uchun ContentType topilmadi. `LogEntry` filtrlanmaydi.")
    

    context = {
        "title": "Mahsulot O'zgarishlari Jurnali (Audit Log)",
        'log_entries': log_entries,
        'is_glavniy_admin': True,
        'is_observer': False, 
    }
    
    return render(request, 'orders/product_audit_log.html', context)

@login_required
@user_passes_test(lambda u: u.is_superuser or is_in_group(u, 'Glavniy Admin'), login_url='/login/')
def export_audit_log_csv(request):
    """Audit Log yozuvlarini CSV formatida eksport qilish."""
    # Kuzatuvchi tekshiruvi
    if is_observer(request.user):
        messages.error(request, "Kuzatuvchi rejimida bu amalni bajarish mumkin emas.")
        return redirect('order_list')
        
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="EcoProm_Audit_Log_Hisoboti.csv"'

    writer = csv.writer(response, delimiter=';') 

    writer.writerow([
        "Harakat Vaqti", 
        "Foydalanuvchi", 
        "Harakat Turi", 
        "Buyurtma Raqami/Obyekt", 
        "O'zgarish Tafsiloti (Change Message)"
    ])

    try:
        order_content_type = ContentType.objects.get(app_label='orders', model='order')
        log_entries = LogEntry.objects.filter(
            content_type=order_content_type
        ).select_related('user').order_by('-action_time')
    except ContentType.DoesNotExist:
        messages.error(request, "Buyurtma (Order) modeli ContentType topilmadi. Eksport qilish bekor qilindi.")
        return redirect('product_audit_log_view') 

    def get_action_type(flag):
        if flag == ADDITION:
            return 'Yaratildi (ADDITION)'
        elif flag == CHANGE:
            return 'Tahrirlandi (CHANGE)'
        elif flag == DELETION:
            return 'Oʻchirildi (DELETION)'
        return 'Nomaʼlum'

    for log in log_entries:
        
        object_identifier = ''
        try:
            object_identifier = Order.objects.get(pk=log.object_id).order_number
        except Order.DoesNotExist:
            object_identifier = f"O'chirilgan obyekti (ID: {log.object_id})"

        
        writer.writerow([
            log.action_time.strftime("%Y-%m-%d %H:%M:%S"),
            log.user.get_full_name() or log.user.username,
            get_action_type(log.action_flag),
            object_identifier,
            log.change_message.replace('\r\n', ' ').replace('\n', ' ') 
        ])

    return response

from django.db.models import F

from django.db.models import F, Q
from django.db.models import F, Q, Sum, DecimalField, ExpressionWrapper, Value
from django.db.models.functions import Coalesce
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator

import io
import pandas as pd
from django.db.models import F, Q, Sum, DecimalField, ExpressionWrapper, Value
from django.db.models.functions import Coalesce
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse

@login_required
def debt_report(request):
    # 1. Ruxsatlarni tekshirish
    user = request.user
    is_admin = user.is_superuser or is_in_group(user, 'Glavniy Admin')
    is_manager = is_in_group(user, 'Menejer/Tasdiqlovchi')
    
    if not (is_admin or is_manager):
        messages.error(request, "Sizga bu sahifani ko'rishga ruxsat yo'q!")
        return redirect('order_list')
    
    # 2. Parametrlarni olish
    search_query = request.GET.get('q', '').strip()
    sort_by = request.GET.get('sort', 'debt_desc')
    
    # 3. Asosiy QuerySet - Hisob-kitobni SQL darajasida qilish
    debts = Order.objects.annotate(
        calculated_debt=ExpressionWrapper(
            Coalesce(F('total_price'), Value(0, output_field=DecimalField())) - 
            Coalesce(F('prepayment'), Value(0, output_field=DecimalField())),
            output_field=DecimalField()
        )
    ).filter(
        calculated_debt__gt=0
    ).exclude(
        status__in=['BEKOR_QILINDI', 'RAD_ETILDI']
    ).select_related('parent_order')

    # 4. Qidiruv
    if search_query:
        debts = debts.filter(
            Q(order_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(customer_unique_id__icontains=search_query) |
            Q(product_name__icontains=search_query)
        )
    
    # 5. Saralash
    sort_dict = {
        'debt_desc': '-calculated_debt',
        'debt_asc': 'calculated_debt',
        'date_desc': '-created_at',
        'date_asc': 'created_at',
        'name_asc': 'customer_name'
    }
    debts = debts.order_by(sort_dict.get(sort_by, '-calculated_debt'))
    
    # 6. Excel Export - Professional Formatlash bilan
    if request.GET.get('export') == 'excel':
        export_data = []
        for order in debts:
            export_data.append({
                'Buyurtma №': order.order_number,
                'Mijoz nomi': order.customer_name,
                'Mijoz ID': order.customer_unique_id,
                'Mahsulot': order.product_name,
                'Umumiy summa ($)': float(order.total_price or 0),
                'To\'langan ($)': float(order.prepayment or 0),
                'Qoldiq qarz ($)': float(order.calculated_debt),
                'Holat': order.get_status_display(),
                'Sana': order.created_at.strftime('%d.%m.%Y') if order.created_at else '-',
            })
        
        df = pd.DataFrame(export_data)
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Qarzlar Ro\'yxati')
        
        workbook  = writer.book
        worksheet = writer.sheets['Qarzlar Ro\'yxati']

        # Formatlar
        header_fmt = workbook.add_format({'bold': True, 'bg_color': '#27ae60', 'font_color': 'white', 'border': 1})
        money_fmt = workbook.add_format({'num_format': '#,##0.00', 'border': 1})
        border_fmt = workbook.add_format({'border': 1})

        # Sarlavhalarni formatlash va kenglikni sozlash
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_fmt)
            column_len = max(df[value].astype(str).str.len().max(), len(value)) + 5
            worksheet.set_column(col_num, col_num, column_len, border_fmt)

        # Pul ustunlariga (E, F, G) format berish
        worksheet.set_column('E:G', 18, money_fmt)
        writer.close()
        output.seek(0)

        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Ecoprom_Qarzlar_{request.user.username}.xlsx'
        return response

    # 7. Statistika va Paginatsiya
    stats = debts.aggregate(
        total_debt_sum=Sum('calculated_debt'),
        total_rev=Sum('total_price'),
        total_paid_sum=Sum('prepayment')
    )
    
    paginator = Paginator(debts, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'page_obj': page_obj,
        'total_debt': stats['total_debt_sum'] or 0,
        'total_orders_count': debts.count(),
        'total_revenue': stats['total_rev'] or 0,
        'total_paid': stats['total_paid_sum'] or 0,
        'debt_customers_count': debts.values('customer_unique_id').distinct().count(),
        'search_query': search_query,
        'sort_by': sort_by,
        'is_admin': is_admin,
    }
    return render(request, 'orders/debt_report.html', context)

def add_prepayment(request, order_id):
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        try:
            amount_str = request.POST.get('amount', '0').replace(',', '.') # Vergulni nuqtaga almashtirish
            amount = float(amount_str)
            
            if amount <= 0:
                messages.error(request, "To'lov summasi 0 dan katta bo'lishi kerak.")
            else:
                # Qarzdan ko'p to'lov kiritilayotganini tekshirish (ixtiyoriy)
                remaining = float(order.total_price) - float(order.prepayment or 0)
                if amount > remaining:
                    messages.warning(request, f"E'tibor bering: Kiritilgan summa qarzdan ({remaining}) ko'proq.")

                # Yangi zalog summasini hisoblash
                order.prepayment = float(order.prepayment or 0) + amount
                order.save()
                
                messages.success(request, f"{order.customer_name} uchun {amount} qo'shildi. Umumiy to'langan: {order.prepayment}")
        
        except ValueError:
            messages.error(request, "Xato: Noto'g'ri raqam kiritildi.")
            
    return redirect('debt_report')


from django.db.models import Sum, Count
from django.shortcuts import render
from .models import Order
from django.db.models import Sum, Count, Max

from django.db.models import Sum, Count, Max, F

from django.shortcuts import render
from django.db.models import Sum, Count, Max, F
from django.http import JsonResponse
from .models import Order

from django.shortcuts import render
from django.db.models import Sum, Count, Max, F, Value, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from .models import Order

from django.shortcuts import render
from django.db.models import Sum, Count, Max, F, Value, DecimalField, ExpressionWrapper, Q
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from .models import Order

from django.shortcuts import render
from django.db.models import Sum, Count, Max, F, Value, DecimalField, ExpressionWrapper, Q
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from .models import Order

from django.db.models import Q, Sum, Count, Max, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import render
from .models import Order

from django.db.models import (
    F, Sum, Count, Max, Avg, Q, 
    Value, DecimalField, Case, When, FloatField
)
from django.db.models.functions import Coalesce, Round, TruncMonth
from django.http import JsonResponse
import json
from datetime import datetime, timedelta
from decimal import Decimal

from django.db.models import (
    F, Sum, Count, Max, Min, Avg, Q, 
    Value, DecimalField, Case, When, FloatField, 
    CharField, ExpressionWrapper, DurationField
)
import json
from datetime import datetime, timedelta
from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import (
    F, Sum, Count, Max, Min, Avg, Q, 
    Value, DecimalField, Case, When, FloatField, 
    CharField, ExpressionWrapper, DurationField
)
from django.db.models.functions import Coalesce, TruncMonth
from .models import Order

import json
from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import (
    Sum, Count, Avg, Max, Min, F, Q, 
    Case, When, Value, FloatField, DecimalField, ExpressionWrapper, CharField
)
from django.db.models.functions import Coalesce
from .models import Order
from django.db.models import DecimalField, Value, Sum, Q
from django.db.models.functions import Coalesce
def customer_rating(request):
    """
    To'liq biznes analitika: Mijozlar reytingi, Mahsulotlar tahlili, 
    PIR panellar, Eshiklar va Panel qalinligi statistikasi.
    """
    
    # ======================== 1. AJAX SO'ROVLAR ========================
    customer_id = request.GET.get('get_orders')
    if customer_id:
        orders = Order.objects.filter(
            customer_unique_id=customer_id, 
            parent_order__isnull=True
        ).order_by('-created_at')
        
        orders_list = [{
            'order_number': o.order_number,
            'product_name': o.product_name or "Eshik/Mebel",
            'panel_kvadrat': float(o.panel_kvadrat or 0),
            'status': o.status,
            'total_price': float(o.total_price or 0),
            'prepayment': float(o.prepayment or 0),
            'created_at': o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else '',
        } for o in orders]
        
        customer_stats = orders.aggregate(
            total_orders=Count('id'),
            total_amount=Coalesce(Sum('total_price'), Value(0, output_field=DecimalField())),
            total_paid=Coalesce(Sum('prepayment'), Value(0, output_field=DecimalField())),
            total_area=Coalesce(Sum('panel_kvadrat'), Value(0, output_field=DecimalField())),
            avg_order_value=Coalesce(Avg('total_price'), Value(0, output_field=DecimalField()))
        )
        
        return JsonResponse({
            'orders': orders_list,
            'stats': customer_stats,
            'customer_id': customer_id
        })

    # ======================== 2. MIJOZLAR REYTINGI ========================
    ratings_query = Order.objects.filter(parent_order__isnull=True).values('customer_unique_id').annotate(
        display_name=Max('customer_name'),
        order_count=Count('id'),
        first_order_date=Min('created_at'),
        last_order_date=Max('created_at'),
        total_m2=Coalesce(Sum('panel_kvadrat'), Value(0, output_field=DecimalField())),
        total_billed=Coalesce(Sum('total_price'), Value(0, output_field=DecimalField())),
        total_paid=Coalesce(Sum('prepayment'), Value(0, output_field=DecimalField())),
    ).annotate(
        payment_ratio=Case(
            When(total_billed__gt=0, then=100.0 * F('total_paid') / F('total_billed')),
            default=Value(0.0),
            output_field=FloatField()
        ),
        avg_order_value=ExpressionWrapper(F('total_billed') / F('order_count'), output_field=DecimalField())
    ).annotate(
        loyalty_score=Case(
            When(Q(order_count__gt=5) & Q(payment_ratio__gt=80), then=Value('A')),
            When(Q(order_count__gt=2) & Q(payment_ratio__gt=60), then=Value('B')),
            default=Value('C'),
            output_field=CharField()
        )
    )

    m2_ratings = list(ratings_query.order_by('-total_m2')[:15])
    sum_ratings = list(ratings_query.order_by('-total_billed')[:15])  # total_paid emas, total_billed
    order_count_ratings = list(ratings_query.order_by('-order_count')[:10])
    loyal_customers = list(ratings_query.filter(loyalty_score='A').order_by('-total_billed')[:10])

    # ======================== 3. UMUMIY STATISTIKA ========================
    base_aggregate = Order.objects.filter(parent_order__isnull=True).aggregate(
        total_orders=Count('id'),
        total_customers=Count('customer_unique_id', distinct=True),
        total_revenue=Coalesce(Sum('total_price'), Value(0, output_field=DecimalField())),
        total_prepayment=Coalesce(Sum('prepayment'), Value(0, output_field=DecimalField())),
        total_area=Coalesce(Sum('panel_kvadrat'), Value(0, output_field=DecimalField())),
        avg_order_value=Coalesce(Avg('total_price'), Value(0, output_field=DecimalField())),
    )

    overall_stats = base_aggregate
    if overall_stats['total_revenue'] > 0:
        overall_stats['avg_prepayment_ratio'] = (float(overall_stats['total_prepayment']) * 100) / float(overall_stats['total_revenue'])
    else:
        overall_stats['avg_prepayment_ratio'] = 0

    # TO'G'RILANGAN QISM - status nomlarini tekshiring
    completed_orders = Order.objects.filter(
        parent_order__isnull=True, 
        status__in=['BAJARILDI', 'TAYYOR']  # 'completed' emas, 'delivered' emas
    ).count()
    overall_stats['completion_rate'] = (completed_orders * 100 / overall_stats['total_orders']) if overall_stats['total_orders'] > 0 else 0

    # ======================== 4. PANEL QALINLIGI STATISTIKASI ========================
    thickness_stat = Order.objects.filter(
        parent_order__isnull=True
    ).exclude(
        Q(panel_thickness__isnull=True) | Q(panel_thickness='')
    ).values('panel_thickness').annotate(
        count=Count('id'),
        total_area=Coalesce(Sum('panel_kvadrat'), Value(0, output_field=DecimalField())),
        eshik_types=Count('product_name', distinct=True)
    ).order_by('panel_thickness')

    # ======================== 5. PIR PANELLAR TAHLILI ========================
    pir_all = Order.objects.filter(Q(panel_type__icontains='PIR') | Q(product_name__icontains='PIR'))
    
    pir_stats = pir_all.values('panel_type').annotate(
        count=Count('id'),
        total_m2=Coalesce(Sum('panel_kvadrat'), Value(0, output_field=DecimalField())),
        total_revenue=Coalesce(Sum('total_price'), Value(0, output_field=DecimalField()))
    ).order_by('-total_m2')

    pir_details = {
    'total_pir': pir_all.count(),
    'total_area': pir_all.aggregate(
        s=Coalesce(
            Sum('panel_kvadrat'), 
            Value(0, output_field=DecimalField()) # Mana bu yerda output_field qo'shildi
        )
    )['s'],
    'tom_panels': pir_all.filter(Q(panel_subtype__icontains='TOM') | Q(product_name__icontains='TOM')).count(),
    'secret_panels': pir_all.filter(Q(panel_subtype__icontains='SECRET') | Q(product_name__icontains='SECRET')).count(),
    'sovut_panels': pir_all.filter(Q(panel_subtype__icontains='SOVUT') | Q(product_name__icontains='SOVUT')).count(),
}

    # ======================== 6. ESHIKLAR TAHLILI ========================
    eshik_stat = Order.objects.filter(parent_order__isnull=True).exclude(
        Q(eshik_turi__isnull=True) | Q(eshik_turi='')
    ).values('eshik_turi').annotate(
        eshik_soni=Count('id'),
        total_revenue=Coalesce(Sum('total_price'), Value(0, output_field=DecimalField()))
    ).order_by('-eshik_soni')

    # ======================== 7. MASHHUR MAHSULOTLAR ========================
    product_rankings = Order.objects.filter(parent_order__isnull=True).values('product_name').annotate(
        order_count=Count('id'),
        total_m2=Coalesce(Sum('panel_kvadrat'), Value(0, output_field=DecimalField())),
        total_revenue=Coalesce(Sum('total_price'), Value(0, output_field=DecimalField()))
    ).order_by('-order_count')[:15]

    product_rankings_list = list(product_rankings)
    if product_rankings_list:
        max_orders = max(p['order_count'] for p in product_rankings_list)
        for p in product_rankings_list:
            p['popularity_score'] = (p['order_count'] * 100) / max_orders if max_orders > 0 else 0

    # ======================== 8. CONTEXT ========================
    context = {
        'm2_ratings': m2_ratings,
        'sum_ratings': sum_ratings,
        'order_count_ratings': order_count_ratings,
        'loyal_customers': loyal_customers,
        'overall_stats': overall_stats,
        'product_rankings': product_rankings_list,
        'thickness_stat': list(thickness_stat),
        'pir_stats': list(pir_stats),
        'pir_details': pir_details,
        'eshik_stat': list(eshik_stat),
        'json_data': {
            'm2_ratings': json.dumps(m2_ratings, default=str),
            'sum_ratings': json.dumps(sum_ratings, default=str),
        }
    }

    return render(request, 'orders/customer_rating.html', context)
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

@login_required
def get_customer_orders(request, customer_id):
    orders = Order.objects.filter(customer_unique_id=customer_id).values(
        'order_number', 'product_name', 'panel_kvadrat', 'status', 'created_at'
    ).order_by('-created_at')
    
    return JsonResponse({'orders': list(orders)})



from django.shortcuts import render
from django.http import JsonResponse
import json
from .models import Order, OrderItem # Modellaringiz nomi

def create_order_view(request):
    if request.method == "POST":
        try:
            # Frontenddan kelgan JSON ma'lumotni yuklaymiz
            data = json.loads(request.body)
            items = data.get('items', [])
            customer_name = data.get('customer') # Agar mijoz ismi ham yuborilsa
            
            # 1. Asosiy buyurtmani yaratamiz
            order = Order.objects.create(
                customer_name=customer_name,
                # boshqa kerakli maydonlar
            )
            
            # 2. Jadvalning har bir qatorini bazaga saqlaymiz
            for item in items:
                OrderItem.objects.create(
                    order=order,
                    product_name=item.get('name'),      # Stevnoi, Krovelnie yoki Eshik
                    product_type=item.get('sub_type'),  # F1..F8 yoki Qalinlik (80mm)
                    length=item.get('length') or 0,
                    quantity=item.get('count') or 0,
                    area=item.get('area') or 0,
                    price=item.get('price') or 0,
                    total_sum=item.get('total') or 0
                )
            
            return JsonResponse({"status": "success", "message": "Ma'lumotlar muvaffaqiyatli saqlandi!"})
        
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    # Agar GET so'rovi bo'lsa, sahifani o'zini qaytaramiz
    return render(request, 'orders/create_order.html')
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt # Osonlik uchun, lekin haqiqiy loyihada CSRF token yuborgan ma'qul
def save_order_ajax(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            items = data.get('items', [])
            
            # Asosiy buyurtmani yaratish
            order = Order.objects.create(customer_name=data.get('customer', 'Noma\'lum'))
            
            # Har bir qatorni saqlash
            for item in items:
                OrderItem.objects.create(
                    order=order,
                    product_name=item['name'],
                    product_type=item['sub_type'],
                    length=float(item['length'] or 0),
                    quantity=int(item['count'] or 0),
                    area=float(item['area'] or 0),
                    price=float(item['price'] or 0),
                    total_sum=float(item['total'] or 0)
                )
            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny # Mini App uchun vaqtincha
from .models import Order

class MyOrdersAPIView(APIView):
    # Dastlab tekshirish oson bo'lishi uchun AllowAny qilamiz
    # Keyinchalik xavfsizlikni kuchaytiramiz
    permission_classes = [AllowAny] 

    def get(self, request):
        # Bot orqali keladigan unique_id (Masalan: 656)
        customer_id = request.query_params.get('customer_id') 
        
        if not customer_id:
            return Response({"error": "ID ko'rsatilmadi"}, status=400)
            
        orders = Order.objects.filter(customer_unique_id=customer_id).order_by('-created_at')
        
        data = [
            {
                "order_number": o.order_number,
                "product_name": o.product_name or "Panel/Eshik",
                "status": o.get_status_display(),
                "total_price": float(o.total_price),
                "remaining": float(o.remaining_amount),
                "created_at": o.created_at.strftime('%d.%m.%Y %H:%M'),
            }
            for o in orders
        ]
        return Response(data)
# orders/views.py - boshida import qismiga qo'shing

from decimal import Decimal, ROUND_HALF_UP  # <-- BU NI QO'SHING

from decimal import Decimal, ROUND_HALF_UP


def order_calculator_list(request):
    # Faqat asosiy panellarni olamiz (parent buyurtmalar)
    orders = Order.objects.filter(parent_order__isnull=True).order_by('-id')

    # Qalinlik bo'yicha koeffitsientlar lug'ati
    SIRYO_COEFFICIENTS = {
        '5': Decimal('2'),
        '8': Decimal('3'),
        '10': Decimal('4'),
        '15': Decimal('6'),
    }

    DEFAULT_HEIGHT = Decimal('3.00')  # Balandlik kiritilmagan bo'lsa ishlatiladigan standart qiymat

    calculated_data = []
    total_kvadrat_all = Decimal('0')
    total_siryo_all = Decimal('0')
    total_zamok_all = 0
    total_stakanchik_all = 0

    for order in orders:
        kv = Decimal(str(order.panel_kvadrat or 0))
        # Qalinlikni string ko'rinishida olamiz (masalan: "10")
        thickness = str(order.panel_thickness or "10").strip()

        # Panel balandligi (height) — kiritilmagan bo'lsa standart qiymat olinadi
        balandlik = Decimal(str(order.panel_balandligi)) if order.panel_balandligi else DEFAULT_HEIGHT
        if balandlik <= 0:
            balandlik = DEFAULT_HEIGHT

        # Siryo hisobi: agar lug'atda qalinlik bo'lsa o'shani, bo'lmasa 10cm koeffitsientini oladi
        coeff = SIRYO_COEFFICIENTS.get(thickness, Decimal('4'))
        siryo_val = kv * coeff

        # List hisobi
        list_val = kv * Decimal('2')  # Har bir qatorda toza kvadrat * 2

        # Panel uzunligi (boyi) — kvadrat / balandlik orqali topiladi
        boyi_val = (kv / balandlik) if kv > 0 else Decimal('0')
        boyi_mm = boyi_val * 1000

        # ZAMOK HISOBI — endi balandlikka (uzunlikka) qarab, kvadraturaga emas
        eni_zamok = 4  # Eni bo'yicha doim 4 ta (boshida 2, oxirida 2)

        # Bo'yi bo'yicha: boshidan va oxiridan 700mm chegara tashlanadi,
        # qolgan qismda har 960mm da 1 tadan (2 tomon uchun)
        ishchi_boyi = boyi_mm - Decimal('1400')
        if ishchi_boyi > 0:
            boyi_zamok_soni = (int((ishchi_boyi / Decimal('960')).to_integral_value(rounding=ROUND_HALF_UP)) + 1) * 2
        else:
            boyi_zamok_soni = 0

        zamok_val = eni_zamok + boyi_zamok_soni

        # STAKANCHIK — har bir zamokka bittadan
        stakan_val = zamok_val

        calculated_data.append({
            'order': order,
            'thickness': thickness,
            'balandligi': balandlik,
            'siryo': siryo_val,
            'list': list_val,
            'zamok': zamok_val,
            'stakanchik': stakan_val,
            'boyi': boyi_val,
        })

        # Jami summalar uchun yig'ib boramiz
        total_kvadrat_all += kv
        total_siryo_all += siryo_val
        total_zamok_all += zamok_val
        total_stakanchik_all += stakan_val

    # JAMI LIST: (Jami Kvadrat * 2) + 10 metr zapas
    total_list_final = (total_kvadrat_all * Decimal('2')) + Decimal('10') if total_kvadrat_all > 0 else 0

    context = {
        'data': calculated_data,
        'total_kvadrat_all': total_kvadrat_all,
        'total_list_all': total_list_final,
        'total_siryo_all': total_siryo_all,
        'total_zamok_all': total_zamok_all,
        'total_stakanchik_all': total_stakanchik_all,
    }

    return render(request, 'orders/calculator.html', context)

from django.shortcuts import render
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from .models import Order, Material, MaterialTransaction

def director_dashboard(request):
    now = timezone.now()
    today = now.date()
    
    # ======================== 1. ORDER STATISTIKASI ========================
    total_orders = Order.objects.count()
    
    # Tugatilgan orderlar (BAJARILDI, TAYYOR, USTA_TUGATDI)
    completed_orders = Order.objects.filter(
        status__in=['BAJARILDI', 'TAYYOR', 'USTA_TUGATDI']
    ).count()
    
    # Jarayondagilar (KIRITILDI, TASDIQLANDI, USTA_QABUL_QILDI, USTA_BOSHLA, ISHDA)
    in_progress = Order.objects.filter(
        status__in=['KIRITILDI', 'TASDIQLANDI', 'USTA_QABUL_QILDI', 'USTA_BOSHLA', 'ISHDA']
    ).count()
    
    # Muddati o'tganlar (deadline < hozir va tugatilmagan)
    overdue_orders = Order.objects.filter(
        deadline__lt=now
    ).exclude(
        status__in=['BAJARILDI', 'TAYYOR', 'USTA_TUGATDI', 'RAD_ETILDI']
    ).count()
    
    # Yangi orderlar (oxirgi 7 kun)
    week_ago = now - timedelta(days=7)
    new_orders = Order.objects.filter(created_at__gte=week_ago).count()
    
    # Protsentlar
    completion_rate = round(completed_orders / total_orders * 100, 1) if total_orders else 0
    overdue_rate = round(overdue_orders / total_orders * 100, 1) if total_orders else 0
    in_progress_rate = round(in_progress / total_orders * 100, 1) if total_orders else 0
    
    # ======================== 2. OMBORXONA HOLATI ========================
    total_materials = Material.objects.count()
    
    # Jami ombor qiymati
    warehouse_value = Material.objects.aggregate(
        total=Sum(F('quantity') * F('price_per_unit'))
    )['total'] or Decimal(0)
    
    # Kategoriya bo'yicha (mavjud bo'lsa)
    try:
        from .models import Category
        raw_materials = Material.objects.filter(category__name__icontains='xom').count()
        semi_finished = Material.objects.filter(category__name__icontains='yarim').count()
        finished = Material.objects.filter(category__name__icontains='tayyor').count()
    except:
        raw_materials = total_materials
        semi_finished = 0
        finished = 0
    
    # Kam qolgan materiallar
    low_stock_materials = Material.objects.filter(quantity__lte=F('min_stock_level')).count()
    
    # ======================== 3. UMUMIY TUSHUM ========================
    # Oy boshidan
    monthly_revenue = Order.objects.filter(
        created_at__month=now.month,
        created_at__year=now.year,
        status__in=['BAJARILDI', 'TAYYOR']
    ).aggregate(total=Sum('total_price'))['total'] or Decimal(0)
    
    # Yil boshidan
    yearly_revenue = Order.objects.filter(
        created_at__year=now.year,
        status__in=['BAJARILDI', 'TAYYOR']
    ).aggregate(total=Sum('total_price'))['total'] or Decimal(0)
    
    # So'nggi 30 kun
    last_30_days = now - timedelta(days=30)
    last_30_revenue = Order.objects.filter(
        created_at__gte=last_30_days,
        status__in=['BAJARILDI', 'TAYYOR']
    ).aggregate(total=Sum('total_price'))['total'] or Decimal(0)
    
    # ======================== 4. QARZDORLAR ========================
    debt_orders = Order.objects.filter(
        total_price__gt=F('prepayment')
    ).exclude(status__in=['RAD_ETILDI', 'BEKOR_QILINDI'])
    
    total_debt = sum(order.total_price - order.prepayment for order in debt_orders)
    debt_count = debt_orders.count()
    
    # ======================== 5. OYLIK GRAFIK ========================
    monthly_data = []
    for i in range(5, -1, -1):
        month_date = now - timedelta(days=30*i)
        revenue = Order.objects.filter(
            created_at__month=month_date.month,
            created_at__year=month_date.year,
            status__in=['BAJARILDI', 'TAYYOR']
        ).aggregate(total=Sum('total_price'))['total'] or Decimal(0)
        monthly_data.append({
            'month': month_date.strftime('%b'),
            'revenue': float(revenue) / 1_000_000
        })
    
    # ======================== 6. SO'NGI HARAKATLAR ========================
    recent_transactions = MaterialTransaction.objects.select_related(
        'material', 'performed_by'
    ).order_by('-timestamp')[:10]
    
    # ======================== 7. USTALAR STATISTIKASI ========================
    from .models import Worker
    workers_count = Worker.objects.count()
    
    # ======================== CONTEXT ========================
    context = {
        # Order ma'lumotlari
        'total_orders': total_orders,
        'overdue_orders': overdue_orders,
        'in_progress': in_progress,
        'new_orders': new_orders,
        'completed_orders': completed_orders,
        'completion_rate': completion_rate,
        'overdue_rate': overdue_rate,
        'in_progress_rate': in_progress_rate,
        
        # Ombor ma'lumotlari
        'total_products': total_materials,
        'warehouse_value': warehouse_value,
        'raw_materials': raw_materials,
        'semi_finished': semi_finished,
        'finished': finished,
        'low_stock_materials': low_stock_materials,
        
        # Tushum ma'lumotlari
        'monthly_revenue': monthly_revenue,
        'yearly_revenue': yearly_revenue,
        'last_30_revenue': last_30_revenue,
        'monthly_data': monthly_data,
        
        # Qarz ma'lumotlari
        'total_debt': total_debt,
        'debt_count': debt_count,
        
        # So'nggi harakatlar
        'recent_transactions': recent_transactions,
        
        # Ustalar soni
        'workers_count': workers_count,
        
        # Joriy vaqt
        'now': now,
    }
    
    return render(request, 'orders/director_dashboard.html', context)

import pandas as pd
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import Material, Category
import json
import traceback
import logging

# Logger yaratish
logger = logging.getLogger(__name__)
import pandas as pd
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import Material, Category
import logging
import traceback

logger = logging.getLogger(__name__)

@login_required
@csrf_exempt
def import_excel_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST mumkin'}, status=405)

    if 'excel_file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'Fayl yuklanmadi'}, status=400)

    try:
        excel_file = request.FILES['excel_file']
        
        # MUHIM: Faylingizda 1-qator sarlavha emas, shuning uchun skiprows=[0] yoki header=1 qilamiz
        # Sarlavhalar 2-qatorda (Index 1) joylashgan
        df = pd.read_excel(excel_file, engine='openpyxl', header=1) 
        
        # Ustun nomlaridagi bo'sh joylarni tozalash
        df.columns = [str(c).strip() for c in df.columns]
        
        logger.info(f"Topilgan ustunlar: {df.columns.tolist()}")

        # Agar 'Наименование' ustuni bo'lmasa, sarlavha noto'g'ri o'qilgan
        if 'Наименование' not in df.columns:
            return JsonResponse({
                'success': False, 
                'message': f'Faylda "Наименование" ustuni topilmadi. Mavjud ustunlar: {", ".join(df.columns)}'
            })

        created_count = 0
        updated_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                name = row.get('Наименование')
                
                # Bo'sh qatorlarni tashlab ketish
                if pd.isna(name) or str(name).strip().lower() in ['nan', '']:
                    continue
                
                name = str(name).strip()

                # Miqdorni aniqlash: Excelingizda 'Количество (шт.)' ustuni bor
                quantity = 0
                qty_val = row.get('Количество (шт.)')
                
                # Agar 'Количество (шт.)' bo'sh bo'lsa, 'Масса (кг)' ni tekshirish
                if pd.isna(qty_val) or qty_val == '':
                    qty_val = row.get('Масса (кг)', 0)

                try:
                    quantity = float(qty_val) if pd.notna(qty_val) else 0
                except:
                    quantity = 0

                # Kategoriya aniqlash
                category_name = detect_category(name)
                category, _ = Category.objects.get_or_create(name=category_name)

                # O'lchov birligini aniqlash
                unit = 'dona'
                if 'Масса (кг)' in row and pd.notna(row['Масса (кг)']) and not pd.notna(row['Количество (шт.)']):
                    unit = 'kg'
                elif any(x in name.lower() for x in ['болт', 'винт', 'припой', 'масло']):
                    unit = 'kg'

                # Materialni saqlash
                material, created = Material.objects.update_or_create(
                    name=name,
                    defaults={
                        'category': category,
                        'unit': unit,
                        'min_stock_level': 0
                    }
                )
                
                if created:
                    material.quantity = quantity
                    created_count += 1
                else:
                    material.quantity += quantity
                    updated_count += 1
                
                material.save()

            except Exception as row_e:
                errors.append(f"Qator {index + 3}: {str(row_e)}")

        return JsonResponse({
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:10]
        })

    except Exception as e:
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f"Xato: {str(e)}"}, status=500)

def detect_category(name):
    """Material nomiga qarab kategoriyani aniqlash"""
    name_lower = name.lower()
    
    if 'трв' in name_lower:
        return 'ТРВ'
    elif 'отвод' in name_lower:
        return 'Отводы'
    elif 'медная труб' in name_lower:
        return 'Медные трубы'
    elif 'фреон' in name_lower:
        return 'Фреоны'
    elif 'щит' in name_lower:
        return 'Щиты'
    elif 'фильтр' in name_lower:
        return 'Фильтры'
    elif 'вибрашланг' in name_lower:
        return 'Вибрашланги'
    elif 'магнитный клапан' in name_lower:
        return 'Магнитные клапаны'
    elif 'манометр' in name_lower:
        return 'Манометры'
    elif 'реле' in name_lower:
        return 'Реле'
    elif 'термо' in name_lower and 'контроллер' in name_lower:
        return 'Термоконтроллеры'
    elif 'масло' in name_lower:
        return 'Масла'
    elif 'конденсатор' in name_lower:
        return 'Конденсаторы'
    elif 'глазок' in name_lower:
        return 'Глазки'
    elif 'сепаратор' in name_lower or 'жидкостный отделитель' in name_lower:
        return 'Сепараторы'
    elif 'вентилятор' in name_lower:
        return 'Вентиляторы'
    elif 'болт' in name_lower or 'винт' in name_lower:
        return 'Крепеж'
    elif 'припой' in name_lower:
        return 'Припои'
    elif 'компрессор' in name_lower:
        return 'Компрессоры'
    elif 'заклепка' in name_lower:
        return 'Заклепки'
    elif 'скотч' in name_lower:
        return 'Скотчи'
    elif 'пена' in name_lower:
        return 'Пена'
    elif 'стакан' in name_lower:
        return 'Стаканы'
    elif 'стрейч' in name_lower or 'ekoprom' in name_lower:
        return 'Стрейч-пленка'
    elif name.startswith('F1') or name.startswith('F5') or name.startswith('F8') or name.startswith('F9'):
        return 'Дверные комплектующие'
    elif 'сверло' in name_lower or 'метчик' in name_lower:
        return 'Инструменты'
    elif 'игла' in name_lower:
        return 'Иглы'
    elif 'штуцер' in name_lower:
        return 'Штуцеры'
    elif 'краник' in name_lower:
        return 'Краники'
    else:
        return 'Разное'
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from .models import Order
import json

@login_required
def chamber_drawing(request, pk):
    """Texnik chizma sahifasi"""
    order = get_object_or_404(Order, pk=pk)
    
    # Spetsifikatsiyani olish
    spec = order.technical_spec_json or {}
    
    context = {
        'order': order,
        'spec': spec,
    }
    return render(request, 'orders/chamber_drawing.html', context)


@login_required
def download_drawing_svg(request, pk):
    """SVG faylni yuklash"""
    order = get_object_or_404(Order, pk=pk)
    
    if order.technical_drawing_svg:
        response = HttpResponse(order.technical_drawing_svg.read(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'attachment; filename="{order.order_number}_drawing.svg"'
        return response
    
    return JsonResponse({'error': 'Fayl topilmadi'}, status=404)


@login_required
def download_drawing_pdf(request, pk):
    """PDF faylni yuklash"""
    import weasyprint
    from django.template.loader import render_to_string
    
    order = get_object_or_404(Order, pk=pk)
    
    html_string = render_to_string('orders/chamber_drawing_pdf.html', {
        'order': order,
        'spec': order.technical_spec_json or {}
    })
    
    pdf = weasyprint.HTML(string=html_string).write_pdf()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{order.order_number}_drawing.pdf"'
    return response


@login_required
def download_spec_json(request, pk):
    """JSON spetsifikatsiyani yuklash"""
    order = get_object_or_404(Order, pk=pk)
    
    response = HttpResponse(
        json.dumps(order.technical_spec_json, indent=2, ensure_ascii=False),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename="{order.order_number}_spec.json"'
    return response


@csrf_exempt
@login_required
def generate_drawing(request, pk):
    """Chizma yaratish"""
    from .signals import generate_chamber_drawing
    
    order = get_object_or_404(Order, pk=pk)
    
    try:
        result = generate_chamber_drawing(order)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
# constructor/views.py
import json
import math
import os
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Tuple

import requests
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import models
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.decorators.csrf import csrf_exempt

from .forms import ProjectForm, QuickCalculatorForm
from .models import Project

# =========================================================
# YORDAMCHI FUNKSIYALAR
# =========================================================

def mm_val(s: str) -> int:
    """'100mm' -> 100"""
    return int(str(s).replace("mm", "").strip())


def m_to_mm(m: float) -> int:
    """Metrni millimetrga o'tkazish"""
    return int(round(m * 1000))


def door_dimensions(door_type: str) -> Tuple[int, int]:
    """Eshik o'lchamlarini qaytaradi (en, bo'y) mm da"""
    door_map = {
        "Bir tabaqali (90x190)": (900, 1900),
        "Surilma (120x200)": (1200, 2000),
        "Muzlatkich eshigi": (960, 2000),
        "Yo'q": (0, 0),
    }
    return door_map.get(door_type, (0, 0))


def panel_count_linear(length_m: float, panel_width_m: float = 1.16) -> Dict:
    """Berilgan uzunlik uchun panel sonini hisoblash"""
    full = int(length_m // panel_width_m)
    rem = round(length_m - (full * panel_width_m), 3)
    total = full + (1 if rem > 0.01 else 0)
    return {
        "full_panels": full,
        "remainder_m": rem,
        "total_panels": total
    }


def split_center_by_960(center_mm: int, module_mm: int = 960) -> List[int]:
    """Markaziy qismni 960mm modullarga bo'lish"""
    if center_mm <= 0:
        return []
    
    parts = []
    remain = center_mm
    
    while remain > module_mm:
        next_remain = remain - module_mm
        if next_remain <= module_mm:
            parts.append(module_mm)
            remain = next_remain
            break
        parts.append(module_mm)
        remain -= module_mm
    
    if remain > 0:
        parts.append(remain)
    
    return parts


def build_side_segments(total_mm: int, corner_mm: int = 480, module_mm: int = 960) -> List[int]:
    """Devor tomonini segmentlarga bo'lish (burchak 480mm + markaziy modullar)"""
    if total_mm <= 0:
        return []
    if total_mm <= corner_mm * 2:
        return [total_mm]
    
    center_mm = total_mm - (corner_mm * 2)
    center_parts = split_center_by_960(center_mm, module_mm)
    return [corner_mm] + center_parts + [corner_mm]


def segment_meta(parts: List[int], has_door: bool = False, door_size: int = 960) -> List[Dict]:
    """Segmentlarni tiplar bilan qaytarish (panel yoki eshik)"""
    result = []
    door_used = False
    
    for p in parts:
        if has_door and (not door_used) and p == door_size:
            result.append({"size": p, "type": "door"})
            door_used = True
        else:
            result.append({"size": p, "type": "panel"})
    
    return result


def calculate_all(project) -> Dict:
    """Loyiha bo'yicha barcha hisob-kitoblarni bajarish"""
    
    L = float(project.length_m)
    W = float(project.width_m)
    H = float(project.height_m)
    
    wall_mm = mm_val(project.wall_thickness)
    ceil_mm = mm_val(project.ceiling_thickness)
    floor_mm = mm_val(project.floor_thickness) if project.has_floor else 0
    
    panel_width = float(project.panel_width)
    door_w_mm, door_h_mm = door_dimensions(project.door_type)
    
    # Hajmlar
    hajm = round(L * W * H, 2)
    inner_L_mm = max(0, m_to_mm(L) - (2 * wall_mm))
    inner_W_mm = max(0, m_to_mm(W) - (2 * wall_mm))
    inner_H_mm = max(0, m_to_mm(H) - ceil_mm - floor_mm)
    inner_hajm = round((inner_L_mm * inner_W_mm * inner_H_mm) / 1_000_000_000, 2)
    
    # Maydonlar
    s_devor = round(2 * (L + W) * H, 2)
    s_patalok = round(L * W, 2)
    s_pol = round(L * W, 2) if project.has_floor else 0
    total_panel_area = round(s_devor + s_patalok + s_pol, 2)
    
    # Panel sonlari
    wall_layout_L = panel_count_linear(L, panel_width)
    wall_layout_W = panel_count_linear(W, panel_width)
    
    devor_panels_total = (wall_layout_L["total_panels"] * 2) + (wall_layout_W["total_panels"] * 2)
    patalok_panels_total = math.ceil(W / panel_width)
    pol_panels_total = math.ceil(W / panel_width) if project.has_floor else 0
    estimated_all_panels = devor_panels_total + patalok_panels_total + pol_panels_total
    
    # Segmentlar
    top_parts = build_side_segments(m_to_mm(L))
    right_parts = build_side_segments(m_to_mm(W))
    
    has_door_top = project.door_type != "Yo'q" and project.door_side in ["Old", "Orqa"]
    has_door_right = project.door_type != "Yo'q" and project.door_side in ["Chap", "O'ng"]
    
    top_meta = segment_meta(top_parts, has_door=has_door_top, door_size=door_w_mm)
    right_meta = segment_meta(right_parts, has_door=has_door_right, door_size=door_w_mm)
    
    # Eshik offset
    door_offset_mm = 0
    if project.door_type != "Yo'q":
        if project.door_side in ["Chap", "O'ng"]:
            door_offset_mm = get_door_offset(right_parts, project.door_position, "vertical", door_h_mm)
        else:
            door_offset_mm = get_door_offset(top_parts, project.door_position, "horizontal", door_w_mm)
    
    return {
        "outer_volume_m3": hajm,
        "inner_volume_m3": inner_hajm,
        "inner_L_mm": inner_L_mm,
        "inner_W_mm": inner_W_mm,
        "inner_H_mm": inner_H_mm,
        "wall_area_m2": s_devor,
        "ceiling_area_m2": s_patalok,
        "floor_area_m2": s_pol,
        "total_panel_area_m2": total_panel_area,
        "wall_panels_total": devor_panels_total,
        "ceiling_panels_total": patalok_panels_total,
        "floor_panels_total": pol_panels_total,
        "total_panels_estimated": estimated_all_panels,
        "top_segments": top_parts,
        "right_segments": right_parts,
        "top_meta": top_meta,
        "right_meta": right_meta,
        "wall_mm": wall_mm,
        "ceil_mm": ceil_mm,
        "floor_mm": floor_mm,
        "door_w_mm": door_w_mm,
        "door_h_mm": door_h_mm,
        "door_offset_mm": door_offset_mm,
        "wall_layout_L": wall_layout_L,
        "wall_layout_W": wall_layout_W,
    }


def get_door_offset(parts: List[int], position: str, side_type: str, door_size_mm: int) -> int:
    """Eshikning aniq joylashuv offsetini hisoblash (mm)"""
    total = sum(parts)
    if total <= 0:
        return 0
    
    if side_type == "vertical":  # Chap/O'ng tomon
        if position == "Tepa":
            offset = 480
        elif position == "Past":
            offset = total - 480 - door_size_mm
        else:  # O'rta
            offset = (total - door_size_mm) / 2
    else:  # Old/Orqa tomon (gorizontal)
        if position == "Chap":
            offset = 480
        elif position == "O'ng":
            offset = total - 480 - door_size_mm
        else:  # O'rta
            offset = (total - door_size_mm) / 2
    
    return int(max(0, min(offset, total - door_size_mm)))


def generate_svg(project, calculations):
    """Loyiha uchun texnik chizma SVG yaratish"""
    
    L = float(project.length_m)
    W = float(project.width_m)
    H = float(project.height_m)
    
    wall_mm = calculations['wall_mm']
    ceil_mm = calculations['ceil_mm']
    floor_mm = calculations['floor_mm']
    
    outer_w_mm = m_to_mm(L)
    outer_h_mm = m_to_mm(W)
    outer_z_mm = m_to_mm(H)
    
    top_meta = calculations['top_meta']
    right_meta = calculations['right_meta']
    
    door_w_mm = calculations['door_w_mm']
    door_h_mm = calculations['door_h_mm']
    door_offset_mm = calculations['door_offset_mm']
    
    # Masshtab
    max_draw_w = 250
    max_draw_h = 185
    scale = min(max_draw_w / outer_w_mm, max_draw_h / outer_h_mm)
    
    draw_w = outer_w_mm * scale
    draw_h = outer_h_mm * scale
    wall_t = max(6, wall_mm * scale)
    
    svg_w = 800
    svg_h = 600
    
    x_center = 400
    px = x_center - draw_w / 2
    py = 100
    
    inner_L_mm = calculations['inner_L_mm']
    inner_W_mm = calculations['inner_W_mm']
    inner_H_mm = calculations['inner_H_mm']
    
    # Ranglar
    c = {
        "sheet": "#ffffff",
        "line": "#111111",
        "dim": "#222222",
        "text": "#111111",
        "muted": "#555555",
        "door": "#111111",
    }
    
    date_str = datetime.now().strftime("%d.%m.%Y")
    
    # ========== SVG YIG'ISH ==========
    svg_parts = []
    
    # Header
    svg_parts.append(f'<svg width="100%" viewBox="0 0 {svg_w} {svg_h}" xmlns="http://www.w3.org/2000/svg">')
    svg_parts.append(f'<rect x="10" y="10" width="{svg_w-20}" height="{svg_h-20}" fill="{c["sheet"]}" stroke="none"/>')
    
    svg_parts.append(f'<text x="{x_center}" y="40" font-size="16" font-weight="bold" text-anchor="middle" fill="{c["text"]}">TEXNIK CHIZMA</text>')
    svg_parts.append(f'<text x="{x_center}" y="60" font-size="12" font-weight="bold" text-anchor="middle" fill="{c["text"]}">{(project.project_name or "").upper()}</text>')
    svg_parts.append(f'<text x="{svg_w-20}" y="40" font-size="11" text-anchor="end" fill="{c["muted"]}">{project.room_code}</text>')
    
    # Tashqi va ichki to'rtburchak
    svg_parts.append(f'<rect x="{px}" y="{py}" width="{draw_w}" height="{draw_h}" fill="none" stroke="{c["line"]}" stroke-width="1.6"/>')
    svg_parts.append(f'<rect x="{px+wall_t}" y="{py+wall_t}" width="{draw_w-2*wall_t}" height="{draw_h-2*wall_t}" fill="none" stroke="{c["line"]}" stroke-width="1.1"/>')
    
    # Ichki o'lcham matni
    svg_parts.append(f'<text x="{px+draw_w/2+8}" y="{py+draw_h/2-8}" font-size="14" font-weight="bold" text-anchor="middle" fill="{c["text"]}" transform="rotate(90 {px+draw_w/2+8},{py+draw_h/2-8})">H-{outer_z_mm}</text>')
    svg_parts.append(f'<text x="{px+draw_w/2}" y="{py+draw_h+40}" font-size="10" text-anchor="middle" fill="{c["muted"]}">Ichki: {inner_L_mm} x {inner_W_mm} x {inner_H_mm} mm</text>')
    
    # Uzunlik o'lchami
    svg_parts.append(dim_h(px, px+draw_w, py-15, f"{outer_w_mm} mm", c["dim"]))
    
    # En o'lchami
    svg_parts.append(dim_v(px+draw_w+15, py, py+draw_h, f"{outer_h_mm} mm", c["dim"]))
    
    # Eshik (agar bor bo'lsa)
    if project.door_type != "Yo'q":
        door_svg = draw_door(project, px, py, draw_w, draw_h, scale, door_offset_mm, door_w_mm, door_h_mm, c)
        svg_parts.append(door_svg)
    
    # Segment chiziqlari
    svg_parts.append(chain_dim_top(px, py-6, top_meta, scale, c["dim"]))
    svg_parts.append(chain_dim_right(px+draw_w+8, py, right_meta, scale, c["dim"]))
    
    # Devor qalinligi
    svg_parts.append(f'<text x="{px+draw_w/2}" y="{py+draw_h+20}" font-size="11" text-anchor="middle" fill="{c["text"]}">Devor: {wall_mm} mm | Patalok: {ceil_mm} mm | Pol: {floor_mm if project.has_floor else 0} mm</text>')
    
    # Title block
    svg_parts.append(title_block(px-50, py+draw_h+60, draw_w+100, 80, project, wall_mm, ceil_mm, floor_mm, date_str, c))
    
    svg_parts.append(f'<text x="{svg_w-20}" y="{svg_h-20}" font-size="10" text-anchor="end" fill="{c["muted"]}">EcoProm Konstruktor</text>')
    svg_parts.append('</svg>')
    
    return '\n'.join(svg_parts)


def dim_h(x1, x2, y, text, color="#222"):
    """Gorizontal o'lcham chizig'i"""
    return f"""
    <g stroke="{color}" fill="none" stroke-width="1">
        <line x1="{x1}" y1="{y}" x2="{x2}" y2="{y}" />
        <line x1="{x1}" y1="{y-5}" x2="{x1}" y2="{y+5}" />
        <line x1="{x2}" y1="{y-5}" x2="{x2}" y2="{y+5}" />
    </g>
    <text x="{(x1+x2)/2}" y="{y-6}" font-size="10" text-anchor="middle" fill="{color}">{text}</text>
    """


def dim_v(x, y1, y2, text, color="#222"):
    """Vertikal o'lcham chizig'i"""
    return f"""
    <g stroke="{color}" fill="none" stroke-width="1">
        <line x1="{x}" y1="{y1}" x2="{x}" y2="{y2}" />
        <line x1="{x-5}" y1="{y1}" x2="{x+5}" y2="{y1}" />
        <line x1="{x-5}" y1="{y2}" x2="{x+5}" y2="{y2}" />
    </g>
    <text x="{x+12}" y="{(y1+y2)/2}" font-size="10" text-anchor="middle" fill="{color}" transform="rotate(90 {x+12},{(y1+y2)/2})">{text}</text>
    """


def chain_dim_top(x, y, parts, scale, color="#222"):
    """Yuqori segment o'lchamlari"""
    svg = ""
    cur = x
    for p in parts:
        nx = cur + p["size"] * scale
        label = f'{p["size"]} ESHIK' if p["type"] == "door" else str(p["size"])
        svg += f'<line x1="{cur}" y1="{y}" x2="{cur}" y2="{y-6}" stroke="{color}" stroke-width="1"/>'
        svg += f'<text x="{(cur+nx)/2}" y="{y-4}" font-size="9" text-anchor="middle" fill="{color}">{label}</text>'
        cur = nx
    svg += f'<line x1="{cur}" y1="{y}" x2="{cur}" y2="{y-6}" stroke="{color}" stroke-width="1"/>'
    return svg


def chain_dim_right(x, y, parts, scale, color="#222"):
    """O'ng segment o'lchamlari"""
    svg = ""
    cur = y
    for p in parts:
        ny = cur + p["size"] * scale
        label = f'{p["size"]} ESHIK' if p["type"] == "door" else str(p["size"])
        svg += f'<line x1="{x}" y1="{cur}" x2="{x+6}" y2="{cur}" stroke="{color}" stroke-width="1"/>'
        svg += f'<text x="{x+10}" y="{(cur+ny)/2}" font-size="9" text-anchor="middle" fill="{color}" transform="rotate(90 {x+10},{(cur+ny)/2})">{label}</text>'
        cur = ny
    svg += f'<line x1="{x}" y1="{cur}" x2="{x+6}" y2="{cur}" stroke="{color}" stroke-width="1"/>'
    return svg


def draw_door(project, px, py, draw_w, draw_h, scale, offset_mm, door_w_mm, door_h_mm, c):
    """Eshik chizish"""
    door_side = project.door_side
    opening = project.door_opening
    
    if door_side == "Chap":
        top = py + offset_mm * scale
        bot = top + door_h_mm * scale
        if opening == "Ichkariga":
            return f"""
            <line x1="{px}" y1="{top}" x2="{px+20}" y2="{top+30}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{px}" y1="{bot}" x2="{px+20}" y2="{bot-30}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{px+20}" y1="{top+30}" x2="{px+20}" y2="{bot-30}" stroke="{c["door"]}" stroke-width="1.5" stroke-dasharray="4,3"/>
            """
        else:
            return f"""
            <line x1="{px}" y1="{top}" x2="{px-20}" y2="{top-30}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{px}" y1="{bot}" x2="{px-20}" y2="{bot+30}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{px-20}" y1="{top-30}" x2="{px-20}" y2="{bot+30}" stroke="{c["door"]}" stroke-width="1.5" stroke-dasharray="4,3"/>
            """
    
    elif door_side == "O'ng":
        rx = px + draw_w
        top = py + offset_mm * scale
        bot = top + door_h_mm * scale
        if opening == "Ichkariga":
            return f"""
            <line x1="{rx}" y1="{top}" x2="{rx-20}" y2="{top+30}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{rx}" y1="{bot}" x2="{rx-20}" y2="{bot-30}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{rx-20}" y1="{top+30}" x2="{rx-20}" y2="{bot-30}" stroke="{c["door"]}" stroke-width="1.5" stroke-dasharray="4,3"/>
            """
        else:
            return f"""
            <line x1="{rx}" y1="{top}" x2="{rx+20}" y2="{top-30}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{rx}" y1="{bot}" x2="{rx+20}" y2="{bot+30}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{rx+20}" y1="{top-30}" x2="{rx+20}" y2="{bot+30}" stroke="{c["door"]}" stroke-width="1.5" stroke-dasharray="4,3"/>
            """
    
    elif door_side == "Old":
        left = px + offset_mm * scale
        right = left + door_w_mm * scale
        by = py + draw_h
        if opening == "Ichkariga":
            return f"""
            <line x1="{left}" y1="{by}" x2="{left+30}" y2="{by-20}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{right}" y1="{by}" x2="{right-30}" y2="{by-20}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{left+30}" y1="{by-20}" x2="{right-30}" y2="{by-20}" stroke="{c["door"]}" stroke-width="1.5" stroke-dasharray="4,3"/>
            """
        else:
            return f"""
            <line x1="{left}" y1="{by}" x2="{left-30}" y2="{by+20}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{right}" y1="{by}" x2="{right+30}" y2="{by+20}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{left-30}" y1="{by+20}" x2="{right+30}" y2="{by+20}" stroke="{c["door"]}" stroke-width="1.5" stroke-dasharray="4,3"/>
            """
    
    elif door_side == "Orqa":
        left = px + offset_mm * scale
        right = left + door_w_mm * scale
        if opening == "Ichkariga":
            return f"""
            <line x1="{left}" y1="{py}" x2="{left+30}" y2="{py+20}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{right}" y1="{py}" x2="{right-30}" y2="{py+20}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{left+30}" y1="{py+20}" x2="{right-30}" y2="{py+20}" stroke="{c["door"]}" stroke-width="1.5" stroke-dasharray="4,3"/>
            """
        else:
            return f"""
            <line x1="{left}" y1="{py}" x2="{left-30}" y2="{py-20}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{right}" y1="{py}" x2="{right+30}" y2="{py-20}" stroke="{c["door"]}" stroke-width="1.5"/>
            <line x1="{left-30}" y1="{py-20}" x2="{right+30}" y2="{py-20}" stroke="{c["door"]}" stroke-width="1.5" stroke-dasharray="4,3"/>
            """
    
    return ""


def title_block(x, y, w, h, project, wall_mm, ceil_mm, floor_mm, date_str, c):
    """Sarlavha bloki"""
    L_mm = m_to_mm(float(project.length_m))
    W_mm = m_to_mm(float(project.width_m))
    H_mm = m_to_mm(float(project.height_m))
    
    return f"""
    <g>
        <rect x="{x}" y="{y}" width="{w}" height="{h}" fill="none" stroke="{c["line"]}" stroke-width="1"/>
        <line x1="{x}" y1="{y+25}" x2="{x+w}" y2="{y+25}" stroke="{c["line"]}" stroke-width="1"/>
        
        <text x="{x+10}" y="{y+18}" font-size="11" font-weight="bold" fill="{c["text"]}">ECOPROM TECHNICAL DRAWING</text>
        
        <text x="{x+10}" y="{y+45}" font-size="10" fill="{c["muted"]}">Loyiha:</text>
        <text x="{x+60}" y="{y+45}" font-size="10" font-weight="bold" fill="{c["text"]}">{project.project_name or "-"}</text>
        
        <text x="{x+10}" y="{y+65}" font-size="10" fill="{c["muted"]}">O'lcham:</text>
        <text x="{x+60}" y="{y+65}" font-size="10" fill="{c["text"]}">{L_mm} x {W_mm} x {H_mm} mm</text>
        
        <text x="{x+250}" y="{y+45}" font-size="10" fill="{c["muted"]}">Kod:</text>
        <text x="{x+280}" y="{y+45}" font-size="10" font-weight="bold" fill="{c["text"]}">{project.room_code}</text>
        
        <text x="{x+250}" y="{y+65}" font-size="10" fill="{c["muted"]}">Devor/Patalok/Pol:</text>
        <text x="{x+360}" y="{y+65}" font-size="10" fill="{c["text"]}">{wall_mm}/{ceil_mm}/{floor_mm} mm</text>
        
        <text x="{x+w-10}" y="{y+45}" font-size="10" text-anchor="end" fill="{c["muted"]}">Sana: {date_str}</text>
        <text x="{x+w-10}" y="{y+65}" font-size="10" text-anchor="end" fill="{c["muted"]}">Sheet: 1/1</text>
    </g>
    """


def get_ai_recommendation(project):
    """Groq API orqali AI tavsiya olish"""
    api_key = getattr(settings, 'GROQ_API_KEY', os.getenv('GROQ_API_KEY', ''))
    
    if not api_key:
        return {"success": False, "message": "GROQ_API_KEY topilmadi"}
    
    prompt = f"""
Siz sovutish kamerasi bo'yicha professional muhandissiz.
Foydalanuvchiga texnik tavsiya bering.

Mijoz ma'lumotlari:
- Mahsulot turi: {project.product_type}
- Talab qilinadigan harorat: {project.storage_temp}
- Kunlik eshik ochilish soni: {project.opening_freq}
- Hudud / iqlim: {project.region}
- Namlik talabi: {project.humidity}
- O'lcham: {project.length_m}m x {project.width_m}m x {project.height_m}m
- Pol paneli: {"Ha" if project.has_floor else "Yo'q"}

JSON formatda qaytaring:
{{
  "rejim": "...",
  "devor_qalinligi_mm": 100,
  "patalok_qalinligi_mm": 80,
  "pol_qalinligi_mm": 100,
  "agregat_turi": "...",
  "eshik_turi": "...",
  "izoh": "...",
  "xulosa": "..."
}}

Faqat JSON qaytaring.
"""
    
    try:
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            },
            json={
                "model": "llama-3.3-70b-versatile",
                "temperature": 0.3,
                "messages": [
                    {"role": "system", "content": "Siz texnik sovutish kamerasi mutaxassisisiz."},
                    {"role": "user", "content": prompt}
                ]
            },
            timeout=60
        )
        response.raise_for_status()
        data = response.json()
        content = data["choices"][0]["message"]["content"]
        
        start = content.find("{")
        end = content.rfind("}")
        if start == -1 or end == -1:
            return {"success": False, "message": f"JSON parse bo'lmadi"}
        
        parsed = json.loads(content[start:end+1])
        return {"success": True, "data": parsed}
    
    except Exception as e:
        return {"success": False, "message": f"Groq xatolik: {e}"}


def send_to_telegram(project, calculations):
    """Telegram kanalga hisobot yuborish"""
    token = getattr(settings, 'TELEGRAM_BOT_TOKEN', os.getenv('TELEGRAM_BOT_TOKEN', ''))
    chat_id = getattr(settings, 'TELEGRAM_CHAT_ID', os.getenv('TELEGRAM_CHAT_ID', '-1002338157363'))
    
    if not token:
        return False, "TELEGRAM_BOT_TOKEN topilmadi"
    
    # Hisobot matnini tayyorlash
    top_report = " + ".join([
        f"{p['size']} ESHIK" if p["type"] == "door" else str(p["size"]) 
        for p in calculations['top_meta']
    ])
    right_report = " + ".join([
        f"{p['size']} ESHIK" if p["type"] == "door" else str(p["size"]) 
        for p in calculations['right_meta']
    ])
    
    message = f"""
<b>🏗 Yangi buyurtma / loyiha</b>

<b>Loyiha:</b> {project.project_name or '-'}
<b>Kod:</b> {project.room_code}
<b>Sana:</b> {datetime.now().strftime('%d.%m.%Y %H:%M')}

<b>Tashqi o'lcham:</b> {project.length_m} × {project.width_m} × {project.height_m} m
<b>Ichki foydali o'lcham:</b> {calculations['inner_L_mm']} × {calculations['inner_W_mm']} × {calculations['inner_H_mm']} mm
<b>Hajm:</b> {calculations['outer_volume_m3']} m³

<b>Devor:</b> {project.wall_type} / {calculations['wall_mm']} mm / {calculations['wall_area_m2']} m²
<b>Patalok:</b> {project.ceiling_type} / {calculations['ceil_mm']} mm / {calculations['ceiling_area_m2']} m²
<b>Pol:</b> {project.floor_type if project.has_floor else 'Mavjud emas'} / {calculations['floor_mm'] if project.has_floor else 0} mm

<b>Eshik:</b> {project.door_type}
<b>Eshik joylashuvi:</b> {project.door_side} / {project.door_position} / {project.door_opening}

<b>Agregat:</b> {project.unit_type} ({project.unit_brand}) / {project.unit_side}

<b>Panel ishchi eni:</b> {project.panel_width} m
<b>Umumiy devor paneli:</b> {calculations['wall_panels_total']} ta
<b>Patalok paneli:</b> {calculations['ceiling_panels_total']} ta
<b>Pol paneli:</b> {calculations['floor_panels_total']} ta

<b>Segmentlar (Uzunlik):</b> {top_report}
<b>Segmentlar (En):</b> {right_report}
""".strip()
    
    try:
        response = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={
                "chat_id": chat_id,
                "text": message,
                "parse_mode": "HTML",
                "disable_web_page_preview": True,
            },
            timeout=30
        )
        if response.status_code != 200:
            return False, f"Telegram xatolik: {response.text}"
        return True, "Yuborildi"
    except Exception as e:
        return False, f"Xatolik: {e}"


# =========================================================
# VIEWS
# =========================================================

@login_required
def constructor_index(request):
    """Konstruktor asosiy sahifasi"""
    
    # So'nggi loyihalar
    recent_projects = Project.objects.all().order_by('-created_at')[:10]
    
    # Tez kalkulyator formasi
    form = QuickCalculatorForm(request.GET or None)
    
    context = {
        'form': form,
        'recent_projects': recent_projects,
        'title': 'Sovutish Kamerasi Konstruktori',
    }
    
    return render(request, 'orders/chizma.html', context)


@login_required
def project_create(request):
    """Yangi loyiha yaratish"""
    
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.created_by = request.user
            
            # Hisob-kitoblarni bajarish
            calculations = calculate_all(project)
            project.calculations = calculations
            project.save()
            
            messages.success(request, f"✅ Loyiha '{project.room_code}' muvaffaqiyatli yaratildi!")
            return redirect('constructor:project_detail', pk=project.pk)
    else:
        form = ProjectForm()
    
    context = {
        'form': form,
        'title': 'Yangi Loyiha Yaratish',
    }
    
    return render(request, 'constructor/project_form.html', context)


@login_required
def project_detail(request, pk):
    """Loyiha tafsilotlari"""
    
    project = get_object_or_404(Project, pk=pk)
    
    # Hisob-kitoblar (agar mavjud bo'lmasa, qayta hisoblash)
    if not project.calculations:
        calculations = calculate_all(project)
        project.calculations = calculations
        project.save()
    else:
        calculations = project.calculations
    
    # SVG yaratish
    svg_content = generate_svg(project, calculations)
    
    context = {
        'project': project,
        'calculations': calculations,
        'svg_content': svg_content,
        'title': f'Loyiha: {project.room_code}',
    }
    
    return render(request, 'constructor/project_detail.html', context)


@login_required
def project_edit(request, pk):
    """Loyihani tahrirlash"""
    
    project = get_object_or_404(Project, pk=pk)
    
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            project = form.save()
            
            # Hisob-kitoblarni qayta bajarish
            calculations = calculate_all(project)
            project.calculations = calculations
            project.save()
            
            messages.success(request, f"✅ Loyiha '{project.room_code}' yangilandi!")
            return redirect('constructor:project_detail', pk=project.pk)
    else:
        form = ProjectForm(instance=project)
    
    context = {
        'form': form,
        'project': project,
        'title': f'Loyihani tahrirlash: {project.room_code}',
    }
    
    return render(request, 'constructor/project_form.html', context)


@login_required
def project_delete(request, pk):
    """Loyihani o'chirish"""
    
    project = get_object_or_404(Project, pk=pk)
    
    if request.method == 'POST':
        room_code = project.room_code
        project.delete()
        messages.success(request, f"✅ Loyiha '{room_code}' o'chirildi!")
        return redirect('constructor:project_list')
    
    context = {
        'project': project,
        'title': f'Loyihani o\'chirish: {project.room_code}',
    }
    
    return render(request, 'constructor/project_confirm_delete.html', context)


@login_required
def project_list(request):
    """Barcha loyihalar ro'yxati"""
    
    projects = Project.objects.all().order_by('-created_at')
    
    # Filtrlar
    search = request.GET.get('search', '')
    if search:
        projects = projects.filter(
            models.Q(project_name__icontains=search) |
            models.Q(room_code__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(projects, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'title': 'Barcha Loyihalar',
    }
    
    return render(request, 'constructor/project_list.html', context)


@login_required
def ai_recommendation(request, pk):
    """AI tavsiya olish"""
    
    project = get_object_or_404(Project, pk=pk)
    
    result = get_ai_recommendation(project)
    
    if result['success']:
        project.ai_result = result['data']
        project.save()
        messages.success(request, "✅ AI tavsiya muvaffaqiyatli olindi!")
    else:
        messages.error(request, f"❌ AI tavsiya olinmadi: {result['message']}")
    
    return redirect('constructor:project_detail', pk=project.pk)


@login_required
def send_report(request, pk):
    """Telegramga hisobot yuborish"""
    
    project = get_object_or_404(Project, pk=pk)
    
    if not project.calculations:
        calculations = calculate_all(project)
        project.calculations = calculations
        project.save()
    else:
        calculations = project.calculations
    
    success, message = send_to_telegram(project, calculations)
    
    if success:
        messages.success(request, "✅ Hisobot Telegram kanalga yuborildi!")
    else:
        messages.error(request, f"❌ Yuborilmadi: {message}")
    
    return redirect('constructor:project_detail', pk=project.pk)


@login_required
def create_order_from_project(request, pk):
    """Loyihadan buyurtma yaratish (orders app ga o'tkazish)"""
    
    project = get_object_or_404(Project, pk=pk)
    
    if not project.calculations:
        calculations = calculate_all(project)
        project.calculations = calculations
        project.save()
    else:
        calculations = project.calculations
    
    # Order modelini import qilish (orders app dan)
    try:
        from orders.models import Order
        
        order = Order.objects.create(
            order_number=f"ORD-{datetime.now().strftime('%Y%m%d')}-{Order.objects.count()+1:04d}",
            customer_name=project.project_name or "Noma'lum",
            customer_unique_id=project.room_code,
            product_name=f"Sovutish kamerasi {project.length_m}x{project.width_m}x{project.height_m}m ({project.wall_type})",
            panel_kvadrat=Decimal(str(calculations['total_panel_area_m2'])),
            panel_thickness=str(calculations['wall_mm']),
            total_price=Decimal('0'),  # Narxni keyinroq kiritish mumkin
            prepayment=Decimal('0'),
            status='KIRITILDI',
            created_by=request.user,
            comment=f"Konstruktor orqali yaratilgan. Loyiha ID: {project.id}",
        )
        
        messages.success(request, f"✅ Buyurtma #{order.order_number} muvaffaqiyatli yaratildi!")
        return redirect('order_detail', pk=order.pk)
        
    except ImportError:
        messages.error(request, "❌ Orders app topilmadi yoki ulanishda xatolik!")
    except Exception as e:
        messages.error(request, f"❌ Xatolik: {str(e)}")
    
    return redirect('constructor:project_detail', pk=project.pk)


# =========================================================
# API ENDPOINTLAR (AJAX uchun)
# =========================================================

@login_required
def api_calculate(request):
    """AJAX orqali hisob-kitob qilish"""
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Faqat POST so\'rovi'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Vaqtinchalik project obyekti yaratish
        class TempProject:
            def __init__(self, data):
                self.length_m = float(data.get('length', 5))
                self.width_m = float(data.get('width', 4))
                self.height_m = float(data.get('height', 3))
                self.wall_thickness = data.get('wall_thickness', '100mm')
                self.ceiling_thickness = data.get('ceiling_thickness', '80mm')
                self.floor_thickness = data.get('floor_thickness', '100mm')
                self.has_floor = data.get('has_floor', True)
                self.panel_width = float(data.get('panel_width', 1.16))
                self.door_type = data.get('door_type', 'Muzlatkich eshigi')
                self.door_side = data.get('door_side', 'Old')
                self.door_position = data.get('door_position', 'O\'rta')
                self.door_opening = data.get('door_opening', 'Ichkariga')
                self.unit_type = data.get('unit_type', 'Split-sistema (Nizkotemp)')
                self.unit_side = data.get('unit_side', 'Old')
                self.unit_brand = data.get('unit_brand', 'Bitzer')
                self.project_name = data.get('project_name', '')
                self.room_code = data.get('room_code', 'EP-001')
                self.product_type = data.get('product_type', 'Go\'sht')
                self.storage_temp = data.get('storage_temp', '-18°C')
                self.opening_freq = data.get('opening_freq', 'Kam')
                self.region = data.get('region', 'Mo\'tadil')
                self.humidity = data.get('humidity', 'Standart')
        
        project = TempProject(data)
        calculations = calculate_all(project)
        
        return JsonResponse({
            'success': True,
            'calculations': calculations
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

# orders/views.py - api_generate_svg funksiyasi

@login_required
def api_generate_svg(request):
    """AJAX orqali SVG yaratish"""
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Faqat POST so\'rovi'}, status=405)
    
    try:
        data = json.loads(request.body)
        print("API generate-svg received:", data.keys())  # DEBUG
        
        # Vaqtinchalik project yaratish
        class TempProject:
            def __init__(self, **kwargs):
                for key, value in kwargs.items():
                    setattr(self, key, value)
        
        # Kerakli maydonlarni olish
        project = TempProject(
            length_m=float(data.get('length', 5)),
            width_m=float(data.get('width', 4)),
            height_m=float(data.get('height', 3)),
            wall_thickness=data.get('wall_thickness', '100mm'),
            ceiling_thickness=data.get('ceiling_thickness', '80mm'),
            floor_thickness=data.get('floor_thickness', '100mm'),
            has_floor=data.get('has_floor', True),
            panel_width=float(data.get('panel_width', 1.16)),
            door_type=data.get('door_type', 'Muzlatkich eshigi'),
            door_side=data.get('door_side', 'Old'),
            door_position=data.get('door_position', 'O\'rta'),
            door_opening=data.get('door_opening', 'Ichkariga'),
            project_name=data.get('project_name', ''),
            room_code=data.get('room_code', 'EP-001'),
        )
        
        # Hisob-kitoblar (agar data da calculations bo'lmasa)
        if 'calculations' in data and data['calculations']:
            calculations = data['calculations']
        else:
            calculations = calculate_all(project)
        
        # SVG yaratish
        svg = generate_svg(project, calculations)
        
        return JsonResponse({
            'success': True,
            'svg': svg
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()  # Konsolga to'liq xatolikni chiqarish
        return JsonResponse({
            'success': False, 
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=400)

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

@login_required
def download_svg(request, pk):
    """
    Constructor loyihasi uchun SVG chizmani yuklab olish.
    Agar Project modeli mavjud bo'lsa, undan foydalanadi.
    """
    try:
        # Constructor app dagi Project modelini import qilishga harakat
        from constructor.models import Project
        from constructor.views import calculate_all, generate_svg
        
        project = get_object_or_404(Project, pk=pk)
        
        if not project.calculations:
            calculations = calculate_all(project)
            project.calculations = calculations
            project.save()
        else:
            calculations = project.calculations
        
        svg_content = generate_svg(project, calculations)
        
        response = HttpResponse(svg_content, content_type='image/svg+xml')
        response['Content-Disposition'] = f'attachment; filename="{project.room_code}_technical_sheet.svg"'
        return response
        
    except ImportError:
        # Agar constructor app hali yaratilmagan bo'lsa
        from django.http import HttpResponseNotFound
        return HttpResponseNotFound("Constructor app hali o'rnatilmagan")
    
# constructor/views.py ga qo'shing:

@login_required
def api_project_detail(request, pk):
    """AJAX orqali loyiha ma'lumotlarini olish"""
    project = get_object_or_404(Project, pk=pk)
    
    data = {
        'id': project.id,
        'project_name': project.project_name,
        'room_code': project.room_code,
        'length_m': float(project.length_m),
        'width_m': float(project.width_m),
        'height_m': float(project.height_m),
        'wall_type': project.wall_type,
        'wall_thickness': project.wall_thickness,
        'ceiling_type': project.ceiling_type,
        'ceiling_thickness': project.ceiling_thickness,
        'has_floor': project.has_floor,
        'floor_type': project.floor_type,
        'floor_thickness': project.floor_thickness,
        'panel_width': float(project.panel_width),
        'door_type': project.door_type,
        'door_side': project.door_side,
        'door_position': project.door_position,
        'door_opening': project.door_opening,
        'unit_type': project.unit_type,
        'unit_side': project.unit_side,
        'unit_brand': project.unit_brand,
        'calculations': project.calculations,
    }
    
    return JsonResponse(data)
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from io import BytesIO
import qrcode

from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, SimpleDocTemplate, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

@login_required
def order_receipt(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    PAGE_WIDTH = 80 * mm  
    PAGE_HEIGHT = 190 * mm  
    
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(PAGE_WIDTH, PAGE_HEIGHT),
        leftMargin=4 * mm,
        rightMargin=4 * mm,
        topMargin=5 * mm,
        bottomMargin=5 * mm
    )
    
    styles = getSampleStyleSheet()
    
    style_center_bold = ParagraphStyle('CenterBold', parent=styles['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=12, leading=14)
    style_center = ParagraphStyle('Center', parent=styles['Normal'], alignment=TA_CENTER, fontName='Helvetica', fontSize=8, leading=11, textColor=colors.HexColor('#222222'))
    style_left = ParagraphStyle('Left', parent=styles['Normal'], alignment=TA_LEFT, fontName='Helvetica', fontSize=8, leading=11)
    style_right = ParagraphStyle('Right', parent=styles['Normal'], alignment=TA_RIGHT, fontName='Helvetica', fontSize=8, leading=11)
    style_left_bold = ParagraphStyle('LeftBold', parent=styles['Normal'], alignment=TA_LEFT, fontName='Helvetica-Bold', fontSize=8, leading=11)
    
    story = []
    printable_width = PAGE_WIDTH - 8 * mm 
    
    story.append(Paragraph("ECO PROM", style_center_bold))
    story.append(Spacer(1, 1 * mm))
    story.append(Paragraph("Sovutish kameralari va panel<br/>ishlab chiqarish do'koni", style_center))
    story.append(Paragraph("+998(78)555-86-16<br/>+998(98)707-86-16<br/>+998(97)926-86-16", style_center))
    story.append(Paragraph("<b>www.ecopromuz.uz</b>", style_center))
    story.append(Spacer(1, 2 * mm))
    
    line_style = TableStyle([
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#333333')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('TOPPADDING', (0,0), (-1,-1), 1)
    ])
    t_line = Table([[""]], colWidths=[printable_width])
    t_line.setStyle(line_style)
    
    kassa_no = f"Kassa №{order.id % 10 + 1:02d}"
    chek_no = f"PRODAJA №{order.order_number}"
    sana_vaqt = timezone.now().strftime('%d.%m.%Y %H:%M:%S')
    sotuvchi = f"SOTUVCHI: {request.user.username}".upper()
    
    meta_data = [
        [Paragraph(kassa_no, style_left), Paragraph(chek_no, style_right)],
        [Paragraph(sana_vaqt, style_left), Paragraph("", style_right)]
    ]
    t_meta = Table(meta_data, colWidths=[printable_width/2, printable_width/2])
    story.append(t_meta)
    story.append(Spacer(1, 1 * mm))
    story.append(Paragraph(sotuvchi, style_left_bold))
    story.append(t_line)
    story.append(Spacer(1, 2 * mm))
    
    story.append(Paragraph(f"Mijoz: {order.customer_name} (ID: {order.customer_unique_id})", style_left))
    story.append(Spacer(1, 2 * mm))
    
    remaining = order.total_price - order.prepayment
    qqs_summa = float(order.total_price) * 12 / 112
    
    calc_data = [
        [Paragraph("Summa 12% QQS bilan:", style_left), Paragraph(f"{qqs_summa:,.2f}", style_right)],
        [Paragraph("<b>Jami:</b>", ParagraphStyle('J', parent=style_left, fontSize=11)), Paragraph(f"<b>{order.total_price:,.0f}</b>", ParagraphStyle('JR', parent=style_right, fontSize=11))],
        [Paragraph("To'landi (Zalog):", style_left), Paragraph(f"{order.prepayment:,.0f}", style_right)],
        [Paragraph("Qoldiq (Qarz):", style_left), Paragraph(f"{remaining:,.0f}", style_right)]
    ]
    t_calc = Table(calc_data, colWidths=[printable_width*0.6, printable_width*0.4])
    t_calc.setStyle(TableStyle([
        ('LINEBELOW', (0,0), (-1,0), 0.5, colors.HexColor('#666666')),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(t_calc)
    story.append(Spacer(1, 1 * mm))
    
    story.append(Paragraph("To'lov turi: PayMe / Naqd", style_left))
    story.append(Spacer(1, 2 * mm))
    
    if remaining > 0:
        story.append(Paragraph("■ DIQQAT: QARZDORLIK MAVJUD!", ParagraphStyle('W', parent=style_center_bold, fontSize=9)))
    else:
        story.append(Paragraph("✓ TO'LOV TO'LIQ QABUL QILINDI", ParagraphStyle('S', parent=style_center_bold, fontSize=9)))
        
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph("Xaridingiz uchun rahmat!!!", ParagraphStyle('Rahmat', parent=style_center_bold, fontName='Helvetica-Oblique', fontSize=11)))
    story.append(Spacer(1, 2 * mm))
    
    qr_link = f"https://taplink.cc/ecopromuz?order={order.order_number}"
    qr = qrcode.QRCode(version=1, box_size=5, border=1)
    qr.add_data(qr_link)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)
    
    qr_size = 32 * mm
    real_qr_image = RLImage(qr_buffer, width=qr_size, height=qr_size)
    
    t_qr = Table([[real_qr_image]], colWidths=[printable_width])
    t_qr.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0)
    ]))
    story.append(t_qr)
    
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="receipt_{order.order_number}.pdf"'
    return response
# orders/views.py ga qo'shing

from decimal import Decimal
import uuid
from datetime import datetime, timedelta
from django.db.models import Sum, Q
from django.core.paginator import Paginator
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==================== KASSA FUNKSIYALARI ====================

def is_cashier(user):
    """Kassa operatori ekanligini tekshirish"""
    if not user.is_authenticated:
        return False
    return (user.is_superuser or 
            is_in_group(user, 'Glavniy Admin') or 
            is_in_group(user, 'Menejer/Tasdiqlovchi'))




@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_dashboard(request):
    """Kassa boshqaruv paneli"""
    from decimal import Decimal
    from django.db.models import Sum
    from datetime import datetime, timedelta
    
    # ✅ Joriy qoldiq
    balance, created = CashRegisterBalance.objects.get_or_create(id=1)
    
    # ✅ TO'G'RI: Bugungi kun boshi va oxirini olish
    today = timezone.now().date()
    
    # Kun boshi (00:00:00) - Toshkent vaqti bilan
    today_start = timezone.make_aware(
        datetime.combine(today, datetime.min.time())
    )
    # Kun oxiri (23:59:59.999999)
    today_end = timezone.make_aware(
        datetime.combine(today, datetime.max.time())
    )
    
    # ✅ Bugungi operatsiyalar
    today_transactions = CashTransaction.objects.filter(
        transaction_date__gte=today_start,
        transaction_date__lte=today_end,
        status='COMPLETED'
    )
    
    # ✅ Bugungi statistikalar
    today_stats = {
        'total_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'total_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'cash_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME'],
            payment_method='CASH'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'cash_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'],
            payment_method='CASH'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'card_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME'],
            payment_method='CARD'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'card_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'],
            payment_method='CARD'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'click_income': today_transactions.filter(
            payment_method='CLICK', 
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'click_expense': today_transactions.filter(
            payment_method='CLICK', 
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'payme_income': today_transactions.filter(
            payment_method='PAYME', 
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'payme_expense': today_transactions.filter(
            payment_method='PAYME', 
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'bank_income': today_transactions.filter(
            payment_method='BANK', 
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'bank_expense': today_transactions.filter(
            payment_method='BANK', 
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
    }
    
    today_stats['net_change'] = today_stats['total_income'] - today_stats['total_expense']
    
    # ✅ So'nggi operatsiyalar
    recent_transactions = CashTransaction.objects.filter(
        status='COMPLETED'
    ).order_by('-transaction_date')[:20]
    
    # ✅ Bugungi hisobot mavjudligini tekshirish
    today_report = DailyCashReport.objects.filter(report_date=today).first()
    
    # ✅ Qarzdorlar soni va qarzdorlik summasi
    debt_stats = {
        'total_debt_usd': Debt.objects.filter(
            is_active=True, 
            currency='USD', 
            remaining__gt=0
        ).aggregate(Sum('remaining'))['remaining__sum'] or Decimal('0'),
        'total_debt_uzs': Debt.objects.filter(
            is_active=True, 
            currency='UZS', 
            remaining__gt=0
        ).aggregate(Sum('remaining'))['remaining__sum'] or Decimal('0'),
        'debt_count': Debt.objects.filter(
            is_active=True, 
            remaining__gt=0
        ).count(),
        'overdue_count': Debt.objects.filter(
            is_active=True, 
            remaining__gt=0, 
            due_date__lt=today
        ).count(),
    }
    
    # ✅ So'nggi qarz operatsiyalari
    recent_debt_transactions = DebtTransaction.objects.all().order_by('-created_at')[:10]
    
    context = {
        'balance': balance,
        'today_stats': today_stats,
        'today_transactions': today_transactions[:30],
        'today_transactions_count': today_transactions.count(),
        'recent_transactions': recent_transactions,
        'today_report': today_report,
        'today': today,
        'today_start': today_start,
        'today_end': today_end,
        'debt_stats': debt_stats,
        'recent_debt_transactions': recent_debt_transactions,
        'title': 'Kassa Boshqaruvi',
        'is_cashier': True,
    }
    
    return render(request, 'orders/cash_dashboard.html', context)


@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_transaction_create(request):
    """Yangi kassa operatsiyasi yaratish - Naqd, Karta, Click, Payme, Bank bilan"""
    from decimal import Decimal, InvalidOperation
    from django.db import transaction as db_transaction
    import uuid
    
    if request.method == 'POST':
        # ✅ POST dan ma'lumotlarni olish
        transaction_type = request.POST.get('transaction_type')
        amount_str = request.POST.get('amount', '0').replace(',', '.')
        currency = request.POST.get('currency', 'UZS')
        payment_method = request.POST.get('payment_method', 'CASH')
        customer_name = request.POST.get('customer_name', '').strip()
        description = request.POST.get('description', '').strip()
        category = request.POST.get('category', 'OTHER')
        
        # ✅ VALIDATSIYA
        if not transaction_type:
            messages.error(request, "❌ Operatsiya turini tanlang!")
            return redirect('cash_management')
        
        # ✅ Summani Decimal ga o'tkazish (xatolikni ushlash)
        try:
            amount = Decimal(amount_str)
        except InvalidOperation:
            messages.error(request, "❌ Summa noto'g'ri formatda! Iltimos, son kiriting.")
            return redirect('cash_management')
        
        if amount <= 0:
            messages.error(request, "❌ Summa 0 dan katta bo'lishi kerak!")
            return redirect('cash_management')
        
        # ✅ Customer name validatsiyasi
        if not customer_name:
            messages.error(request, "❌ Kim oldi/berdi maydonini to'ldiring!")
            return redirect('cash_management')
        
        # ✅ Description validatsiyasi
        if not description:
            messages.error(request, "❌ Izoh maydonini to'ldiring!")
            return redirect('cash_management')
        
        # ✅ Payment method validatsiyasi
        valid_payment_methods = ['CASH', 'CARD', 'CLICK', 'PAYME', 'BANK']
        if payment_method not in valid_payment_methods:
            payment_method = 'CASH'
        
        try:
            with db_transaction.atomic():
                # ✅ Tranzaksiyani yaratish
                transaction = CashTransaction.objects.create(
                    transaction_id=f"TXN-{uuid.uuid4().hex[:12].upper()}",
                    transaction_type=transaction_type,
                    amount=amount,
                    currency=currency,
                    payment_method=payment_method,
                    customer_name=customer_name,
                    description=description,
                    category=category,
                    status='COMPLETED',
                    performed_by=request.user,
                    transaction_date=timezone.now()
                )
                
                # ✅ Kassa qoldig'ini yangilash (FAQAT NAQD PUL UCHUN)
                balance, created = CashRegisterBalance.objects.get_or_create(id=1)
                
                if payment_method == 'CASH':
                    if transaction_type in ['INCOME', 'EXTERNAL_INCOME']:
                        if currency == 'USD':
                            balance.cash_balance_usd += amount
                        else:
                            balance.cash_balance += amount
                    else:  # EXPENSE yoki EXTERNAL_EXPENSE
                        if currency == 'USD':
                            balance.cash_balance_usd -= amount
                        else:
                            balance.cash_balance -= amount
                    
                    balance.updated_by = request.user
                    balance.save()
                    messages.success(
                        request, 
                        f"✅ Naqd pul operatsiyasi: {amount:,.2f} {currency} "
                        f"(Yangi qoldiq: {balance.cash_balance:,.2f} UZS / {balance.cash_balance_usd:,.2f} USD)"
                    )
                else:
                    # Karta, Click, Payme, Bank orqali to'lov
                    payment_names = {
                        'CASH': 'Naqd pul',
                        'CARD': 'Plastik karta',
                        'CLICK': 'Click',
                        'PAYME': 'Payme',
                        'BANK': 'Bank'
                    }
                    payment_name = payment_names.get(payment_method, payment_method)
                    messages.success(
                        request, 
                        f"✅ {payment_name} orqali operatsiya: {amount:,.2f} {currency} "
                        f"(Kassa balansiga ta'sir qilmaydi)"
                    )
                
        except Exception as e:
            messages.error(request, f"❌ Tizim xatoligi: {str(e)}")
            return redirect('cash_management')
        
        return redirect('cash_management')
    
    # ✅ GET so'rovi bo'lsa cash_management ga qaytarish
    return redirect('cash_management')


def update_cash_balance(transaction):
    """Kassa qoldig'ini yangilash - FAQAT NAQD PUL UCHUN"""
    balance, _ = CashRegisterBalance.objects.get_or_create(id=1)
    
    # ✅ FAQAT NAQD PUL operatsiyalari kassa qoldig'iga ta'sir qiladi
    if transaction.payment_method == 'CASH':
        if transaction.transaction_type in ['INCOME', 'EXTERNAL_INCOME']:
            if transaction.currency == 'USD':
                balance.cash_balance_usd += transaction.amount
            else:
                balance.cash_balance += transaction.amount
        else:  # EXPENSE
            if transaction.currency == 'USD':
                balance.cash_balance_usd -= transaction.amount
            else:
                balance.cash_balance -= transaction.amount
        
        balance.updated_by = transaction.performed_by
        balance.save()
    
    return balance


@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_transaction_list(request):
    """Kassa operatsiyalari ro'yxati"""
    
    # Filtrlar
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    transaction_type = request.GET.get('type', '')
    category = request.GET.get('category', '')
    payment_method = request.GET.get('payment_method', '')  # ✅ QO'SHILDI
    
    transactions = CashTransaction.objects.filter(status='COMPLETED')
    
    if start_date:
        transactions = transactions.filter(transaction_date__date__gte=start_date)
    if end_date:
        transactions = transactions.filter(transaction_date__date__lte=end_date)
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    if category:
        transactions = transactions.filter(category=category)
    if payment_method:
        transactions = transactions.filter(payment_method=payment_method)  # ✅ QO'SHILDI
    
    # Pagination
    paginator = Paginator(transactions.order_by('-transaction_date'), 50)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    # Statistikalar (to'lov usuli bo'yicha)
    stats = transactions.aggregate(
        total_income=Sum('amount', filter=Q(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'])),
        total_expense=Sum('amount', filter=Q(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'])),
        cash_total=Sum('amount', filter=Q(payment_method='CASH')),
        card_total=Sum('amount', filter=Q(payment_method='CARD')),  # ✅ QO'SHILDI
        click_total=Sum('amount', filter=Q(payment_method='CLICK')),
        payme_total=Sum('amount', filter=Q(payment_method='PAYME')),
        bank_total=Sum('amount', filter=Q(payment_method='BANK')),  # ✅ QO'SHILDI
    )
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'start_date': start_date,
        'end_date': end_date,
        'selected_type': transaction_type,
        'selected_category': category,
        'selected_payment': payment_method,
        'title': 'Kassa Operatsiyalari',
    }
    
    return render(request, 'orders/cash_transaction_list.html', context)
# views.py faylining boshiga import qismiga qo'shing
from .forms import CashTransactionForm, DailyReportForm
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def daily_report_create(request):
    """Kunlik hisobot yaratish (modal POST dan ishlaydi)"""
    
    if request.method == 'POST':
        today = timezone.now().date()
        
        # Bugungi hisobot mavjudligini tekshirish
        existing_report = DailyCashReport.objects.filter(report_date=today).first()
        if existing_report:
            messages.warning(request, f"{today} uchun hisobot allaqachon mavjud!")
            return redirect('cash_management')
        
        # POST dan ma'lumotlarni olish
        opening_balance = Decimal(request.POST.get('opening_balance', '0'))
        actual_balance = Decimal(request.POST.get('actual_balance', '0'))
        notes = request.POST.get('notes', '')
        
        # Bugungi operatsiyalar
        today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))
        
        today_transactions = CashTransaction.objects.filter(
            transaction_date__range=[today_start, today_end],
            status='COMPLETED'
        )
        
        # Bugungi statistikalar
        cash_income = today_transactions.filter(
            transaction_type='INCOME'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        
        cash_expense = today_transactions.filter(
            transaction_type='EXPENSE'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        
        total_income = cash_income
        total_expense = cash_expense
        expected_balance = opening_balance + total_income - total_expense
        difference = actual_balance - expected_balance
        
        # Hisobotni yaratish
        report = DailyCashReport.objects.create(
            report_date=today,
            opening_balance=opening_balance,
            cash_income=cash_income,
            cash_expense=cash_expense,
            click_income=Decimal('0'),
            payme_income=Decimal('0'),
            bank_income=Decimal('0'),
            click_expense=Decimal('0'),
            payme_expense=Decimal('0'),
            bank_expense=Decimal('0'),
            total_income=total_income,
            total_expense=total_expense,
            expected_balance=expected_balance,
            actual_balance=actual_balance,
            difference=difference,
            notes=notes,
            created_by=request.user
        )
        
        messages.success(request, f"{today} uchun kunlik hisobot yaratildi!")
        return redirect('cash_management')
    
    # GET so'rovi bo'lsa, cash_management sahifasiga qaytarish
    return redirect('cash_management')

@login_required
@user_passes_test(is_cashier, login_url='/login/')
def daily_report_detail(request, pk):
    """Kunlik hisobot tafsilotlari"""
    
    report = get_object_or_404(DailyCashReport, pk=pk)
    
    # Shu kundagi operatsiyalar
    start = timezone.make_aware(datetime.combine(report.report_date, datetime.min.time()))
    end = timezone.make_aware(datetime.combine(report.report_date, datetime.max.time()))
    
    transactions = CashTransaction.objects.filter(
        transaction_date__range=[start, end],
        status='COMPLETED'
    ).order_by('transaction_type', '-transaction_date')
    
    # Kategoriyalar bo'yicha guruhlash
    category_groups = {}
    for cat, _ in CashTransaction.CATEGORY_CHOICES:
        cat_transactions = transactions.filter(category=cat)
        if cat_transactions.exists():
            category_groups[cat] = {
                'total': cat_transactions.aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
                'count': cat_transactions.count(),
                'transactions': cat_transactions
            }
    
    context = {
        'report': report,
        'transactions': transactions,
        'category_groups': category_groups,
        'title': f"Kunlik Hisobot - {report.report_date}",
    }
    
    return render(request, 'orders/daily_report_detail.html', context)


@login_required
@user_passes_test(is_cashier, login_url='/login/')
def daily_report_list(request):
    """Kunlik hisobotlar ro'yxati"""
    
    reports = DailyCashReport.objects.all().order_by('-report_date')
    
    # Filtr
    year = request.GET.get('year')
    month = request.GET.get('month')
    
    if year:
        reports = reports.filter(report_date__year=year)
    if month:
        reports = reports.filter(report_date__month=month)
    
    paginator = Paginator(reports, 30)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    context = {
        'page_obj': page_obj,
        'title': 'Kunlik Hisobotlar',
        'years': range(2020, timezone.now().year + 1),
        'months': range(1, 13),
        'current_year': request.GET.get('year', str(timezone.now().year)),
        'current_month': request.GET.get('month', ''),
    }
    
    return render(request, 'orders/daily_report_list.html', context)

@login_required
@user_passes_test(is_cashier, login_url='/login/')
def export_cash_report_excel(request):
    """Kassa hisobotini Excel formatida eksport qilish - Barcha to'lov usullari bo'yicha alohida"""
    from datetime import datetime, timedelta
    from decimal import Decimal
    from django.db.models import Sum, Q
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.contrib import messages
    from django.shortcuts import redirect
    from django.utils import timezone
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date or not end_date:
        messages.error(request, "Iltimos, sana oralig'ini tanlang!")
        return redirect('cash_transaction_list')
    
    # Sana oralig'idagi operatsiyalar
    start = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
    end = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
    
    all_transactions = CashTransaction.objects.filter(
        transaction_date__range=[start, end],
        status='COMPLETED'
    ).order_by('-transaction_date')
    
    # ============ VALYUTALAR BO'YICHA ALOHIDA ============
    usd_transactions = all_transactions.filter(currency='USD')
    uzs_transactions = all_transactions.filter(currency='UZS')
    null_transactions = all_transactions.filter(currency__isnull=True)
    
    # ============ USD HISOBLAR ============
    usd_income = usd_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME']).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_expense = usd_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_net = usd_income - usd_expense
    
    # USD to'lov usullari bo'yicha
    usd_cash_income = usd_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CASH').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_cash_expense = usd_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CASH').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_card_income = usd_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CARD').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_card_expense = usd_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CARD').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_click_income = usd_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CLICK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_click_expense = usd_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CLICK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_payme_income = usd_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='PAYME').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_payme_expense = usd_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='PAYME').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_bank_income = usd_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='BANK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    usd_bank_expense = usd_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='BANK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    # ============ UZS HISOBLAR (null bilan birga) ============
    uzs_income = uzs_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME']).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_expense = uzs_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    null_income = null_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME']).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_expense = null_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    total_uzs_income = uzs_income + null_income
    total_uzs_expense = uzs_expense + null_expense
    uzs_net = total_uzs_income - total_uzs_expense
    
    # UZS to'lov usullari bo'yicha
    uzs_cash_income = uzs_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CASH').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_cash_expense = uzs_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CASH').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_card_income = uzs_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CARD').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_card_expense = uzs_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CARD').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_click_income = uzs_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CLICK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_click_expense = uzs_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CLICK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_payme_income = uzs_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='PAYME').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_payme_expense = uzs_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='PAYME').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_bank_income = uzs_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='BANK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    uzs_bank_expense = uzs_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='BANK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    # null bo'lganlarni qo'shamiz
    null_cash_income = null_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CASH').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_cash_expense = null_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CASH').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_card_income = null_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CARD').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_card_expense = null_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CARD').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_click_income = null_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CLICK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_click_expense = null_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CLICK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_payme_income = null_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='PAYME').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_payme_expense = null_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='PAYME').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_bank_income = null_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='BANK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    null_bank_expense = null_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='BANK').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    # UZS ga null larni qo'shamiz
    total_uzs_cash_income = uzs_cash_income + null_cash_income
    total_uzs_cash_expense = uzs_cash_expense + null_cash_expense
    total_uzs_card_income = uzs_card_income + null_card_income
    total_uzs_card_expense = uzs_card_expense + null_card_expense
    total_uzs_click_income = uzs_click_income + null_click_income
    total_uzs_click_expense = uzs_click_expense + null_click_expense
    total_uzs_payme_income = uzs_payme_income + null_payme_income
    total_uzs_payme_expense = uzs_payme_expense + null_payme_expense
    total_uzs_bank_income = uzs_bank_income + null_bank_income
    total_uzs_bank_expense = uzs_bank_expense + null_bank_expense
    
    # ============ JAMI HISOBLAR ============
    total_income = usd_income + total_uzs_income
    total_expense = usd_expense + total_uzs_expense
    total_net = total_income - total_expense
    
    # Jami to'lov usullari bo'yicha
    total_cash_income = usd_cash_income + total_uzs_cash_income
    total_cash_expense = usd_cash_expense + total_uzs_cash_expense
    total_card_income = usd_card_income + total_uzs_card_income
    total_card_expense = usd_card_expense + total_uzs_card_expense
    total_click_income = usd_click_income + total_uzs_click_income
    total_click_expense = usd_click_expense + total_uzs_click_expense
    total_payme_income = usd_payme_income + total_uzs_payme_income
    total_payme_expense = usd_payme_expense + total_uzs_payme_expense
    total_bank_income = usd_bank_income + total_uzs_bank_income
    total_bank_expense = usd_bank_expense + total_uzs_bank_expense
    
    # ============ EXCEL FAYL YARATISH ============
    wb = openpyxl.Workbook()
    
    # ============ STILLAR ============
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    stats_font = Font(name='Segoe UI', size=12, bold=True)
    
    usd_header_fill = PatternFill(start_color="B45309", end_color="B45309", fill_type="solid")
    uzs_header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    total_header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    payment_header_fill = PatternFill(start_color="6B7280", end_color="6B7280", fill_type="solid")
    
    income_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    expense_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    usd_bg_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    uzs_bg_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    total_bg_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    money_format = '#,##0.00'
    payment_headers = ['To\'lov usuli', 'Kirim', 'Chiqim', 'Sof']
    
    # ==============================================================
    # SHEET 1: USD (AQSH DOLLARI)
    # ==============================================================
    ws_usd = wb.active
    ws_usd.title = "USD"
    
    # Sarlavha
    ws_usd.merge_cells('A1:H1')
    ws_usd['A1'] = f"USD (AQSH DOLLARI) OPERATSIYALARI"
    ws_usd['A1'].font = Font(bold=True, size=16, color='B45309')
    ws_usd['A1'].alignment = center_alignment
    
    ws_usd.merge_cells('A2:H2')
    ws_usd['A2'] = f"Davr: {start_date} - {end_date}"
    ws_usd['A2'].font = Font(size=11, color='666666')
    ws_usd['A2'].alignment = center_alignment
    
    # ======== 1. KATTA STATISTIKA (TEPADA) ========
    row = 4
    
    # JAMI KIRIM
    ws_usd.merge_cells(f'A{row}:B{row}')
    ws_usd.cell(row=row, column=1, value="JAMI KIRIM")
    ws_usd.cell(row=row, column=1).font = Font(bold=True, size=14, color='FFFFFF')
    ws_usd.cell(row=row, column=1).fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws_usd.cell(row=row, column=1).alignment = center_alignment
    
    ws_usd.merge_cells(f'C{row}:D{row}')
    ws_usd.cell(row=row, column=3, value=f"{float(usd_income):,.2f}")
    ws_usd.cell(row=row, column=3).font = Font(bold=True, size=16, color='10B981')
    ws_usd.cell(row=row, column=3).alignment = right_alignment
    
    # JAMI CHIQIM
    ws_usd.merge_cells(f'E{row}:F{row}')
    ws_usd.cell(row=row, column=5, value="JAMI CHIQIM")
    ws_usd.cell(row=row, column=5).font = Font(bold=True, size=14, color='FFFFFF')
    ws_usd.cell(row=row, column=5).fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    ws_usd.cell(row=row, column=5).alignment = center_alignment
    
    ws_usd.merge_cells(f'G{row}:H{row}')
    ws_usd.cell(row=row, column=7, value=f"{float(usd_expense):,.2f}")
    ws_usd.cell(row=row, column=7).font = Font(bold=True, size=16, color='EF4444')
    ws_usd.cell(row=row, column=7).alignment = right_alignment
    
    # SOF NATIJA
    ws_usd.merge_cells(f'I{row}:J{row}')
    ws_usd.cell(row=row, column=9, value="SOF NATIJA")
    ws_usd.cell(row=row, column=9).font = Font(bold=True, size=14, color='FFFFFF')
    ws_usd.cell(row=row, column=9).fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws_usd.cell(row=row, column=9).alignment = center_alignment
    
    ws_usd.merge_cells(f'K{row}:L{row}')
    net_color = "10B981" if usd_net >= 0 else "EF4444"
    ws_usd.cell(row=row, column=11, value=f"{float(usd_net):,.2f}")
    ws_usd.cell(row=row, column=11).font = Font(bold=True, size=16, color=net_color)
    ws_usd.cell(row=row, column=11).alignment = right_alignment
    
    # ======== 2. TO'LOV USULLARI BO'YICHA STATISTIKA ========
    row += 2
    ws_usd.cell(row=row, column=1, value="TO'LOV USULLARI BO'YICHA STATISTIKA")
    ws_usd.cell(row=row, column=1).font = stats_font
    ws_usd.merge_cells(f'A{row}:L{row}')
    row += 1
    
    for col, header in enumerate(payment_headers, 1):
        cell = ws_usd.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = payment_header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    usd_payment_data = [
        ('Naqd pul', float(usd_cash_income), float(usd_cash_expense), float(usd_cash_income - usd_cash_expense)),
        ('Plastik karta', float(usd_card_income), float(usd_card_expense), float(usd_card_income - usd_card_expense)),
        ('Click', float(usd_click_income), float(usd_click_expense), float(usd_click_income - usd_click_expense)),
        ('Payme', float(usd_payme_income), float(usd_payme_expense), float(usd_payme_income - usd_payme_expense)),
        ('Bank', float(usd_bank_income), float(usd_bank_expense), float(usd_bank_income - usd_bank_expense)),
    ]
    
    row += 1
    for name, inc, exp, net in usd_payment_data:
        ws_usd.cell(row=row, column=1, value=name)
        ws_usd.cell(row=row, column=2, value=inc)
        ws_usd.cell(row=row, column=3, value=exp)
        ws_usd.cell(row=row, column=4, value=net)
        
        if inc > 0:
            ws_usd.cell(row=row, column=2).fill = income_fill
        if exp > 0:
            ws_usd.cell(row=row, column=3).fill = expense_fill
        ws_usd.cell(row=row, column=4).font = Font(bold=True, color='10B981' if net >= 0 else 'EF4444')
        
        for col in range(1, 5):
            ws_usd.cell(row=row, column=col).border = thin_border
            if col > 1:
                ws_usd.cell(row=row, column=col).number_format = money_format
                ws_usd.cell(row=row, column=col).alignment = right_alignment
            else:
                ws_usd.cell(row=row, column=col).alignment = left_alignment
        row += 1
    
    # Jami qator
    ws_usd.cell(row=row, column=1, value="JAMI")
    ws_usd.cell(row=row, column=1).font = Font(bold=True)
    ws_usd.cell(row=row, column=2, value=float(usd_income))
    ws_usd.cell(row=row, column=3, value=float(usd_expense))
    ws_usd.cell(row=row, column=4, value=float(usd_net))
    ws_usd.cell(row=row, column=4).font = Font(bold=True, color='10B981' if usd_net >= 0 else 'EF4444')
    for col in range(1, 5):
        ws_usd.cell(row=row, column=col).border = thin_border
        if col > 1:
            ws_usd.cell(row=row, column=col).number_format = money_format
            ws_usd.cell(row=row, column=col).alignment = right_alignment
        else:
            ws_usd.cell(row=row, column=col).alignment = center_alignment
            ws_usd.cell(row=row, column=col).fill = total_bg_fill
    
    # ======== 3. OPERATSIYALAR RO'YXATI (PASTDA) ========
    row += 2
    ws_usd.cell(row=row, column=1, value="USD OPERATSIYALARI RO'YXATI")
    ws_usd.cell(row=row, column=1).font = stats_font
    ws_usd.merge_cells(f'A{row}:L{row}')
    row += 1
    
    headers = ['No', 'Sana', 'Operatsiya ID', 'Tur', "To'lov usuli", 'Summa (USD)', 'Kim oldi/berdi', 'Izoh']
    for col, header in enumerate(headers, 1):
        cell = ws_usd.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = usd_header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    ws_usd.freeze_panes = ws_usd['A' + str(row + 1)]
    
    for idx, trans in enumerate(usd_transactions.order_by('-transaction_date'), 1):
        row += 1
        ws_usd.cell(row=row, column=1, value=idx)
        ws_usd.cell(row=row, column=2, value=trans.transaction_date.strftime('%d.%m.%Y %H:%M'))
        ws_usd.cell(row=row, column=3, value=trans.transaction_id)
        
        tur = "Kirim" if trans.transaction_type in ['INCOME', 'EXTERNAL_INCOME'] else "Chiqim"
        ws_usd.cell(row=row, column=4, value=tur)
        ws_usd.cell(row=row, column=4).fill = income_fill if tur == "Kirim" else expense_fill
        
        ws_usd.cell(row=row, column=5, value=trans.get_payment_method_display())
        ws_usd.cell(row=row, column=6, value=float(trans.amount))
        ws_usd.cell(row=row, column=7, value=trans.customer_name or '-')
        ws_usd.cell(row=row, column=8, value=trans.description[:60] if trans.description else '-')
        
        for col in range(1, 9):
            ws_usd.cell(row=row, column=col).border = thin_border
            if col == 6:
                ws_usd.cell(row=row, column=6).number_format = money_format
                ws_usd.cell(row=row, column=6).alignment = right_alignment
            elif col in [4, 5]:
                ws_usd.cell(row=row, column=col).alignment = center_alignment
    
    # Ustun kengliklari
    ws_usd.column_dimensions['A'].width = 6
    ws_usd.column_dimensions['B'].width = 18
    ws_usd.column_dimensions['C'].width = 22
    ws_usd.column_dimensions['D'].width = 10
    ws_usd.column_dimensions['E'].width = 16
    ws_usd.column_dimensions['F'].width = 18
    ws_usd.column_dimensions['G'].width = 22
    ws_usd.column_dimensions['H'].width = 50
    
    # ==============================================================
    # SHEET 2: UZS (SO'M) - RASMDAGIDEK
    # ==============================================================
    ws_uzs = wb.create_sheet("UZS")
    
    # Sarlavha
    ws_uzs.merge_cells('A1:H1')
    ws_uzs['A1'] = f"UZS (SO'M) OPERATSIYALARI"
    ws_uzs['A1'].font = Font(bold=True, size=16, color='1E40AF')
    ws_uzs['A1'].alignment = center_alignment
    
    ws_uzs.merge_cells('A2:H2')
    ws_uzs['A2'] = f"Davr: {start_date} - {end_date}"
    ws_uzs['A2'].font = Font(size=11, color='666666')
    ws_uzs['A2'].alignment = center_alignment
    
    # ======== 1. KATTA STATISTIKA (TEPADA) ========
    row = 4
    
    # JAMI KIRIM
    ws_uzs.merge_cells(f'A{row}:B{row}')
    ws_uzs.cell(row=row, column=1, value="JAMI KIRIM")
    ws_uzs.cell(row=row, column=1).font = Font(bold=True, size=14, color='FFFFFF')
    ws_uzs.cell(row=row, column=1).fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws_uzs.cell(row=row, column=1).alignment = center_alignment
    
    ws_uzs.merge_cells(f'C{row}:D{row}')
    ws_uzs.cell(row=row, column=3, value=f"{float(total_uzs_income):,.2f}")
    ws_uzs.cell(row=row, column=3).font = Font(bold=True, size=16, color='10B981')
    ws_uzs.cell(row=row, column=3).alignment = right_alignment
    
    # JAMI CHIQIM
    ws_uzs.merge_cells(f'E{row}:F{row}')
    ws_uzs.cell(row=row, column=5, value="JAMI CHIQIM")
    ws_uzs.cell(row=row, column=5).font = Font(bold=True, size=14, color='FFFFFF')
    ws_uzs.cell(row=row, column=5).fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    ws_uzs.cell(row=row, column=5).alignment = center_alignment
    
    ws_uzs.merge_cells(f'G{row}:H{row}')
    ws_uzs.cell(row=row, column=7, value=f"{float(total_uzs_expense):,.2f}")
    ws_uzs.cell(row=row, column=7).font = Font(bold=True, size=16, color='EF4444')
    ws_uzs.cell(row=row, column=7).alignment = right_alignment
    
    # SOF NATIJA
    ws_uzs.merge_cells(f'I{row}:J{row}')
    ws_uzs.cell(row=row, column=9, value="SOF NATIJA")
    ws_uzs.cell(row=row, column=9).font = Font(bold=True, size=14, color='FFFFFF')
    ws_uzs.cell(row=row, column=9).fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws_uzs.cell(row=row, column=9).alignment = center_alignment
    
    ws_uzs.merge_cells(f'K{row}:L{row}')
    net_color_uzs = "10B981" if uzs_net >= 0 else "EF4444"
    ws_uzs.cell(row=row, column=11, value=f"{float(uzs_net):,.2f}")
    ws_uzs.cell(row=row, column=11).font = Font(bold=True, size=16, color=net_color_uzs)
    ws_uzs.cell(row=row, column=11).alignment = right_alignment
    
    # ======== 2. TO'LOV USULLARI BO'YICHA STATISTIKA ========
    row += 2
    ws_uzs.cell(row=row, column=1, value="TO'LOV USULLARI BO'YICHA STATISTIKA")
    ws_uzs.cell(row=row, column=1).font = stats_font
    ws_uzs.merge_cells(f'A{row}:L{row}')
    row += 1
    
    for col, header in enumerate(payment_headers, 1):
        cell = ws_uzs.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = payment_header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    uzs_payment_data = [
        ('Naqd pul', float(total_uzs_cash_income), float(total_uzs_cash_expense), float(total_uzs_cash_income - total_uzs_cash_expense)),
        ('Plastik karta', float(total_uzs_card_income), float(total_uzs_card_expense), float(total_uzs_card_income - total_uzs_card_expense)),
        ('Click', float(total_uzs_click_income), float(total_uzs_click_expense), float(total_uzs_click_income - total_uzs_click_expense)),
        ('Payme', float(total_uzs_payme_income), float(total_uzs_payme_expense), float(total_uzs_payme_income - total_uzs_payme_expense)),
        ('Bank', float(total_uzs_bank_income), float(total_uzs_bank_expense), float(total_uzs_bank_income - total_uzs_bank_expense)),
    ]
    
    row += 1
    for name, inc, exp, net in uzs_payment_data:
        ws_uzs.cell(row=row, column=1, value=name)
        ws_uzs.cell(row=row, column=2, value=inc)
        ws_uzs.cell(row=row, column=3, value=exp)
        ws_uzs.cell(row=row, column=4, value=net)
        
        if inc > 0:
            ws_uzs.cell(row=row, column=2).fill = income_fill
        if exp > 0:
            ws_uzs.cell(row=row, column=3).fill = expense_fill
        ws_uzs.cell(row=row, column=4).font = Font(bold=True, color='10B981' if net >= 0 else 'EF4444')
        
        for col in range(1, 5):
            ws_uzs.cell(row=row, column=col).border = thin_border
            if col > 1:
                ws_uzs.cell(row=row, column=col).number_format = money_format
                ws_uzs.cell(row=row, column=col).alignment = right_alignment
            else:
                ws_uzs.cell(row=row, column=col).alignment = left_alignment
        row += 1
    
    # Jami qator
    ws_uzs.cell(row=row, column=1, value="JAMI")
    ws_uzs.cell(row=row, column=1).font = Font(bold=True)
    ws_uzs.cell(row=row, column=2, value=float(total_uzs_income))
    ws_uzs.cell(row=row, column=3, value=float(total_uzs_expense))
    ws_uzs.cell(row=row, column=4, value=float(uzs_net))
    ws_uzs.cell(row=row, column=4).font = Font(bold=True, color='10B981' if uzs_net >= 0 else 'EF4444')
    for col in range(1, 5):
        ws_uzs.cell(row=row, column=col).border = thin_border
        if col > 1:
            ws_uzs.cell(row=row, column=col).number_format = money_format
            ws_uzs.cell(row=row, column=col).alignment = right_alignment
        else:
            ws_uzs.cell(row=row, column=col).alignment = center_alignment
            ws_uzs.cell(row=row, column=col).fill = total_bg_fill
    
    # ======== 3. OPERATSIYALAR RO'YXATI (PASTDA) ========
    row += 2
    ws_uzs.cell(row=row, column=1, value="UZS OPERATSIYALARI RO'YXATI")
    ws_uzs.cell(row=row, column=1).font = stats_font
    ws_uzs.merge_cells(f'A{row}:L{row}')
    row += 1
    
    for col, header in enumerate(headers, 1):
        cell = ws_uzs.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = uzs_header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    ws_uzs.freeze_panes = ws_uzs['A' + str(row + 1)]
    
    uzs_all = list(uzs_transactions) + list(null_transactions)
    uzs_all.sort(key=lambda x: x.transaction_date, reverse=True)
    
    for idx, trans in enumerate(uzs_all, 1):
        row += 1
        ws_uzs.cell(row=row, column=1, value=idx)
        ws_uzs.cell(row=row, column=2, value=trans.transaction_date.strftime('%d.%m.%Y %H:%M'))
        ws_uzs.cell(row=row, column=3, value=trans.transaction_id)
        
        tur = "Kirim" if trans.transaction_type in ['INCOME', 'EXTERNAL_INCOME'] else "Chiqim"
        ws_uzs.cell(row=row, column=4, value=tur)
        ws_uzs.cell(row=row, column=4).fill = income_fill if tur == "Kirim" else expense_fill
        
        ws_uzs.cell(row=row, column=5, value=trans.get_payment_method_display())
        ws_uzs.cell(row=row, column=6, value=float(trans.amount))
        ws_uzs.cell(row=row, column=7, value=trans.customer_name or '-')
        ws_uzs.cell(row=row, column=8, value=trans.description[:60] if trans.description else '-')
        
        for col in range(1, 9):
            ws_uzs.cell(row=row, column=col).border = thin_border
            if col == 6:
                ws_uzs.cell(row=row, column=6).number_format = money_format
                ws_uzs.cell(row=row, column=6).alignment = right_alignment
            elif col in [4, 5]:
                ws_uzs.cell(row=row, column=col).alignment = center_alignment
    
    # Ustun kengliklari
    ws_uzs.column_dimensions['A'].width = 6
    ws_uzs.column_dimensions['B'].width = 18
    ws_uzs.column_dimensions['C'].width = 22
    ws_uzs.column_dimensions['D'].width = 10
    ws_uzs.column_dimensions['E'].width = 16
    ws_uzs.column_dimensions['F'].width = 18
    ws_uzs.column_dimensions['G'].width = 22
    ws_uzs.column_dimensions['H'].width = 50
    
    # ==============================================================
    # SHEET 3: JAMI (UMUMIY)
    # ==============================================================
    ws_total = wb.create_sheet("JAMI")
    
    ws_total.merge_cells('A1:H1')
    ws_total['A1'] = f"UMUMIY HISOBOT"
    ws_total['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws_total['A1'].fill = total_header_fill
    ws_total['A1'].alignment = center_alignment
    
    ws_total.merge_cells('A2:H2')
    ws_total['A2'] = f"Davr: {start_date} - {end_date}"
    ws_total['A2'].font = Font(size=11, color='666666')
    ws_total['A2'].alignment = center_alignment
    
    # ======== 1. KATTA STATISTIKA ========
    row = 4
    
    ws_total.merge_cells(f'A{row}:B{row}')
    ws_total.cell(row=row, column=1, value="JAMI KIRIM")
    ws_total.cell(row=row, column=1).font = Font(bold=True, size=14, color='FFFFFF')
    ws_total.cell(row=row, column=1).fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws_total.cell(row=row, column=1).alignment = center_alignment
    
    ws_total.merge_cells(f'C{row}:D{row}')
    ws_total.cell(row=row, column=3, value=f"{float(total_income):,.2f}")
    ws_total.cell(row=row, column=3).font = Font(bold=True, size=16, color='10B981')
    ws_total.cell(row=row, column=3).alignment = right_alignment
    
    ws_total.merge_cells(f'E{row}:F{row}')
    ws_total.cell(row=row, column=5, value="JAMI CHIQIM")
    ws_total.cell(row=row, column=5).font = Font(bold=True, size=14, color='FFFFFF')
    ws_total.cell(row=row, column=5).fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    ws_total.cell(row=row, column=5).alignment = center_alignment
    
    ws_total.merge_cells(f'G{row}:H{row}')
    ws_total.cell(row=row, column=7, value=f"{float(total_expense):,.2f}")
    ws_total.cell(row=row, column=7).font = Font(bold=True, size=16, color='EF4444')
    ws_total.cell(row=row, column=7).alignment = right_alignment
    
    ws_total.merge_cells(f'I{row}:J{row}')
    ws_total.cell(row=row, column=9, value="SOF NATIJA")
    ws_total.cell(row=row, column=9).font = Font(bold=True, size=14, color='FFFFFF')
    ws_total.cell(row=row, column=9).fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws_total.cell(row=row, column=9).alignment = center_alignment
    
    ws_total.merge_cells(f'K{row}:L{row}')
    net_color_total = "10B981" if total_net >= 0 else "EF4444"
    ws_total.cell(row=row, column=11, value=f"{float(total_net):,.2f}")
    ws_total.cell(row=row, column=11).font = Font(bold=True, size=16, color=net_color_total)
    ws_total.cell(row=row, column=11).alignment = right_alignment
    
    # ======== 2. TO'LOV USULLARI BO'YICHA ========
    row += 2
    ws_total.cell(row=row, column=1, value="TO'LOV USULLARI BO'YICHA STATISTIKA")
    ws_total.cell(row=row, column=1).font = stats_font
    ws_total.merge_cells(f'A{row}:L{row}')
    row += 1
    
    for col, header in enumerate(payment_headers, 1):
        cell = ws_total.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = payment_header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    total_payment_data = [
        ('Naqd pul', float(total_cash_income), float(total_cash_expense), float(total_cash_income - total_cash_expense)),
        ('Plastik karta', float(total_card_income), float(total_card_expense), float(total_card_income - total_card_expense)),
        ('Click', float(total_click_income), float(total_click_expense), float(total_click_income - total_click_expense)),
        ('Payme', float(total_payme_income), float(total_payme_expense), float(total_payme_income - total_payme_expense)),
        ('Bank', float(total_bank_income), float(total_bank_expense), float(total_bank_income - total_bank_expense)),
    ]
    
    row += 1
    for name, inc, exp, net in total_payment_data:
        ws_total.cell(row=row, column=1, value=name)
        ws_total.cell(row=row, column=2, value=inc)
        ws_total.cell(row=row, column=3, value=exp)
        ws_total.cell(row=row, column=4, value=net)
        
        if inc > 0:
            ws_total.cell(row=row, column=2).fill = income_fill
        if exp > 0:
            ws_total.cell(row=row, column=3).fill = expense_fill
        ws_total.cell(row=row, column=4).font = Font(bold=True, color='10B981' if net >= 0 else 'EF4444')
        
        for col in range(1, 5):
            ws_total.cell(row=row, column=col).border = thin_border
            if col > 1:
                ws_total.cell(row=row, column=col).number_format = money_format
                ws_total.cell(row=row, column=col).alignment = right_alignment
            else:
                ws_total.cell(row=row, column=col).alignment = left_alignment
        row += 1
    
    ws_total.cell(row=row, column=1, value="JAMI")
    ws_total.cell(row=row, column=1).font = Font(bold=True)
    ws_total.cell(row=row, column=2, value=float(total_income))
    ws_total.cell(row=row, column=3, value=float(total_expense))
    ws_total.cell(row=row, column=4, value=float(total_net))
    ws_total.cell(row=row, column=4).font = Font(bold=True, color='10B981' if total_net >= 0 else 'EF4444')
    for col in range(1, 5):
        ws_total.cell(row=row, column=col).border = thin_border
        if col > 1:
            ws_total.cell(row=row, column=col).number_format = money_format
            ws_total.cell(row=row, column=col).alignment = right_alignment
        else:
            ws_total.cell(row=row, column=col).alignment = center_alignment
            ws_total.cell(row=row, column=col).fill = total_bg_fill
    
    # ======== 3. VALYUTALAR BO'YICHA TAQQOSLASH ========
    row += 2
    ws_total.cell(row=row, column=1, value="VALYUTALAR BO'YICHA TAQQOSLASH")
    ws_total.cell(row=row, column=1).font = stats_font
    ws_total.merge_cells(f'A{row}:L{row}')
    row += 1
    
    compare_headers = ['Valyuta', 'Jami kirim', 'Jami chiqim', 'Sof natija', 'Operatsiyalar soni']
    for col, header in enumerate(compare_headers, 1):
        cell = ws_total.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.alignment = center_alignment
        cell.border = thin_border
    
    row += 1
    ws_total.cell(row=row, column=1, value="USD")
    ws_total.cell(row=row, column=2, value=float(usd_income))
    ws_total.cell(row=row, column=3, value=float(usd_expense))
    ws_total.cell(row=row, column=4, value=float(usd_net))
    ws_total.cell(row=row, column=5, value=usd_transactions.count())
    ws_total.cell(row=row, column=1).fill = usd_bg_fill
    
    row += 1
    ws_total.cell(row=row, column=1, value="UZS")
    ws_total.cell(row=row, column=2, value=float(total_uzs_income))
    ws_total.cell(row=row, column=3, value=float(total_uzs_expense))
    ws_total.cell(row=row, column=4, value=float(uzs_net))
    ws_total.cell(row=row, column=5, value=uzs_transactions.count() + null_transactions.count())
    ws_total.cell(row=row, column=1).fill = uzs_bg_fill
    
    row += 1
    ws_total.cell(row=row, column=1, value="JAMI")
    ws_total.cell(row=row, column=1).font = Font(bold=True)
    ws_total.cell(row=row, column=2, value=float(total_income))
    ws_total.cell(row=row, column=3, value=float(total_expense))
    ws_total.cell(row=row, column=4, value=float(total_net))
    ws_total.cell(row=row, column=5, value=all_transactions.count())
    ws_total.cell(row=row, column=1).fill = total_bg_fill
    ws_total.cell(row=row, column=4).font = Font(bold=True, color='10B981' if total_net >= 0 else 'EF4444')
    
    for r in range(row-2, row+1):
        for col in range(2, 6):
            ws_total.cell(row=r, column=col).number_format = money_format
            ws_total.cell(row=r, column=col).alignment = right_alignment
        ws_total.cell(row=r, column=1).alignment = center_alignment
    
    # Ustun kengliklari
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws_total.column_dimensions[col].width = 18
    
    # Faylni yuborish
    filename = f"kassa_hisoboti_{start_date}_{end_date}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


@login_required
@user_passes_test(is_cashier, login_url='/login/')
def order_payment_click(request, order_id):
    """Buyurtma uchun Click orqali to'lov qilish (simulyatsiya)"""
    
    order = get_object_or_404(Order, pk=order_id)
    remaining = order.total_price - order.prepayment
    
    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', '0'))
        
        if amount <= 0:
            messages.error(request, "To'lov summasi 0 dan katta bo'lishi kerak!")
        elif amount > remaining:
            messages.error(request, f"To'lov summasi qarzdan ({remaining:,.0f}) ko'p bo'lishi mumkin emas!")
        else:
            # Click to'lov simulyatsiyasi
            transaction_id = f"CLICK-{uuid.uuid4().hex[:12].upper()}"
            
            # Operatsiyani yaratish
            transaction = CashTransaction.objects.create(
                transaction_type='EXTERNAL_INCOME',
                category='CUSTOMER_PAYMENT',
                amount=amount,
                payment_method='CLICK',
                order=order,
                customer_name=order.customer_name,
                description=f"Buyurtma #{order.order_number} uchun Click orqali to'lov",
                external_payment_id=transaction_id,
                performed_by=request.user,
                external_payment_data={
                    'merchant_id': 'ECOPROM',
                    'order_id': order.order_number,
                    'payment_type': 'click'
                }
            )
            
            # Order prepaymentni yangilash
            order.prepayment += amount
            order.save()
            
            messages.success(request, f"✅ Click orqali {amount:,.0f} so'm to'lov qabul qilindi! Tranzaksiya ID: {transaction_id}")
            
            return redirect('order_detail', pk=order.pk)
    
    context = {
        'order': order,
        'remaining': remaining,
        'title': f"Click To'lov - #{order.order_number}",
    }
    
    return render(request, 'orders/order_payment_click.html', context)


@login_required
@user_passes_test(is_cashier, login_url='/login/')
def order_payment_payme(request, order_id):
    """Buyurtma uchun Payme orqali to'lov qilish (simulyatsiya)"""
    
    order = get_object_or_404(Order, pk=order_id)
    remaining = order.total_price - order.prepayment
    
    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', '0'))
        
        if amount <= 0:
            messages.error(request, "To'lov summasi 0 dan katta bo'lishi kerak!")
        elif amount > remaining:
            messages.error(request, f"To'lov summasi qarzdan ({remaining:,.0f}) ko'p bo'lishi mumkin emas!")
        else:
            # Payme to'lov simulyatsiyasi
            transaction_id = f"PAYME-{uuid.uuid4().hex[:12].upper()}"
            
            transaction = CashTransaction.objects.create(
                transaction_type='EXTERNAL_INCOME',
                category='CUSTOMER_PAYMENT',
                amount=amount,
                payment_method='PAYME',
                order=order,
                customer_name=order.customer_name,
                description=f"Buyurtma #{order.order_number} uchun Payme orqali to'lov",
                external_payment_id=transaction_id,
                performed_by=request.user,
                external_payment_data={
                    'merchant_id': 'ECOPROM',
                    'order_id': order.order_number,
                    'payment_type': 'payme'
                }
            )
            
            order.prepayment += amount
            order.save()
            
            messages.success(request, f"✅ Payme orqali {amount:,.0f} so'm to'lov qabul qilindi! Tranzaksiya ID: {transaction_id}")
            
            return redirect('order_detail', pk=order.pk)
    
    context = {
        'order': order,
        'remaining': remaining,
        'title': f"Payme To'lov - #{order.order_number}",
    }
    
    return render(request, 'orders/order_payment_payme.html', context)
# orders/views.py ga qo'shing

@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_management(request):
    """Yagona kassa boshqaruv sahifasi"""
    
    from datetime import datetime, timedelta
    from decimal import Decimal
    from django.core.paginator import Paginator
    from django.db.models import Q, Sum
    
    # Joriy qoldiq
    balance, created = CashRegisterBalance.objects.get_or_create(id=1)
    
    # Bugungi sana
    today = timezone.now().date()
    today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))
    
    # Bugungi operatsiyalar
    today_transactions = CashTransaction.objects.filter(
        transaction_date__range=[today_start, today_end],
        status='COMPLETED'
    )
    
    # Bugungi statistika
    today_stats = {
        'total_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'total_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'cash_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CASH'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'cash_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CASH'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'click_income': today_transactions.filter(
            payment_method='CLICK', transaction_type__in=['INCOME', 'EXTERNAL_INCOME']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'payme_income': today_transactions.filter(
            payment_method='PAYME', transaction_type__in=['INCOME', 'EXTERNAL_INCOME']
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
    }
    today_stats['net_change'] = today_stats['total_income'] - today_stats['total_expense']
    
    # Filtrlar
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    transaction_type = request.GET.get('type', '')
    category = request.GET.get('category', '')
    payment_method = request.GET.get('payment_method', '')
    
    # Operatsiyalar
    transactions = CashTransaction.objects.filter(status='COMPLETED')
    
    if start_date:
        start = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        transactions = transactions.filter(transaction_date__gte=start)
    if end_date:
        end = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
        transactions = transactions.filter(transaction_date__lt=end)
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    if category:
        transactions = transactions.filter(category=category)
    if payment_method:
        transactions = transactions.filter(payment_method=payment_method)
    
    # Pagination
    paginator = Paginator(transactions.order_by('-transaction_date'), 30)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    # Bugungi hisobot
    today_report = DailyCashReport.objects.filter(report_date=today).first()
    
    # Faol buyurtmalar
    active_orders = Order.objects.filter(
        status__in=['KIRITILDI', 'TASDIQLANDI', 'ISHDA', 'TAYYOR']
    )[:20]
    
    # Kategoriyalar
    categories = CashTransaction.CATEGORY_CHOICES
    
    # URL parametrlarini saqlash
    current_params = ''
    if start_date:
        current_params += f'&start_date={start_date}'
    if end_date:
        current_params += f'&end_date={end_date}'
    if transaction_type:
        current_params += f'&type={transaction_type}'
    if category:
        current_params += f'&category={category}'
    if payment_method:
        current_params += f'&payment_method={payment_method}'
    
    context = {
        'balance': balance,
        'today_stats': today_stats,
        'today_transactions_count': today_transactions.count(),
        'today_income_count': today_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME']).count(),
        'today_expense_count': today_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']).count(),
        'today': today,
        'today_report': today_report,
        'page_obj': page_obj,
        'active_orders': active_orders,
        'categories': categories,
        'start_date': start_date,
        'end_date': end_date,
        'current_params': current_params,
        'opening_balance': today_report.expected_balance if today_report else balance.cash_balance,
    }
    
    return render(request, 'orders/cash_management.html', context)


def cash_transaction_json(request, transaction_id):
    transaction = get_object_or_404(CashTransaction, transaction_id=transaction_id)
    data = {
        'transaction_id': transaction.transaction_id,
        'transaction_date': transaction.transaction_date.strftime('%d.%m.%Y %H:%M:%S'),
        'amount': float(transaction.amount),
        'currency': transaction.currency,    # ← qo'shildi
        'transaction_type': transaction.transaction_type,
        'customer_name': transaction.customer_name or '-',
        'description': transaction.description,
        'performed_by': transaction.performed_by.get_full_name() or transaction.performed_by.username if transaction.performed_by else '-',
    }
    return JsonResponse(data)
from .models import (
    Order, Notification, Worker, Customer, 
    Material, MaterialTransaction, Category,
    CashTransaction, DailyCashReport, CashRegisterBalance  # ✅ Kassa modellari
)

@login_required
@user_passes_test(is_cashier, login_url='/login/')
def daily_report_json(request, pk):
    """AJAX uchun hisobot ma'lumotlarini JSON qaytarish"""
    report = get_object_or_404(DailyCashReport, pk=pk)
    
    data = {
        'report_date': report.report_date.strftime('%d.%m.%Y'),
        'opening_balance': float(report.opening_balance),
        'total_income': float(report.total_income),
        'total_expense': float(report.total_expense),
        'expected_balance': float(report.expected_balance),
        'actual_balance': float(report.actual_balance),
        'difference': float(report.difference),
        'notes': report.notes,
        'created_by': report.created_by.get_full_name() or report.created_by.username if report.created_by else '-',
    }
    
    return JsonResponse(data)

def cash_transaction_json(request, transaction_id):
    transaction = get_object_or_404(CashTransaction, transaction_id=transaction_id)
    data = {
        'transaction_id': transaction.transaction_id,
        'transaction_date': transaction.transaction_date.strftime('%d.%m.%Y %H:%M:%S'),
        'amount': float(transaction.amount),
        'currency': transaction.currency,    # ← qo'shildi
        'transaction_type': transaction.transaction_type,
        'customer_name': transaction.customer_name or '-',
        'description': transaction.description,
        'performed_by': transaction.performed_by.get_full_name() or transaction.performed_by.username if transaction.performed_by else '-',
    }
    return JsonResponse(data)
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def daily_report_json(request, pk):
    """AJAX uchun hisobot ma'lumotlarini JSON qaytarish"""
    report = get_object_or_404(DailyCashReport, pk=pk)
    
    data = {
        'report_date': report.report_date.strftime('%d.%m.%Y'),
        'opening_balance': float(report.opening_balance),
        'total_income': float(report.total_income),
        'total_expense': float(report.total_expense),
        'expected_balance': float(report.expected_balance),
        'actual_balance': float(report.actual_balance),
        'difference': float(report.difference),
        'notes': report.notes,
        'created_by': report.created_by.get_full_name() or report.created_by.username if report.created_by else '-',
    }
    
    return JsonResponse(data)
# ==================== KASSA AJAX API ENDPOINTLAR ====================
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_stats(request):
    from django.db.models import Sum
    from decimal import Decimal

    today = timezone.now().date()
    today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))

    today_transactions = CashTransaction.objects.filter(
        transaction_date__range=[today_start, today_end],
        status='COMPLETED'
    )

    balance, _ = CashRegisterBalance.objects.get_or_create(id=1)
    
    # Bugungi statistika - BARCHA KIRIM/CHIQIM TURLARI UCHUN
    today_usd_income = today_transactions.filter(
        transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], currency='USD'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    today_usd_expense = today_transactions.filter(
        transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], currency='USD'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    today_uzs_income = today_transactions.filter(
        transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], currency='UZS'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    today_uzs_expense = today_transactions.filter(
        transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], currency='UZS'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    # Null bo'lganlarni UZS ga qo'shamiz
    today_null_income = today_transactions.filter(
        transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], currency__isnull=True
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    today_null_expense = today_transactions.filter(
        transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], currency__isnull=True
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    today_uzs_income += today_null_income
    today_uzs_expense += today_null_expense

    # ✅ TO'LOV USULLARI BO'YICHA STATISTIKA - BARCHA TURLAR UCHUN
    payment_stats = {
        # Naqd pul
        'cash_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CASH'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'cash_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CASH'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        
        # Plastik karta
        'card_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CARD'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'card_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CARD'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        
        # Click
        'click_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='CLICK'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'click_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='CLICK'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        
        # Payme
        'payme_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='PAYME'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'payme_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='PAYME'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        
        # Bank
        'bank_income': today_transactions.filter(
            transaction_type__in=['INCOME', 'EXTERNAL_INCOME'], payment_method='BANK'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
        'bank_expense': today_transactions.filter(
            transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE'], payment_method='BANK'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0'),
    }

    # Qarz statistikasi
    debt_stats = {
        'total_usd': Debt.objects.filter(is_active=True, currency='USD').aggregate(Sum('remaining'))['remaining__sum'] or Decimal('0'),
        'total_uzs': Debt.objects.filter(is_active=True, currency='UZS').aggregate(Sum('remaining'))['remaining__sum'] or Decimal('0'),
        'count_usd': Debt.objects.filter(is_active=True, currency='USD', remaining__gt=0).count(),
        'count_uzs': Debt.objects.filter(is_active=True, currency='UZS', remaining__gt=0).count(),
    }

    # Muddati o'tgan qarzlar
    overdue_debts = Debt.objects.filter(is_active=True, remaining__gt=0, due_date__lt=today)
    overdue_list = [
        {
            'full_name': d.full_name,
            'remaining': float(d.remaining),
            'currency': d.currency,
            'due_date': d.due_date.strftime('%d.%m.%Y'),
            'debt_id': str(d.debt_id),
        }
        for d in overdue_debts[:10]
    ]

    # Bugungi hisobot
    today_report = DailyCashReport.objects.filter(report_date=today).first()
    expected_balance = float(today_report.expected_balance) if today_report else 0
    actual_balance = float(today_report.actual_balance) if today_report else 0

    data = {
        'success': True,
        'cash_balance_uzs': float(balance.cash_balance),
        'cash_balance_usd': float(balance.cash_balance_usd),
        'last_updated': balance.last_updated.strftime('%H:%M %d.%m.%Y'),
        
        # Bugungi ma'lumotlar
        'today_income_uzs': float(today_uzs_income),
        'today_income_usd': float(today_usd_income),
        'today_expense_uzs': float(today_uzs_expense),
        'today_expense_usd': float(today_usd_expense),
        
        'today_income': float(today_uzs_income + today_usd_income),
        'today_expense': float(today_uzs_expense + today_usd_expense),
        'net_change': float((today_uzs_income + today_usd_income) - (today_uzs_expense + today_usd_expense)),
        
        'today_count': today_transactions.count(),
        'today_income_count': today_transactions.filter(transaction_type__in=['INCOME', 'EXTERNAL_INCOME']).count(),
        'today_expense_count': today_transactions.filter(transaction_type__in=['EXPENSE', 'EXTERNAL_EXPENSE']).count(),
        'today': today.strftime('%d.%m.%Y'),
        'today_report': today_report is not None,
        'expected_balance': expected_balance,
        'actual_balance': actual_balance,
        
        # ✅ TO'LOV USULLARI BO'YICHA STATISTIKA
        'payment_stats': payment_stats,
        
        # Qarz statistikasi
        'debt_stats': debt_stats,
        'overdue_debts': overdue_list,
    }

    return JsonResponse(data)
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_transactions(request):
    """AJAX uchun operatsiyalar ro'yxati - payment_method bilan"""
    from django.core.paginator import Paginator
    
    page = request.GET.get('page', 1)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    trans_type = request.GET.get('type')
    currency = request.GET.get('currency')
    payment_method = request.GET.get('payment_method', '')  # ✅ QO'SHILDI
    
    transactions = CashTransaction.objects.filter(status='COMPLETED')
    
    if start_date:
        transactions = transactions.filter(transaction_date__date__gte=start_date)
    if end_date:
        transactions = transactions.filter(transaction_date__date__lte=end_date)
    if trans_type:
        transactions = transactions.filter(transaction_type=trans_type)
    if currency:
        transactions = transactions.filter(currency=currency)
    if payment_method:
        transactions = transactions.filter(payment_method=payment_method)  # ✅ QO'SHILDI
    
    paginator = Paginator(transactions.order_by('-transaction_date'), 50)
    page_obj = paginator.get_page(page)
    
    data = {
        'success': True,
        'transactions': [
            {
                'transaction_id': t.transaction_id,
                'date': t.transaction_date.strftime('%d.%m.%Y %H:%M'),
                'type': t.transaction_type,
                'currency': getattr(t, 'currency', 'UZS'),
                'payment_method': t.payment_method,  # ✅ QO'SHILDI
                'amount': float(t.amount),
                'customer_name': t.customer_name,
                'description': t.description,
                'performed_by': t.performed_by.get_full_name() if t.performed_by else '-',
            }
            for t in page_obj
        ],
        'total': paginator.count,
        'total_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
    }
    
    return JsonResponse(data)
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_transaction_create(request):
    """AJAX orqali yangi kassa operatsiyasi yaratish"""
    from decimal import Decimal, InvalidOperation
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': "Faqat POST so'rovi qabul qilinadi"}, status=405)

    transaction_type = request.POST.get('transaction_type')
    amount_str = request.POST.get('amount', '0')
    customer_name = request.POST.get('customer_name', '').strip()
    description = request.POST.get('description', '').strip()
    currency = request.POST.get('currency', 'UZS')
    payment_method = request.POST.get('payment_method', 'CASH')
    category = request.POST.get('category', 'OTHER')

    # Validatsiya
    if not transaction_type:
        return JsonResponse({'success': False, 'message': "Operatsiya turini tanlang!"})
    
    # ✅ Decimal ga o'tkazishda xatolikni ushlash
    try:
        amount = Decimal(amount_str)
    except InvalidOperation:
        return JsonResponse({'success': False, 'message': "Summa noto'g'ri formatda! Iltimos, son kiriting."})
    
    if amount <= 0:
        return JsonResponse({'success': False, 'message': "Summa 0 dan katta bo'lishi kerak!"})
    
    if not customer_name:
        return JsonResponse({'success': False, 'message': "Kim oldi/berdi maydonini to'ldiring!"})
    
    if not description:
        return JsonResponse({'success': False, 'message': "Izoh maydonini to'ldiring!"})

    try:
        # Tranzaksiyani yaratish
        transaction = CashTransaction.objects.create(
            transaction_type=transaction_type,
            category=category,
            amount=amount,
            currency=currency,
            payment_method=payment_method,
            customer_name=customer_name,
            description=description,
            status='COMPLETED',
            performed_by=request.user,
            transaction_date=timezone.now()
        )

        # FAQAT NAQD PUL kassa balansiga ta'sir qiladi
        balance, _ = CashRegisterBalance.objects.get_or_create(id=1)
        
        if payment_method == 'CASH':
            if transaction_type == 'INCOME':
                if currency == 'USD':
                    balance.cash_balance_usd += amount
                else:
                    balance.cash_balance += amount
            else:  # EXPENSE
                if currency == 'USD':
                    balance.cash_balance_usd -= amount
                else:
                    balance.cash_balance -= amount
            balance.updated_by = request.user
            balance.save()
        
        # Qaysi to'lov usuli ekanligini log'ga yozish
        payment_names = {
            'CASH': 'Naqd pul',
            'CARD': 'Plastik karta',
            'CLICK': 'Click',
            'PAYME': 'Payme',
            'BANK': 'Bank pul o\'tkazmasi'
        }
        payment_name = payment_names.get(payment_method, payment_method)
        
        return JsonResponse({
            'success': True,
            'message': f"✅ {payment_name} orqali {amount:,.2f} {currency} operatsiya bajarildi!",
            'transaction': {
                'id': transaction.transaction_id,
                'amount': float(amount),
                'currency': currency,
                'type': transaction_type,
                'payment_method': payment_method
            }
        })
    except Exception as e:
        print(f"❌ Xatolik: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Xatolik: {str(e)}'})

@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_daily_report_create(request):
    """AJAX orqali kunlik hisobot yaratish"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rovi qabul qilinadi'}, status=405)
    
    today = timezone.now().date()
    
    # Bugungi hisobot mavjudligini tekshirish
    if DailyCashReport.objects.filter(report_date=today).exists():
        return JsonResponse({'success': False, 'message': f'{today} uchun hisobot allaqachon mavjud!'})
    
    opening_balance = Decimal(request.POST.get('opening_balance', '0'))
    actual_balance = Decimal(request.POST.get('actual_balance', '0'))
    notes = request.POST.get('notes', '')
    
    # Bugungi operatsiyalar
    today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))
    
    today_transactions = CashTransaction.objects.filter(
        transaction_date__range=[today_start, today_end],
        status='COMPLETED'
    )
    
    cash_income = today_transactions.filter(transaction_type='INCOME').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    cash_expense = today_transactions.filter(transaction_type='EXPENSE').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    total_income = cash_income
    total_expense = cash_expense
    expected_balance = opening_balance + total_income - total_expense
    difference = actual_balance - expected_balance
    
    try:
        report = DailyCashReport.objects.create(
            report_date=today,
            opening_balance=opening_balance,
            cash_income=cash_income,
            cash_expense=cash_expense,
            click_income=Decimal('0'),
            payme_income=Decimal('0'),
            bank_income=Decimal('0'),
            click_expense=Decimal('0'),
            payme_expense=Decimal('0'),
            bank_expense=Decimal('0'),
            total_income=total_income,
            total_expense=total_expense,
            expected_balance=expected_balance,
            actual_balance=actual_balance,
            difference=difference,
            notes=notes,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{today} uchun kunlik hisobot yaratildi!',
            'report_id': report.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Xatolik: {str(e)}'})
    

@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_transaction_delete(request, transaction_id):
    """Kassa operatsiyasini o'chirish"""
    from decimal import Decimal
    
    if request.method == 'POST':
        try:
            transaction = CashTransaction.objects.get(transaction_id=transaction_id, status='COMPLETED')
            
            # Kassa qoldig'ini qaytarish (operatsiyani bekor qilish)
            balance, _ = CashRegisterBalance.objects.get_or_create(id=1)
            
            if transaction.transaction_type == 'INCOME':
                if transaction.currency == 'USD':
                    balance.cash_balance_usd -= transaction.amount
                else:
                    balance.cash_balance -= transaction.amount
            else:  # EXPENSE
                if transaction.currency == 'USD':
                    balance.cash_balance_usd += transaction.amount
                else:
                    balance.cash_balance += transaction.amount
            
            balance.updated_by = request.user
            balance.save()
            
            # Operatsiyani o'chirish
            transaction.status = 'CANCELLED'
            transaction.save()
            
            messages.success(request, f"✅ Operatsiya o'chirildi! ID: {transaction_id}")
            return JsonResponse({'success': True, 'message': 'Operatsiya o\'chirildi'})
            
        except CashTransaction.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Operatsiya topilmadi'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_transaction_edit(request, transaction_id):
    """Kassa operatsiyasini tahrirlash - AJAX uchun"""
    from decimal import Decimal
    from datetime import datetime
    
    try:
        transaction = CashTransaction.objects.get(transaction_id=transaction_id, status='COMPLETED')
    except CashTransaction.DoesNotExist:
        if request.method == 'GET':
            return JsonResponse({'success': False, 'message': 'Operatsiya topilmadi'}, status=404)
        messages.error(request, "Operatsiya topilmadi")
        return redirect('cash_management')
    
    if request.method == 'GET':
        data = {
            'success': True,
            'transaction_id': transaction.transaction_id,
            'transaction_type': transaction.transaction_type,
            'amount': float(transaction.amount),
            'currency': getattr(transaction, 'currency', 'UZS'),
            'payment_method': getattr(transaction, 'payment_method', 'CASH'),
            'customer_name': transaction.customer_name,
            'description': transaction.description,
            'date': transaction.transaction_date.strftime('%Y-%m-%dT%H:%M'),
        }
        return JsonResponse(data)
    
    if request.method == 'POST':
        try:
            # POST dan ma'lumotlarni olish
            new_type = request.POST.get('transaction_type')
            new_amount = Decimal(request.POST.get('amount', '0'))
            new_currency = request.POST.get('currency', 'UZS')
            new_payment_method = request.POST.get('payment_method', 'CASH')
            new_customer = request.POST.get('customer_name', '').strip()
            new_description = request.POST.get('description', '').strip()
            new_date_str = request.POST.get('transaction_date', '')
            
            # Validatsiya
            if not new_type:
                return JsonResponse({'success': False, 'message': 'Operatsiya turini tanlang!'})
            if new_amount <= 0:
                return JsonResponse({'success': False, 'message': 'Summa 0 dan katta bo\'lishi kerak!'})
            if not new_customer:
                return JsonResponse({'success': False, 'message': 'Kim oldi/berdi maydonini to\'ldiring!'})
            
            # Kassa qoldig'ini yangilash (faqat eski va yangi CASH bo'lsa)
            balance, _ = CashRegisterBalance.objects.get_or_create(id=1)
            
            # Eski operatsiyani bekor qilish (faqat eski CASH bo'lsa)
            if transaction.payment_method == 'CASH':
                if transaction.transaction_type == 'INCOME':
                    if transaction.currency == 'USD':
                        balance.cash_balance_usd -= transaction.amount
                    else:
                        balance.cash_balance -= transaction.amount
                else:
                    if transaction.currency == 'USD':
                        balance.cash_balance_usd += transaction.amount
                    else:
                        balance.cash_balance += transaction.amount
            
            # Yangi operatsiyani qo'shish (faqat yangi CASH bo'lsa)
            if new_payment_method == 'CASH':
                if new_type == 'INCOME':
                    if new_currency == 'USD':
                        balance.cash_balance_usd += new_amount
                    else:
                        balance.cash_balance += new_amount
                else:
                    if new_currency == 'USD':
                        balance.cash_balance_usd -= new_amount
                    else:
                        balance.cash_balance -= new_amount
            
            balance.updated_by = request.user
            balance.save()
            
            # Tranzaksiyani yangilash
            transaction.transaction_type = new_type
            transaction.amount = new_amount
            transaction.currency = new_currency
            transaction.payment_method = new_payment_method
            transaction.customer_name = new_customer
            transaction.description = new_description
            
            if new_date_str:
                new_date = timezone.make_aware(datetime.strptime(new_date_str, '%Y-%m-%dT%H:%M'))
                transaction.transaction_date = new_date
            
            transaction.save()
            
            return JsonResponse({'success': True, 'message': 'Operatsiya yangilandi'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_transaction_delete(request, transaction_id):
    """Kassa operatsiyasini o'chirish - AJAX uchun"""
    from decimal import Decimal
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    try:
        transaction = CashTransaction.objects.get(transaction_id=transaction_id, status='COMPLETED')
        
        # Kassa qoldig'ini qaytarish (faqat CASH bo'lsa)
        if transaction.payment_method == 'CASH':
            balance, _ = CashRegisterBalance.objects.get_or_create(id=1)
            
            if transaction.transaction_type == 'INCOME':
                if transaction.currency == 'USD':
                    balance.cash_balance_usd -= transaction.amount
                else:
                    balance.cash_balance -= transaction.amount
            else:
                if transaction.currency == 'USD':
                    balance.cash_balance_usd += transaction.amount
                else:
                    balance.cash_balance += transaction.amount
            
            balance.updated_by = request.user
            balance.save()
        
        # Operatsiyani o'chirish
        transaction.status = 'CANCELLED'
        transaction.save()
        
        return JsonResponse({'success': True, 'message': 'Operatsiya o\'chirildi'})
        
    except CashTransaction.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Operatsiya topilmadi'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ==================== QARZDORLAR (DEBT) FUNKSIYALARI ====================
from .models import (
    Order, Notification, Worker, Customer, 
    Material, MaterialTransaction, Category,
    CashTransaction, DailyCashReport, CashRegisterBalance,
    Debt, DebtTransaction,  # ✅ QARZDORLAR MODELLARI - SHUNI QO'SHING
)
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_debts(request):
    """AJAX uchun qarzdorlar ro'yxati"""
    from django.core.paginator import Paginator
    from django.db.models import Q, Sum
    
    page = request.GET.get('page', 1)
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    currency = request.GET.get('currency', '')
    
    # Qarzdorlarni olish
    debts = Debt.objects.filter(is_active=True)
    
    # Qidiruv
    if search:
        debts = debts.filter(
            Q(full_name__icontains=search) |
            Q(phone__icontains=search)
        )
    
    # Valyuta filtri
    if currency:
        debts = debts.filter(currency=currency)
    
    # Holat filtri
    today = timezone.now().date()
    if status == 'active':
        debts = debts.filter(remaining__gt=0, due_date__gte=today)
    elif status == 'overdue':
        debts = debts.filter(remaining__gt=0, due_date__lt=today)
    elif status == 'paid':
        debts = debts.filter(remaining=0)
    
    # Pagination
    paginator = Paginator(debts.order_by('-created_at'), 50)
    page_obj = paginator.get_page(page)
    
    data = {
        'success': True,
        'debts': [
            {
                'debt_id': str(d.debt_id),
                'full_name': d.full_name,
                'phone': d.phone or '',
                'amount': float(d.amount),
                'remaining': float(d.remaining),
                'currency': d.currency,
                'due_date': d.due_date.strftime('%Y-%m-%d') if d.due_date else None,
                'description': d.description or '',
                'created_at': d.created_at.strftime('%d.%m.%Y'),
                'is_overdue': d.due_date and d.due_date < today and d.remaining > 0,
            }
            for d in page_obj
        ],
        'total': paginator.count,
        'total_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
    }
    
    return JsonResponse(data)

@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_debt_create(request):
    """AJAX orqali yangi qarzdor qo'shish - Kassa balansiga ta'sir qilmaydi"""
    from decimal import Decimal
    import uuid
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rovi qabul qilinadi'}, status=405)
    
    full_name = request.POST.get('full_name', '').strip()
    phone = request.POST.get('phone', '').strip()
    amount = request.POST.get('amount', '0')
    currency = request.POST.get('currency', 'UZS')
    due_date = request.POST.get('due_date')
    description = request.POST.get('description', '').strip()
    
    # Validatsiya
    if not full_name:
        return JsonResponse({'success': False, 'message': "To'liq ism majburiy!"})
    
    try:
        amount = Decimal(amount)
        if amount <= 0:
            return JsonResponse({'success': False, 'message': "Qarz summasi 0 dan katta bo'lishi kerak!"})
    except:
        return JsonResponse({'success': False, 'message': "Summa noto'g'ri formatda!"})
    
    if not due_date:
        return JsonResponse({'success': False, 'message': "Qaytarish muddati majburiy!"})
    
    try:
        due_date = timezone.datetime.strptime(due_date, '%Y-%m-%d').date()
    except:
        return JsonResponse({'success': False, 'message': "Sana noto'g'ri formatda!"})
    
    # Qarzdorni yaratish
    debt = Debt.objects.create(
        debt_id=uuid.uuid4().hex[:12].upper(),
        full_name=full_name,
        phone=phone or None,
        amount=amount,
        remaining=amount,
        currency=currency,
        due_date=due_date,
        description=description,
        created_by=request.user
    )
    
    # ❌ KASSA BALANSIGA TA'SIR QILMAYDI - OLIB TASHLANDI
    # balance, _ = CashRegisterBalance.objects.get_or_create(id=1)
    # if currency == 'USD':
    #     balance.cash_balance_usd -= amount
    # else:
    #     balance.cash_balance -= amount
    # balance.updated_by = request.user
    # balance.save()
    
    # 🔥 Qarz operatsiyasini yozish (faqat tarix uchun)
    DebtTransaction.objects.create(
        debt=debt,
        transaction_type='DEBT_GIVEN',
        amount=amount,
        currency=currency,
        remaining_after=amount,
        description=f"{full_name} ga qarz berildi - {description}" if description else f"{full_name} ga qarz berildi",
        created_by=request.user
    )
    
    return JsonResponse({
        'success': True,
        'message': f"✅ {full_name} ga {amount:,.2f} {currency} qarz berildi! (Kassa balansiga ta'sir qilmaydi)",
        'debt': {
            'id': str(debt.debt_id),
            'full_name': debt.full_name,
            'amount': float(debt.amount),
            'remaining': float(debt.remaining),
            'currency': debt.currency,
            'due_date': debt.due_date.strftime('%Y-%m-%d'),
        }
    })

@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_debt_json(request, debt_id):
    """Qarzdor ma'lumotlarini JSON qaytarish"""
    try:
        debt = Debt.objects.get(debt_id=debt_id, is_active=True)
        
        today = timezone.now().date()
        is_overdue = debt.due_date and debt.due_date < today and debt.remaining > 0
        
        data = {
            'success': True,
            'debt_id': str(debt.debt_id),
            'full_name': debt.full_name,
            'phone': debt.phone or '',
            'amount': float(debt.amount),
            'remaining': float(debt.remaining),
            'currency': debt.currency,
            'due_date': debt.due_date.strftime('%Y-%m-%d') if debt.due_date else None,
            'description': debt.description or '',
            'created_at': debt.created_at.strftime('%d.%m.%Y %H:%M'),
            'is_overdue': is_overdue,
            'paid_amount': float(debt.amount - debt.remaining),
        }
        return JsonResponse(data)
        
    except Debt.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Qarzdor topilmadi'}, status=404)


@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_debt_edit(request, debt_id):
    """Qarzdor ma'lumotlarini tahrirlash - Kassa balansiga ta'sir qilmaydi"""
    from decimal import Decimal
    
    try:
        debt = Debt.objects.get(debt_id=debt_id, is_active=True)
    except Debt.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Qarzdor topilmadi'}, status=404)
    
    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        amount = request.POST.get('amount', '0')
        currency = request.POST.get('currency', 'UZS')
        due_date = request.POST.get('due_date')
        description = request.POST.get('description', '').strip()
        
        if not full_name:
            return JsonResponse({'success': False, 'message': "To'liq ism majburiy!"})
        
        try:
            new_amount = Decimal(amount)
            if new_amount <= 0:
                return JsonResponse({'success': False, 'message': "Qarz summasi 0 dan katta bo'lishi kerak!"})
        except:
            return JsonResponse({'success': False, 'message': "Summa noto'g'ri formatda!"})
        
        if not due_date:
            return JsonResponse({'success': False, 'message': "Qaytarish muddati majburiy!"})
        
        try:
            new_due_date = timezone.datetime.strptime(due_date, '%Y-%m-%d').date()
        except:
            return JsonResponse({'success': False, 'message': "Sana noto'g'ri formatda!"})
        
        # ❌ KASSA BALANSIGA TA'SIR QILMAYDI - OLIB TASHLANDI
        # old_amount = debt.amount
        # balance, _ = CashRegisterBalance.objects.get_or_create(id=1)
        # if old_amount != new_amount:
        #     difference = new_amount - old_amount
        #     if debt.currency == 'USD':
        #         balance.cash_balance_usd -= difference
        #     else:
        #         balance.cash_balance -= difference
        #     balance.updated_by = request.user
        #     balance.save()
        
        # Qarzdorni yangilash
        debt.full_name = full_name
        debt.phone = phone or None
        debt.amount = new_amount
        debt.currency = currency
        debt.due_date = new_due_date
        debt.description = description
        debt.save()
        
        return JsonResponse({
            'success': True,
            'message': f"✅ {full_name} ma'lumotlari yangilandi! (Kassa balansiga ta'sir qilmaydi)"
        })
    
    data = {
        'success': True,
        'full_name': debt.full_name,
        'phone': debt.phone or '',
        'amount': float(debt.amount),
        'currency': debt.currency,
        'due_date': debt.due_date.strftime('%Y-%m-%d') if debt.due_date else None,
        'description': debt.description or '',
    }
    return JsonResponse(data)
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_debt_delete(request, debt_id):
    """Qarzdorni o'chirish - Qolgan qarz bo'lsa ham o'chirish mumkin"""
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    try:
        debt = Debt.objects.get(debt_id=debt_id, is_active=True)
        
        # ❌ QOLGAN QARZNI TEKSHIRMAYMIZ - O'CHIRISH MUMKIN
        # if debt.remaining > 0:
        #     return JsonResponse({
        #         'success': False,
        #         'message': f"Bu qarzdorning hali {debt.remaining:,.2f} {debt.currency} qarzi bor! Avval to'lashing kerak."
        #     })
        
        # ✅ Qarzdorni o'chirish (faqat statusini o'zgartiramiz)
        debt.is_active = False
        debt.save()
        
        # Qarzdor bilan bog'liq qarz operatsiyalarini ham o'chirish (yoki faqat qarzdorni)
        # DebtTransaction.objects.filter(debt=debt).delete()
        
        return JsonResponse({'success': True, 'message': 'Qarzdor o\'chirildi'})
        
    except Debt.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Qarzdor topilmadi'}, status=404)
@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_debt_payment(request):
    """Qarz to'lovini qabul qilish - AJAX"""
    from decimal import Decimal
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rovi qabul qilinadi'}, status=405)
    
    debt_id = request.POST.get('debt_id')
    amount = request.POST.get('amount', '0')
    payment_method = request.POST.get('payment_method', 'CASH')
    description = request.POST.get('description', '').strip()
    
    if not debt_id:
        return JsonResponse({'success': False, 'message': 'Qarzdor ID si majburiy!'})
    
    try:
        debt = Debt.objects.get(debt_id=debt_id, is_active=True)
    except Debt.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Qarzdor topilmadi'}, status=404)
    
    try:
        amount = Decimal(amount)
        if amount <= 0:
            return JsonResponse({'success': False, 'message': "To'lov summasi 0 dan katta bo'lishi kerak!"})
    except:
        return JsonResponse({'success': False, 'message': "Summa noto'g'ri formatda!"})
    
    if amount > debt.remaining:
        return JsonResponse({
            'success': False,
            'message': f"To'lov summasi qolgan qarzdan ({debt.remaining:,.2f}) ko'p bo'lishi mumkin emas!"
        })
    
    # 🔥 Kassa qoldig'iga kirim (qarz to'landi)

    
    # Qolgan qarzni hisoblash
    old_remaining = debt.remaining
    debt.remaining -= amount
    debt.save()
    
    # 🔥 Qarz operatsiyasini yozish
    payment_names = {
        'CASH': 'Naqd pul',
        'CARD': 'Plastik karta',
        'CLICK': 'Click',
        'PAYME': 'Payme',
        'BANK': 'Bank'
    }
    payment_name = payment_names.get(payment_method, 'Naqd pul')
    
    transaction = DebtTransaction.objects.create(
        debt=debt,
        transaction_type='DEBT_PAID',
        amount=amount,
        currency=debt.currency,
        remaining_after=debt.remaining,
        description=f"{debt.full_name} dan qarz to'lovi - {payment_name} orqali. {description}" if description else f"{debt.full_name} dan qarz to'lovi - {payment_name} orqali",
        created_by=request.user
    )
    
    return JsonResponse({
        'success': True,
        'message': f"✅ {debt.full_name} dan {amount:,.2f} {debt.currency} qarz to'lovi qabul qilindi! Qolgan qarz: {debt.remaining:,.2f}",
        'remaining': float(debt.remaining),
        'is_paid': debt.remaining == 0
    })


@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_debt_transactions(request):
    """Qarz operatsiyalari ro'yxati - AJAX"""
    from django.core.paginator import Paginator
    
    page = request.GET.get('page', 1)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    debtor_id = request.GET.get('debtor')
    
    transactions = DebtTransaction.objects.all().order_by('-created_at')
    
    if start_date:
        transactions = transactions.filter(created_at__date__gte=start_date)
    if end_date:
        transactions = transactions.filter(created_at__date__lte=end_date)
    if debtor_id:
        transactions = transactions.filter(debt__debt_id=debtor_id)
    
    paginator = Paginator(transactions, 50)
    page_obj = paginator.get_page(page)
    
    data = {
        'success': True,
        'transactions': [
            {
                'transaction_id': str(t.transaction_id),
                'date': t.created_at.strftime('%d.%m.%Y %H:%M'),
                'debtor_name': t.debt.full_name,
                'type': t.transaction_type,
                'currency': t.currency,
                'amount': float(t.amount),
                'remaining_after': float(t.remaining_after),
                'description': t.description or '',
            }
            for t in page_obj
        ],
        'total': paginator.count,
        'total_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
    }
    
    return JsonResponse(data)


@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_debt_transaction_json(request, transaction_id):
    """Qarz operatsiyasi tafsilotlarini JSON qaytarish"""
    from django.shortcuts import get_object_or_404
    
    transaction = get_object_or_404(DebtTransaction, transaction_id=transaction_id)
    
    data = {
        'success': True,
        'transaction_id': str(transaction.transaction_id),
        'date': transaction.created_at.strftime('%d.%m.%Y %H:%M'),
        'debtor_name': transaction.debt.full_name,
        'type': transaction.transaction_type,
        'currency': transaction.currency,
        'amount': float(transaction.amount),
        'remaining_after': float(transaction.remaining_after),
        'description': transaction.description or '',
        'debt_id': str(transaction.debt.debt_id),
    }
    
    return JsonResponse(data)


@login_required
@user_passes_test(is_cashier, login_url='/login/')
def cash_api_debtors_list(request):
    """Select uchun qarzdorlar ro'yxati (faol qarzi borlar)"""
    debts = Debt.objects.filter(is_active=True, remaining__gt=0).order_by('full_name')
    
    data = {
        'success': True,
        'debtors': [
            {
                'debt_id': str(d.debt_id),
                'full_name': d.full_name,
                'remaining': float(d.remaining),
                'currency': d.currency,
            }
            for d in debts
        ]
    }
    
    return JsonResponse(data)


# =======================================================================
# OSHXONA (KITCHEN) VIEW'LARI
# =======================================================================

from decimal import Decimal
from django.db.models import Sum
from .models import (
    KitchenIngredient, KitchenIngredientTransaction,
    DailyMeal, DailyMealIngredient, KitchenOrder
)
from .forms import (
    KitchenIngredientForm, DailyMealForm,
    DailyMealIngredientFormSet, KitchenOrderForm
)

def is_kitchen_staff(user):
    return user.groups.filter(name='Kitchen').exists() or user.is_superuser

@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def kitchen_dashboard(request):
    """Oshxona bosh sahifasi"""
    today = timezone.now().date()
    today_meals = DailyMeal.objects.filter(date=today).order_by('meal_type')
    ingredients = KitchenIngredient.objects.all().order_by('name')
    low_ingredients = [ing for ing in ingredients if ing.is_low()]
    today_total_persons = today_meals.aggregate(Sum('person_count'))['person_count__sum'] or 0
    
    context = {
        'today_meals': today_meals,
        'ingredients': ingredients,
        'low_ingredients': low_ingredients,
        'today_total_persons': today_total_persons,
        'low_count': len(low_ingredients),
    }
    return render(request, 'orders/kitchen_dashboard.html', context)

@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def kitchen_ingredient_list(request):
    """Masalliqlar ro'yxati"""
    ingredients = KitchenIngredient.objects.all().order_by('name')
    search = request.GET.get('search', '')
    if search:
        ingredients = ingredients.filter(name__icontains=search)
    
    filter_type = request.GET.get('filter', 'all')
    if filter_type == 'low':
        ingredients = [ing for ing in ingredients if ing.is_low()]
    
    context = {
        'ingredients': ingredients,
        'search': search,
        'filter_type': filter_type,
    }
    return render(request, 'orders/kitchen_ingredient_list.html', context)

@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def kitchen_ingredient_add(request):
    """Yangi masalliq qo'shish"""
    if request.method == 'POST':
        form = KitchenIngredientForm(request.POST)
        if form.is_valid():
            ingredient = form.save()
            KitchenIngredientTransaction.objects.create(
                ingredient=ingredient,
                transaction_type='IN',
                amount=ingredient.quantity,
                previous_quantity=0,
                new_quantity=ingredient.quantity,
                description="Yangi masalliq qo'shildi",
                created_by=request.user
            )
            messages.success(request, f"{ingredient.name} muvaffaqiyatli qo'shildi!")
            return redirect('kitchen_ingredient_list')
    else:
        form = KitchenIngredientForm()
    
    return render(request, 'orders/kitchen_ingredient_form.html', {'form': form, 'title': 'Yangi masalliq'})

@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def kitchen_ingredient_edit(request, pk):
    """Masalliq tahrirlash"""
    ingredient = get_object_or_404(KitchenIngredient, pk=pk)
    
    if request.method == 'POST':
        form = KitchenIngredientForm(request.POST, instance=ingredient)
        if form.is_valid():
            old_quantity = ingredient.quantity
            ingredient = form.save()
            
            if old_quantity != ingredient.quantity:
                KitchenIngredientTransaction.objects.create(
                    ingredient=ingredient,
                    transaction_type='IN' if ingredient.quantity > old_quantity else 'OUT',
                    amount=abs(ingredient.quantity - old_quantity),
                    previous_quantity=old_quantity,
                    new_quantity=ingredient.quantity,
                    description="Miqdor tahrirlandi",
                    created_by=request.user
                )
            
            messages.success(request, f"{ingredient.name} muvaffaqiyatli tahrirlandi!")
            return redirect('kitchen_ingredient_list')
    else:
        form = KitchenIngredientForm(instance=ingredient)
    
    return render(request, 'orders/kitchen_ingredient_form.html', {'form': form, 'title': 'Masalliq tahrirlash', 'ingredient': ingredient})

@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def kitchen_ingredient_add_quantity(request, pk):
    """Masalliq miqdorini oshirish"""
    ingredient = get_object_or_404(KitchenIngredient, pk=pk)
    
    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', 0))
        description = request.POST.get('description', '')
        
        if amount <= 0:
            messages.error(request, "Miqdor 0 dan katta bo'lishi kerak!")
            return redirect('kitchen_ingredient_list')
        
        old_quantity = ingredient.quantity
        ingredient.add_quantity(amount)
        
        KitchenIngredientTransaction.objects.create(
            ingredient=ingredient,
            transaction_type='IN',
            amount=amount,
            previous_quantity=old_quantity,
            new_quantity=ingredient.quantity,
            description=description or "Masalliq qo'shildi",
            created_by=request.user
        )
        
        messages.success(request, f"{ingredient.name} ga {amount} {ingredient.get_unit_display()} qo'shildi!")
        return redirect('kitchen_ingredient_list')
    
    return render(request, 'orders/kitchen_ingredient_add_quantity.html', {'ingredient': ingredient})

@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def daily_meal_create(request):
    """Kunlik ovqat qo'shish"""
    from django.utils import timezone
    
    if request.method == 'POST':
        form = DailyMealForm(request.POST)
        formset = DailyMealIngredientFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            meal = form.save(commit=False)
            meal.created_by = request.user
            meal.save()
            
            formset.instance = meal
            formset.save()
            
            # Har bir masalliqni hisobdan chiqarish
            for meal_ingredient in meal.ingredients.all():
                ingredient = meal_ingredient.ingredient
                total_quantity = meal_ingredient.total_quantity
                old_quantity = ingredient.quantity
                
                if ingredient.subtract_quantity(total_quantity):
                    KitchenIngredientTransaction.objects.create(
                        ingredient=ingredient,
                        transaction_type='OUT',
                        amount=total_quantity,
                        previous_quantity=old_quantity,
                        new_quantity=ingredient.quantity,
                        description=f"Ovqat uchun: {meal.meal_name} ({meal.get_meal_type_display()})",
                        created_by=request.user
                    )
                else:
                    messages.warning(request, f"{ingredient.name} da yetarli miqdor yo'q! Joriy: {ingredient.quantity}")
            
            messages.success(request, f"{meal.meal_name} muvaffaqiyatli qo'shildi!")
            return redirect('kitchen_dashboard')
        else:
            messages.error(request, "Iltimos, formani to'g'ri to'ldiring!")
    else:
        form = DailyMealForm(initial={'date': timezone.now().date()})
        formset = DailyMealIngredientFormSet()
    
    context = {
        'form': form,
        'formset': formset,
        'title': 'Yangi ovqat qo\'shish',
        'today': timezone.now().date(),  # ✅ today qo'shildi
    }
    return render(request, 'orders/kitchen_meal_form.html', context)



@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def daily_meal_list(request):
    """Ovqatlar ro'yxati"""
    meals = DailyMeal.objects.all().order_by('-date', '-created_at')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        meals = meals.filter(date__gte=date_from)
    if date_to:
        meals = meals.filter(date__lte=date_to)
    
    context = {
        'meals': meals,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'orders/kitchen_meal_list.html', context)

@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def kitchen_order_create(request):
    """Oshxonaga buyurtma berish"""
    if request.method == 'POST':
        form = KitchenOrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.created_by = request.user
            order.save()
            messages.success(request, f"{order.order_number} buyurtma yaratildi!")
            return redirect('kitchen_order_list')
    else:
        form = KitchenOrderForm()
    
    return render(request, 'orders/kitchen_order_form.html', {'form': form, 'title': 'Yangi buyurtma'})

@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def kitchen_order_list(request):
    """Buyurtmalar ro'yxati"""
    orders = KitchenOrder.objects.all().order_by('-created_at')
    context = {'orders': orders}
    return render(request, 'orders/kitchen_order_list.html', context)

@login_required
@user_passes_test(is_kitchen_staff, login_url='/login/')
def kitchen_order_approve(request, pk):
    """Buyurtmani tasdiqlash va qabul qilish"""
    order = get_object_or_404(KitchenOrder, pk=pk)
    
    if request.method == 'POST':
        received = Decimal(request.POST.get('received_quantity', 0))
        
        if received <= 0:
            messages.error(request, "Qabul qilingan miqdor 0 dan katta bo'lishi kerak!")
            return redirect('kitchen_order_list')
        
        order.received_quantity = received
        order.status = 'COMPLETED'
        order.approved_by = request.user
        order.save()
        
        ingredient = order.ingredient
        old_quantity = ingredient.quantity
        ingredient.add_quantity(received)
        
        KitchenIngredientTransaction.objects.create(
            ingredient=ingredient,
            transaction_type='IN',
            amount=received,
            previous_quantity=old_quantity,
            new_quantity=ingredient.quantity,
            description=f"Buyurtma bo'yicha: {order.order_number}",
            created_by=request.user
        )
        
        messages.success(request, f"{order.order_number} buyurtma qabul qilindi!")
        return redirect('kitchen_order_list')
    
    return render(request, 'orders/kitchen_order_approve.html', {'order': order})







